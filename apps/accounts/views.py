from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group
from django.contrib import messages
from django.http import JsonResponse, HttpResponseRedirect, HttpResponseForbidden
from django.core.paginator import Paginator
from django.db.models import Q
from django.urls import reverse
from .models import User, Department, Role, Permission, LoginLog
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from .decorators import permission_required
import json
import threading
import uuid
import time

# 全局变量存储导入进度
import_progress = {}


def user_login(request):
    from django.contrib.auth import get_backends
    
    if request.method == 'POST':
        emp_no = request.POST.get('emp_no', '').strip()
        password = request.POST.get('password', '')
        remember = request.POST.get('remember', False)
        
        user = authenticate(request, username=emp_no, password=password)
        
        if user is not None:
            if user.is_active:
                backend = get_backends()[0]
                login(request, user, backend='django.contrib.auth.backends.ModelBackend')
                
                if remember:
                    request.session.set_expiry(60 * 60 * 24 * 14)  # 记住我：14天
                else:
                    from apps.settings.views import get_config_value
                    timeout_minutes = get_config_value('session_timeout_minutes', 120)
                    request.session.set_expiry(60 * timeout_minutes)  # 使用系统配置的会话超时
                
                LoginLog.objects.create(
                    user=user,
                    username=emp_no,
                    ip_address=get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', ''),
                    action='login',
                    message='登录成功'
                )
                next_url = request.GET.get('next', '/')
                return HttpResponseRedirect(next_url)
            else:
                messages.error(request, '账户已被禁用')
        else:
            LoginLog.objects.create(
                username=emp_no,
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                action='login_fail',
                message='登录失败：用户名或密码错误'
            )
            messages.error(request, '工号或密码错误')
    
    return render(request, 'accounts/login.html')


def user_logout(request):
    if request.user.is_authenticated:
        LoginLog.objects.create(
            user=request.user,
            username=request.user.emp_no,
            ip_address=get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', ''),
            action='logout',
            message='退出登录'
        )
    logout(request)
    return redirect('login')


def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


@login_required
def index(request):
    return redirect('dashboard')


@login_required
def dashboard(request):
    from apps.assets.models import Device, AssetCategory
    from apps.inventory.models import InventoryPlan, InventoryTask
    from apps.todos.models import Todo
    
    # 全部设备（排除已报废设备）
    all_devices = Device.objects.exclude(status='scrapped')
    all_stats = {
        'total': all_devices.count(),
        'normal': all_devices.filter(status='normal').count(),
        'fault': all_devices.filter(status='fault').count(),
        'unused': all_devices.filter(status='unused').count(),
    }
    
    # 在账设备（固资在账为是，排除已报废）
    fixed_devices = Device.objects.filter(is_fixed=True).exclude(status='scrapped')
    fixed_stats = {
        'total': fixed_devices.count(),
        'normal': fixed_devices.filter(status='normal').count(),
        'fault': fixed_devices.filter(status='fault').count(),
        'unused': fixed_devices.filter(status='unused').count(),
    }
    
    # 在账电脑类（资产分类包含台式机，排除已报废）
    desktop_categories = AssetCategory.objects.filter(name__icontains='台式机')
    computer_devices = fixed_devices.filter(category__in=desktop_categories)
    computer_stats = {
        'total': computer_devices.count(),
        'normal': computer_devices.filter(status='normal').count(),
        'fault': computer_devices.filter(status='fault').count(),
        'unused': computer_devices.filter(status='unused').count(),
    }
    
    stats = {
        'all': all_stats,
        'fixed': fixed_stats,
        'computer': computer_stats,
        'pending_tasks': InventoryTask.objects.filter(status='pending').count(),
        'my_todos': Todo.objects.filter(assignee=request.user, status='pending').count(),
    }
    
    return render(request, 'accounts/dashboard.html', {'stats': stats})


@login_required
@permission_required('user')
def user_list(request):
    search = request.GET.get('search', '')
    dept_id = request.GET.get('department', '')
    role_id = request.GET.get('role', '')
    
    users = User.objects.select_related('department', 'role').all()
    
    if search:
        users = users.filter(Q(emp_no__icontains=search) | Q(realname__icontains=search) | Q(phone__icontains=search))
    if dept_id:
        users = users.filter(department_id=dept_id)
    if role_id:
        users = users.filter(role_id=role_id)
    
    # 获取页码大小，默认20
    page_size = int(request.GET.get('page_size', 20))
    if page_size not in [20, 50, 100, 200]:
        page_size = 20
    
    paginator = Paginator(users, page_size)
    page = request.GET.get('page', 1)
    users = paginator.get_page(page)
    
    # 计算分页范围
    current_page = users.number
    total_pages = paginator.num_pages
    
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        page_range.extend([1, 2])
        start = max(3, current_page - 2)
        end = min(total_pages - 1, current_page + 2)
        if start > 3:
            page_range.append('...')
        page_range.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_range.append('...')
        page_range.extend([total_pages - 1, total_pages])
        
        # 去重（保留顺序，解决窗口与首尾重叠问题）
        seen = set()
        deduped = []
        for p in page_range:
            if p not in seen:
                seen.add(p)
                deduped.append(p)
        page_range = deduped
    
    departments = Department.objects.all()
    roles = Role.objects.all()
    
    return render(request, 'accounts/user_list.html', {
        'users': users,
        'departments': departments,
        'roles': roles,
        'page_range': page_range,
        'page_size': page_size,
    })


@login_required
@permission_required('user_create')
def user_create(request):
    if request.method == 'POST':
        emp_no = request.POST.get('emp_no')
        realname = request.POST.get('realname')
        password = request.POST.get('password', 'password')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        gender = request.POST.get('gender')
        department_id = request.POST.get('department')
        role_id = request.POST.get('role')
        
        if User.objects.filter(emp_no=emp_no).exists():
            messages.error(request, '工号已存在')
        else:
            user = User.objects.create_user(
                username=emp_no,
                emp_no=emp_no,
                realname=realname,
                password=password,
                email=email,
                phone=phone,
                gender=gender,
                department_id=department_id if department_id else None,
                role_id=role_id if role_id else None,
            )
            messages.success(request, '用户创建成功')
            return redirect('user_list')
    
    departments = Department.objects.all()
    roles = Role.objects.all()
    default_role = Role.objects.filter(code='user').first()
    return render(request, 'accounts/user_form.html', {
        'departments': departments,
        'roles': roles,
        'default_role_id': default_role.id if default_role else None,
    })


@login_required
@permission_required('user_edit')
def user_edit(request, pk):
    user = User.objects.get(pk=pk)
    
    if request.method == 'POST':
        user.realname = request.POST.get('realname')
        user.email = request.POST.get('email')
        user.phone = request.POST.get('phone')
        user.gender = request.POST.get('gender')
        user.department_id = request.POST.get('department') or None
        user.role_id = request.POST.get('role') or None
        user.is_active = request.POST.get('is_active') == 'on'
        
        password = request.POST.get('password')
        if password:
            user.set_password(password)
        
        user.save()
        messages.success(request, '用户更新成功')
        return redirect('user_list')
    
    departments = Department.objects.all()
    roles = Role.objects.all()
    return render(request, 'accounts/user_form.html', {'edit_user': user, 'departments': departments, 'roles': roles})


@login_required
@permission_required('user_delete')
def user_delete(request, pk):
    if request.method == 'POST':
        user = User.objects.get(pk=pk)
        if user.is_superuser:
            messages.error(request, '不能删除超级管理员')
        else:
            user.delete()
            messages.success(request, '用户删除成功')
    return redirect('user_list')


@login_required
@permission_required('user')
def user_detail(request, pk):
    user = User.objects.select_related('department', 'role').get(pk=pk)
    return render(request, 'accounts/user_detail.html', {'user': user})


@login_required
@permission_required('user_reset_password')
def user_reset_password(request, pk):
    user = User.objects.get(pk=pk)
    if request.method == 'POST':
        password = request.POST.get('password', 'password')
        user.set_password(password)
        user.save()
        messages.success(request, f'密码已重置为: {password}')
        return redirect('user_list')
    return render(request, 'accounts/user_reset_password.html', {'user': user})


@login_required
@permission_required('user_import')
def user_download_template(request):
    import openpyxl
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '用户导入模板'
    
    headers = ['工号', '姓名', '性别', '部门', '邮箱', '电话']
    ws.append(headers)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 10)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="user_import_template.xlsx"'
    wb.save(response)
    return response


def process_import_task(task_id, file_content):
    """后台处理导入任务"""
    import openpyxl
    from io import BytesIO
    
    progress = import_progress[task_id]
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
    except Exception as e:
        progress['status'] = 'error'
        progress['errors'].append(f'文件读取失败: {str(e)}')
        return
    
    default_role = Role.objects.filter(code='user').first()
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    progress['total'] = len(rows)
    
    for idx, row in enumerate(rows, start=2):
        progress['current'] = idx - 1
        
        emp_no = str(row[0]).strip() if row[0] else ''
        realname = str(row[1]).strip() if row[1] else ''
        gender = str(row[2]).strip() if row[2] else ''
        dept_name = str(row[3]).strip() if row[3] else ''
        email = str(row[4]).strip() if row[4] else ''
        phone = str(row[5]).strip() if row[5] else ''
        
        progress['current_emp_no'] = emp_no
        
        if not emp_no:
            progress['error'] += 1
            progress['errors'].append(f'第{idx}行：工号不能为空')
            continue
        if not realname:
            progress['error'] += 1
            progress['errors'].append(f'第{idx}行：姓名不能为空')
            continue
        
        department = None
        if dept_name:
            department = Department.objects.filter(name=dept_name).first()
            if not department:
                progress['error'] += 1
                progress['errors'].append(f'第{idx}行：部门"{dept_name}"不存在')
                continue
        
        # 检查用户是否已存在（同时检查 emp_no 和 username）
        user = User.objects.filter(Q(emp_no=emp_no) | Q(username=emp_no)).first()
        
        if user:
            need_update = False
            if gender and gender != user.gender:
                user.gender = gender
                need_update = True
            if email and email != user.email:
                user.email = email
                need_update = True
            if phone and phone != user.phone:
                user.phone = phone
                need_update = True
            if department and department != user.department:
                user.department = department
                need_update = True
            
            if need_update:
                user.save()
                progress['update'] += 1
        else:
            try:
                User.objects.create_user(
                    username=emp_no,
                    emp_no=emp_no,
                    realname=realname,
                    password='password',
                    gender=gender,
                    email=email,
                    phone=phone,
                    department=department,
                    role=default_role,
                )
                progress['success'] += 1
            except Exception as e:
                progress['error'] += 1
                progress['errors'].append(f'第{idx}行：创建失败 - {str(e)}')
        
        time.sleep(0.01)  # 模拟处理时间，让前端能看到进度
    
    progress['current'] = progress['total']
    progress['status'] = 'completed'


@login_required
@csrf_exempt
def user_import(request):
    if request.method == 'POST':
        import openpyxl
        
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'message': '请选择要导入的文件'})
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': '请上传Excel文件(.xlsx或.xls)'})
        
        # 生成任务ID
        task_id = str(uuid.uuid4())[:8]
        
        # 读取文件内容
        file_content = file.read()
        
        # 初始化进度
        import_progress[task_id] = {
            'status': 'processing',
            'total': 0,
            'current': 0,
            'success': 0,
            'update': 0,
            'error': 0,
            'errors': [],
            'current_emp_no': ''
        }
        
        # 启动后台任务处理导入
        thread = threading.Thread(target=process_import_task, args=(task_id, file_content))
        thread.daemon = True
        thread.start()
        
        return JsonResponse({'success': True, 'task_id': task_id})
    
    return render(request, 'accounts/user_import.html')


@login_required
def import_progress_api(request):
    task_id = request.GET.get('task_id')
    if task_id and task_id in import_progress:
        return JsonResponse(import_progress[task_id])
    return JsonResponse({'status': 'not_found'})


@login_required
@permission_required('user_batch_delete')
def user_batch_delete(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        User.objects.filter(id__in=ids).delete()
        messages.success(request, f'成功删除 {len(ids)} 个用户')
    return redirect('user_list')


@login_required
@permission_required('user_batch_enable')
def user_batch_enable(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        User.objects.filter(id__in=ids).update(is_active=True)
        messages.success(request, f'成功启用 {len(ids)} 个用户')
    return redirect('user_list')


@login_required
@permission_required('user_batch_disable')
def user_batch_disable(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        User.objects.filter(id__in=ids).update(is_active=False)
        messages.success(request, f'成功禁用 {len(ids)} 个用户')
    return redirect('user_list')


@login_required
@permission_required('department')
def department_list(request):
    departments = Department.objects.filter(parent__isnull=True).prefetch_related('children')
    return render(request, 'accounts/department_list.html', {'departments': departments})


@login_required
@permission_required('department_create')
def department_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        parent_id = request.POST.get('parent')
        description = request.POST.get('description')
        sort = request.POST.get('sort', 0)
        
        if Department.objects.filter(code=code).exists():
            messages.error(request, '部门编码已存在')
        else:
            Department.objects.create(
                name=name,
                code=code,
                parent_id=parent_id if parent_id else None,
                description=description,
                sort=sort
            )
            messages.success(request, '部门创建成功')
            return redirect('department_list')
    
    departments = Department.objects.all()
    return render(request, 'accounts/department_form.html', {'departments': departments})


@login_required
@permission_required('department_edit')
def department_edit(request, pk):
    dept = Department.objects.get(pk=pk)
    
    if request.method == 'POST':
        dept.name = request.POST.get('name')
        dept.code = request.POST.get('code')
        dept.parent_id = request.POST.get('parent') or None
        dept.description = request.POST.get('description')
        dept.sort = request.POST.get('sort', 0)
        dept.save()
        messages.success(request, '部门更新成功')
        return redirect('department_list')
    
    departments = Department.objects.exclude(pk=pk)
    return render(request, 'accounts/department_form.html', {'department': dept, 'departments': departments})


@login_required
@permission_required('department_delete')
def department_delete(request, pk):
    if request.method == 'POST':
        dept = Department.objects.get(pk=pk)
        if dept.children.exists() or dept.users.exists():
            messages.error(request, '该部门下有子部门或用户，无法删除')
        else:
            dept.delete()
            messages.success(request, '部门删除成功')
    return redirect('department_list')


@login_required
@permission_required('role')
def role_list(request):
    search = request.GET.get('search', '')
    roles = Role.objects.prefetch_related('permissions', 'users')
    if search:
        roles = roles.filter(Q(name__icontains=search) | Q(code__icontains=search))
    roles = roles.order_by('sort', 'id')
    
    paginator = Paginator(roles, 20)
    page = request.GET.get('page', 1)
    roles = paginator.get_page(page)
    
    return render(request, 'accounts/role_list.html', {'roles': roles})


@login_required
@permission_required('role_create')
def role_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        description = request.POST.get('description')
        sort = request.POST.get('sort', 0)
        permission_ids = request.POST.getlist('permissions')
        
        role = Role.objects.create(
            name=name,
            code=code,
            description=description,
            sort=sort
        )
        if permission_ids:
            role.permissions.set(permission_ids)
        messages.success(request, '角色创建成功')
        return redirect('role_list')
    
    perm_tree = build_permission_tree()
    
    return render(request, 'accounts/role_form.html', {'perm_tree': perm_tree})


def build_permission_tree():
    """构建权限树结构"""
    all_perms = Permission.objects.all().order_by('sort', 'id')
    
    # 构建权限树结构：一级菜单 -> 二级菜单 -> 按钮/三级菜单
    perm_tree = []
    top_menus = all_perms.filter(parent__isnull=True)
    
    for top_menu in top_menus:
        menu_data = {
            'menu': top_menu,
            'children': []
        }
        # 获取二级菜单/按钮
        second_level = all_perms.filter(parent=top_menu)
        for second_item in second_level:
            if second_item.type == 'menu':
                second_data = {
                    'menu': second_item,
                    'buttons': [],
                    'children': []
                }
                # 获取三级按钮
                buttons = all_perms.filter(parent=second_item, type='button')
                second_data['buttons'] = list(buttons)
                
                # 获取三级菜单
                third_menus = all_perms.filter(parent=second_item, type='menu')
                for third_menu in third_menus:
                    third_data = {
                        'menu': third_menu,
                        'buttons': list(all_perms.filter(parent=third_menu, type='button'))
                    }
                    second_data['children'].append(third_data)
                
                menu_data['children'].append(second_data)
            else:
                # 一级菜单下的直接按钮
                menu_data.setdefault('buttons', []).append(second_item)
        
        perm_tree.append(menu_data)
    
    return perm_tree


@login_required
@permission_required('role_edit')
def role_edit(request, pk):
    role = get_object_or_404(Role, pk=pk)
    
    # 禁止编辑内置超级管理员角色
    if role.code == 'superuser':
        messages.error(request, '系统内置角色不可编辑')
        return redirect('role_list')
    
    if request.method == 'POST':
        role.name = request.POST.get('name')
        role.code = request.POST.get('code')
        role.description = request.POST.get('description')
        role.sort = request.POST.get('sort', 0)
        role.permissions.set(request.POST.getlist('permissions'))
        role.save()
        messages.success(request, '角色更新成功')
        return redirect('role_list')
    
    perm_tree = build_permission_tree()
    role_perm_ids = set(role.permissions.values_list('id', flat=True))
    
    return render(request, 'accounts/role_form.html', {
        'role': role,
        'perm_tree': perm_tree,
        'role_perm_ids': role_perm_ids,
        'is_system_role': role.code == 'superuser'
    })


@login_required
@permission_required('role')
def role_detail(request, pk):
    role = get_object_or_404(Role.objects.prefetch_related('permissions', 'users'), pk=pk)
    users = role.users.all().select_related('department')
    
    perm_tree = build_permission_tree()
    role_perm_ids = set(role.permissions.values_list('id', flat=True))
    
    return render(request, 'accounts/role_detail.html', {
        'role': role,
        'perm_tree': perm_tree,
        'role_perm_ids': role_perm_ids,
        'users': users,
        'perm_count': role.permissions.count(),
        'user_count': users.count(),
    })


@login_required
@permission_required('role_delete')
def role_delete(request, pk):
    if request.method == 'POST':
        role = get_object_or_404(Role, pk=pk)
        # 禁止删除内置角色
        if role.code == 'superuser':
            messages.error(request, '系统内置角色不可删除')
        elif role.users.exists():
            messages.error(request, '该角色下有用户，无法删除')
        else:
            role.delete()
            messages.success(request, '角色删除成功')
    return redirect('role_list')


def get_permissions(request):
    if not request.user.is_authenticated:
        return JsonResponse({'permissions': []})
    
    permissions = []
    if request.user.is_superuser:
        perms = Permission.objects.filter(is_visible=True)
    else:
        perms = Permission.objects.filter(
            roles__users=request.user,
            is_visible=True
        ).distinct()
    
    for p in perms:
        permissions.append({
            'id': p.id,
            'name': p.name,
            'code': p.code,
            'type': p.type,
            'module': p.module,
        })
    
    return JsonResponse({'permissions': permissions})


def get_menu(request):
    if not request.user.is_authenticated:
        return JsonResponse({'menu': []})
    
    if request.user.is_superuser:
        menus = Permission.objects.filter(type='menu', parent__isnull=True, is_visible=True).prefetch_related('children__children__children')
    else:
        menus = Permission.objects.filter(
            roles__users=request.user,
            type='menu',
            parent__isnull=True,
            is_visible=True
        ).distinct().prefetch_related('children__children__children')
    
    result = []
    for menu in menus:
        item = {
            'id': menu.id,
            'name': menu.name,
            'code': menu.code,
            'children': []
        }
        for child in menu.children.all():
            child_item = {
                'id': child.id,
                'name': child.name,
                'code': child.code,
                'children': []
            }
            for c in child.children.all():
                child_item['children'].append({
                    'id': c.id,
                    'name': c.name,
                    'code': c.code,
                })
            item['children'].append(child_item)
        result.append(item)
    
    return JsonResponse({'menu': result})
