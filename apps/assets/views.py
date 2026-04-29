from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.conf import settings
import json
import qrcode
import os
import uuid
import string
from io import BytesIO
import random
from datetime import datetime
import threading
import time

from .models import (
    Device, AssetCategory, AssetLocation, DeviceField, DeviceFieldValue,
    Workstation, MapElement, MapBackground, LocationAreaBinding,
    Software, SoftwareCategory, SoftwareLicense, SoftwareField, SoftwareFieldValue,
    Consumable, ConsumableCategory, ConsumableRecord,
    ServiceType, ServiceRequest, ServiceLog, ServiceContract, AssetLog, LabelTemplate
)
from apps.accounts.models import User, Department
from apps.accounts.decorators import permission_required
from apps.settings.views import get_config_value
from apps.settings.llm_service import is_llm_enabled, call_llm, get_llm_config

# 全局变量存储导入进度
device_import_progress = {}
software_import_progress = {}


def is_point_in_polygon(x, y, points):
    """使用射线法判断点是否在多边形内"""
    n = len(points)
    if n < 3:
        return False
    inside = False
    p1x, p1y = points[0]['x'], points[0]['y']
    for i in range(1, n + 1):
        p2x, p2y = points[i % n]['x'], points[i % n]['y']
        if y > min(p1y, p2y):
            if y <= max(p1y, p2y):
                if x <= max(p1x, p2x):
                    if p1y != p2y:
                        xinters = (y - p1y) * (p2x - p1x) / (p2y - p1y) + p1x
                    if p1x == p2x or x <= xinters:
                        inside = not inside
        p1x, p1y = p2x, p2y
    return inside


def update_workstation_area(workstation):
    """更新工位的区域关联"""
    if not workstation.location_id or workstation.location.level != 3:
        return
    
    # 获取当前位置下的所有区域
    regions = AssetLocation.objects.filter(parent=workstation.location, level=4).exclude(area_points='')
    
    for region in regions:
        try:
            points = json.loads(region.area_points)
            if is_point_in_polygon(workstation.x, workstation.y, points):
                if workstation.area_id != region.id:
                    workstation.area = region
                    workstation.save(update_fields=['area'])
                return
        except:
            continue
    
    # 如果不在任何区域内，清空area字段
    if workstation.area_id:
        workstation.area = None
        workstation.save(update_fields=['area'])


def build_location_tree(location):
    children = location.children.all().order_by('sort', 'code')
    return {
        'id': location.id,
        'name': location.name,
        'code': location.code,
        'level': location.level,
        'children': [build_location_tree(child) for child in children]
    }


def get_location_tree_data():
    roots = AssetLocation.objects.filter(parent__isnull=True).order_by('sort', 'code')
    return [build_location_tree(loc) for loc in roots]


@login_required
@permission_required('device')
def device_list(request):
    # 检查是否有搜索筛选参数传入（不含分页参数）
    has_filter_params = any([
        request.GET.get('search'),
        request.GET.get('category'),
        request.GET.get('location'),
        request.GET.get('secret_level'),
        request.GET.get('status'),
        request.GET.get('is_fixed'),
        request.GET.get('is_secret'),
        request.GET.get('secret_category'),
    ])
    
    # 如果没有筛选参数且 session 中有存储的筛选条件，则重定向带参数
    if not has_filter_params and 'device_list_filters' in request.session:
        filters = request.session['device_list_filters']
        # 只保留筛选参数，排除分页参数
        filter_keys = ['search', 'category', 'location', 'secret_level', 'status', 'is_fixed', 'is_secret', 'secret_category']
        params = '&'.join([f'{k}={v}' for k, v in filters.items() if v and k in filter_keys])
        if params:
            return redirect(f'/assets/?{params}')
        else:
            # 没有有效筛选参数，清除 session
            del request.session['device_list_filters']
    
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    location_id = request.GET.get('location', '')
    secret_level = request.GET.get('secret_level', '')
    status = request.GET.get('status', '')
    is_fixed = request.GET.get('is_fixed', '')
    is_secret = request.GET.get('is_secret', '')
    secret_category = request.GET.get('secret_category', '')
    page = request.GET.get('page', '1')
    page_size_param = request.GET.get('page_size', '')
    
    # 存储筛选参数到 session（只存搜索筛选，不存分页）
    if has_filter_params:
        request.session['device_list_filters'] = {
            'search': search,
            'category': category_id,
            'location': location_id,
            'secret_level': secret_level,
            'status': status,
            'is_fixed': is_fixed,
            'is_secret': is_secret,
            'secret_category': secret_category,
        }
    
    devices = Device.objects.select_related('category', 'location', 'user', 'department', 'workstation').exclude(status='scrapped').order_by('id').all()
    
    if search:
        devices = devices.filter(
            Q(asset_no__icontains=search) | 
            Q(name__icontains=search) | 
            Q(serial_no__icontains=search) |
            Q(device_no__icontains=search) |
            Q(model__icontains=search) |
            Q(user__realname__icontains=search) |
            Q(department__name__icontains=search) |
            Q(mac_address__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(remarks__icontains=search) |
            Q(category__name__icontains=search) |
            Q(location__name__icontains=search) |
            Q(location_text__icontains=search) |
            Q(workstation__workstation_code__icontains=search) |
            Q(os_name__icontains=search) |
            Q(asset_card_no__icontains=search) |
            Q(disk_serial__icontains=search)
        ).distinct()
    if category_id:
        devices = devices.filter(category_id=category_id)
    if location_id:
        devices = devices.filter(location_id=location_id)
    if secret_level:
        devices = devices.filter(secret_level=secret_level)
    if status:
        devices = devices.filter(status=status)
    if is_fixed:
        devices = devices.filter(is_fixed=True)
    if is_secret:
        devices = devices.filter(is_secret=True)
    if secret_category:
        devices = devices.filter(secret_category=secret_category)
    
    # 获取页码大小，默认20
    page_size = int(request.GET.get('page_size', 20))
    if page_size not in [20, 50, 100, 200]:
        page_size = 20
    
    paginator = Paginator(devices, page_size)
    page = request.GET.get('page', 1)
    devices = paginator.get_page(page)
    
    # 计算分页范围
    current_page = devices.number
    total_pages = paginator.num_pages
    
    # 生成页码列表：前2页 ... 当前页前后3页 ... 后2页
    page_range = []
    if total_pages <= 7:
        # 总页数少于7页，显示所有页码
        page_range = list(range(1, total_pages + 1))
    else:
        # 总是显示前2页
        page_range.extend([1, 2])
        
        # 计算中间范围
        start = max(3, current_page - 2)
        end = min(total_pages - 1, current_page + 2)
        
        # 添加省略号或页码
        if start > 3:
            page_range.append('...')
        page_range.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_range.append('...')
        
        # 总是显示后2页
        page_range.extend([total_pages - 1, total_pages])
        
        # 去重（保留顺序，解决窗口与首尾重叠问题）
        seen = set()
        deduped = []
        for p in page_range:
            if p not in seen:
                seen.add(p)
                deduped.append(p)
        page_range = deduped
    
    categories = AssetCategory.objects.all().order_by('code')
    locations = AssetLocation.objects.filter(parent__isnull=True).prefetch_related('children')
    
    # 获取台账分类选项
    secret_categories = Device.objects.filter(
        is_secret=True, 
        secret_category__isnull=False
    ).exclude(secret_category='').values_list('secret_category', flat=True).distinct().order_by('secret_category')
    
    # 获取用户自定义的字段显示配置
    visible_field_keys = request.session.get('device_visible_fields', None)
    if visible_field_keys:
        visible_fields = DeviceField.objects.filter(field_key__in=visible_field_keys).order_by('sort')
    else:
        visible_fields = DeviceField.objects.filter(is_visible=True).order_by('sort')
    
    # 获取卡片可见字段
    card_visible_fields = DeviceField.objects.filter(is_card_visible=True).order_by('sort')
    
    all_fields = DeviceField.objects.all().order_by('sort')
    
    return render(request, 'assets/device_list.html', {
        'devices': devices,
        'categories': categories,
        'locations': locations,
        'visible_fields': visible_fields,
        'card_visible_fields': card_visible_fields,
        'all_fields': all_fields,
        'secret_categories': secret_categories,
        'page_range': page_range,
        'page_size': page_size,
    })


@login_required
def device_fault_list(request):
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    location_id = request.GET.get('location', '')
    secret_level = request.GET.get('secret_level', '')
    is_fixed = request.GET.get('is_fixed', '')
    is_secret = request.GET.get('is_secret', '')
    secret_category = request.GET.get('secret_category', '')
    
    devices = Device.objects.select_related('category', 'location', 'user', 'department', 'workstation').filter(status='fault').order_by('id')
    
    if search:
        devices = devices.filter(
            Q(asset_no__icontains=search) | 
            Q(name__icontains=search) | 
            Q(serial_no__icontains=search) |
            Q(device_no__icontains=search) |
            Q(model__icontains=search) |
            Q(user__realname__icontains=search) |
            Q(department__name__icontains=search) |
            Q(mac_address__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(remarks__icontains=search) |
            Q(category__name__icontains=search) |
            Q(location__name__icontains=search) |
            Q(location_text__icontains=search) |
            Q(workstation__workstation_code__icontains=search) |
            Q(os_name__icontains=search) |
            Q(asset_card_no__icontains=search) |
            Q(disk_serial__icontains=search)
        ).distinct()
    if category_id:
        devices = devices.filter(category_id=category_id)
    if location_id:
        devices = devices.filter(location_id=location_id)
    if secret_level:
        devices = devices.filter(secret_level=secret_level)
    if is_fixed:
        devices = devices.filter(is_fixed=True)
    if is_secret:
        devices = devices.filter(is_secret=True)
    if secret_category:
        devices = devices.filter(secret_category=secret_category)
    
    # 获取页码大小，默认20
    page_size = int(request.GET.get('page_size', 20))
    if page_size not in [20, 50, 100, 200]:
        page_size = 20
    
    paginator = Paginator(devices, page_size)
    page = request.GET.get('page', 1)
    devices = paginator.get_page(page)
    
    # 计算分页范围
    current_page = devices.number
    total_pages = paginator.num_pages
    
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        page_range.extend([1, 2])
        start = max(3, current_page - 2)
        end = min(total_pages - 1, current_page + 2)
        if start > 3:
            page_range.append('...')
        page_range.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_range.append('...')
        page_range.extend([total_pages - 1, total_pages])
    
    categories = AssetCategory.objects.all().order_by('code')
    locations = AssetLocation.objects.filter(parent__isnull=True).prefetch_related('children')
    secret_categories = Device.objects.filter(
        is_secret=True, 
        secret_category__isnull=False
    ).exclude(secret_category='').values_list('secret_category', flat=True).distinct().order_by('secret_category')
    
    visible_field_keys = request.session.get('device_visible_fields', None)
    if visible_field_keys:
        visible_fields = DeviceField.objects.filter(field_key__in=visible_field_keys).order_by('sort')
    else:
        visible_fields = DeviceField.objects.filter(is_visible=True).order_by('sort')
    
    all_fields = DeviceField.objects.all().order_by('sort')
    
    return render(request, 'assets/device_fault_list.html', {
        'devices': devices,
        'categories': categories,
        'locations': locations,
        'visible_fields': visible_fields,
        'all_fields': all_fields,
        'secret_categories': secret_categories,
        'page_range': page_range,
        'page_size': page_size,
    })


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def device_fault(request, pk):
    device = get_object_or_404(Device, pk=pk)
    fault_reason = request.POST.get('fault_reason', '')
    
    old_status = device.status
    device.status = 'fault'
    device.fault_reason = fault_reason
    device.save()
    
    AssetLog.objects.create(
        device=device,
        user=request.user,
        action='fault',
        field_name='status',
        old_value=old_status,
        new_value='fault',
        remarks=fault_reason
    )
    
    return JsonResponse({'success': True, 'message': '报障成功'})


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def device_repair(request, pk):
    device = get_object_or_404(Device, pk=pk)
    repair_reason = request.POST.get('repair_reason', '')
    
    old_status = device.status
    # 根据设备是否有使用人判断状态
    new_status = 'normal' if device.user_id else 'unused'
    
    device.status = new_status
    device.fault_reason = ''
    device.save()
    
    AssetLog.objects.create(
        device=device,
        user=request.user,
        action='repair',
        field_name='status',
        old_value=old_status,
        new_value=new_status,
        remarks=repair_reason
    )
    
    return JsonResponse({'success': True, 'message': '维修成功'})


@login_required
def device_scrap_list(request):
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    location_id = request.GET.get('location', '')
    secret_level = request.GET.get('secret_level', '')
    is_fixed = request.GET.get('is_fixed', '')
    is_secret = request.GET.get('is_secret', '')
    secret_category = request.GET.get('secret_category', '')
    page_size_param = request.GET.get('page_size', '')
    page = request.GET.get('page', 1)
    
    devices = Device.objects.filter(status='scrapped').order_by('id')
    
    if search:
        devices = devices.filter(
            Q(asset_no__icontains=search) | 
            Q(name__icontains=search) | 
            Q(serial_no__icontains=search) |
            Q(device_no__icontains=search) |
            Q(model__icontains=search) |
            Q(mac_address__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(remarks__icontains=search) |
            Q(category__name__icontains=search) |
            Q(os_name__icontains=search) |
            Q(asset_card_no__icontains=search) |
            Q(disk_serial__icontains=search)
        ).distinct()
    if category_id:
        devices = devices.filter(category_id=category_id)
    if location_id:
        devices = devices.filter(location_id=location_id)
    if secret_level:
        devices = devices.filter(secret_level=secret_level)
    if is_fixed:
        devices = devices.filter(is_fixed=True)
    if is_secret:
        devices = devices.filter(is_secret=True)
    if secret_category:
        devices = devices.filter(secret_category=secret_category)
    
    devices_list = list(devices.select_related('category'))
    for device in devices_list:
        device.pre_department = None
        device.pre_user = None
        device.pre_location = None
        device.pre_workstation = None
        
        if device.scrap_pre_scrap_data:
            try:
                pre_data = json.loads(device.scrap_pre_scrap_data)
                if pre_data.get('department_id'):
                    device.pre_department = Department.objects.filter(id=pre_data['department_id']).first()
                if pre_data.get('user_id'):
                    device.pre_user = User.objects.filter(id=pre_data['user_id']).first()
                if pre_data.get('location_id'):
                    device.pre_location = AssetLocation.objects.filter(id=pre_data['location_id']).first()
                if pre_data.get('workstation_id'):
                    device.pre_workstation = Workstation.objects.filter(id=pre_data['workstation_id']).first()
            except:
                pass
    
    # 获取页码大小，默认20
    page_size = int(page_size_param) if page_size_param else 20
    if page_size not in [20, 50, 100, 200]:
        page_size = 20
    
    paginator = Paginator(devices, page_size)
    devices = paginator.get_page(page)
    
    # 计算分页范围
    current_page = devices.number
    total_pages = paginator.num_pages
    
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        page_range.extend([1, 2])
        start = max(3, current_page - 2)
        end = min(total_pages - 1, current_page + 2)
        if start > 3:
            page_range.append('...')
        page_range.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_range.append('...')
        page_range.extend([total_pages - 1, total_pages])
    
    categories = AssetCategory.objects.all().order_by('code')
    locations = AssetLocation.objects.filter(parent__isnull=True).prefetch_related('children')
    secret_categories = Device.objects.filter(
        is_secret=True, 
        secret_category__isnull=False
    ).exclude(secret_category='').values_list('secret_category', flat=True).distinct().order_by('secret_category')
    
    visible_field_keys = request.session.get('device_visible_fields', None)
    if visible_field_keys:
        visible_fields = DeviceField.objects.filter(field_key__in=visible_field_keys).order_by('sort')
    else:
        visible_fields = DeviceField.objects.filter(is_visible=True).order_by('sort')
    
    all_fields = DeviceField.objects.all().order_by('sort')
    
    return render(request, 'assets/device_scrap_list.html', {
        'devices': devices,
        'categories': categories,
        'locations': locations,
        'visible_fields': visible_fields,
        'all_fields': all_fields,
        'secret_categories': secret_categories,
        'page_range': page_range,
        'page_size': page_size,
    })


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def device_scrap(request, pk):
    device = get_object_or_404(Device, pk=pk)
    
    pre_scrap_data = {
        'department_id': device.department_id,
        'user_id': device.user_id,
        'location_id': device.location_id,
        'workstation_id': device.workstation_id,
    }
    
    old_status = device.status
    device.status = 'scrapped'
    device.scrap_date = timezone.now()
    device.scrap_pre_scrap_data = json.dumps(pre_scrap_data)
    device.department_id = None
    device.user_id = None
    device.location_id = None
    device.workstation_id = None
    device.save()
    
    if pre_scrap_data.get('workstation_id'):
        Workstation.objects.filter(id=pre_scrap_data['workstation_id']).update(status='available')
    
    AssetLog.objects.create(
        device=device,
        user=request.user,
        action='scrap',
        field_name='status',
        old_value=old_status,
        new_value='scrapped',
    )
    
    return JsonResponse({'success': True, 'message': '报废成功'})


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def device_recall(request, pk):
    device = get_object_or_404(Device, pk=pk)
    recall_reason = request.POST.get('recall_reason', '')
    
    pre_scrap_data = {}
    if device.scrap_pre_scrap_data:
        try:
            pre_scrap_data = json.loads(device.scrap_pre_scrap_data)
        except:
            pre_scrap_data = {}
    
    old_status = device.status
    new_status = 'normal' if pre_scrap_data.get('user_id') else 'unused'
    
    device.status = new_status
    device.scrap_date = None
    device.scrap_pre_scrap_data = ''
    device.department_id = pre_scrap_data.get('department_id')
    device.user_id = pre_scrap_data.get('user_id')
    device.location_id = pre_scrap_data.get('location_id')
    device.workstation_id = pre_scrap_data.get('workstation_id')
    device.save()
    
    if pre_scrap_data.get('workstation_id'):
        Workstation.objects.filter(id=pre_scrap_data['workstation_id']).update(status='occupied')
    
    AssetLog.objects.create(
        device=device,
        user=request.user,
        action='revoke',
        field_name='status',
        old_value=old_status,
        new_value=new_status,
        remarks=recall_reason
    )
    
    return JsonResponse({'success': True, 'message': '撤回成功'})


@login_required
@csrf_exempt
def api_save_field_visibility(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            field_keys = body.get('visible_fields', [])
            request.session['device_visible_fields'] = field_keys
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})
    return JsonResponse({'success': False, 'message': '仅支持POST请求'})


@login_required
@csrf_exempt
def api_clear_device_filters(request):
    if request.method == 'POST':
        if 'device_list_filters' in request.session:
            del request.session['device_list_filters']
        return JsonResponse({'success': True})
    return JsonResponse({'success': False, 'message': '仅支持POST请求'})


@login_required
@permission_required('device_create')
def device_create(request):
    if request.method == 'POST':
        category_id = request.POST.get('category')
        category = AssetCategory.objects.get(pk=category_id)
        
        asset_no_input = request.POST.get('asset_no', '').strip()
        auto_number = get_config_value('asset_auto_number', True)
        if asset_no_input:
            asset_no = asset_no_input
        elif auto_number:
            asset_no = generate_asset_no(category)
        else:
            asset_no = ''
        
        user_id = request.POST.get('user') or None
        device = Device.objects.create(
            asset_no=asset_no,
            device_no=request.POST.get('device_no'),
            serial_no=request.POST.get('serial_no'),
            name=request.POST.get('name'),
            model=request.POST.get('model'),
            category=category,
            status='normal' if user_id else 'unused',
            secret_level=request.POST.get('secret_level', 'public'),
            user_id=user_id,
            department_id=request.POST.get('department') or None,
            location_id=request.POST.get('location') or None,
            workstation_id=request.POST.get('workstation') or None,
            location_text=request.POST.get('location_text') or '',
            purchase_date=request.POST.get('purchase_date') or None,
            enable_date=request.POST.get('enable_date') or None,
            install_date=request.POST.get('install_date') or None,
            mac_address=request.POST.get('mac_address'),
            ip_address=request.POST.get('ip_address'),
            os_name=request.POST.get('os_name'),
            disk_serial=request.POST.get('disk_serial'),
            purpose=request.POST.get('purpose'),
            remarks=request.POST.get('remarks'),
            is_fixed=request.POST.get('is_fixed') == 'on',
            asset_card_no=request.POST.get('asset_card_no'),
            is_secret=request.POST.get('is_secret') == 'on',
            secret_category=request.POST.get('secret_category'),
            created_by=request.user,
        )
        
        if device.workstation_id:
            device.workstation.status = 'occupied'
            device.workstation.save()
            # 根据工位的area_id更新设备位置
            if device.workstation.area_id:
                # 工位关联在4级内，设备位置记录4级位置
                device.location = device.workstation.area
            else:
                # 工位没有关联区域，设备位置记录工位的location（3级楼层）
                device.location = device.workstation.location
            device.save(update_fields=['location'])
        
        app_url = get_config_value('app_url', 'http://127.0.0.1:8000').rstrip('/')
        qr = qrcode.QRCode(version=1, box_size=10, border=1)
        qr.add_data(f'{app_url}/assets/view/{device.asset_no}')
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        buffer = BytesIO()
        img.save(buffer, 'PNG')
        if device.qrcode:
            device.qrcode.delete(save=False)
        device.qrcode.save(f'QR-{asset_no}.png', buffer)
        
        if request.FILES.get('photo'):
            save_photo_with_asset_no(device, request.FILES.get('photo'))
            device.photo_updated_at = timezone.now()
            device.save(update_fields=['photo', 'photo_updated_at'])
        
        AssetLog.objects.create(
            device=device,
            user=request.user,
            action='create',
            new_value='新增设备',
            remarks=f'资产编号: {asset_no}'
        )
        
        messages.success(request, f'设备创建成功，资产编号: {asset_no}')
        return redirect('device_list')
    
    categories = AssetCategory.objects.filter(parent__isnull=True).prefetch_related('children')
    locations = AssetLocation.objects.all()
    users = User.objects.all()
    departments = Department.objects.all()
    return render(request, 'assets/device_form.html', {
        'categories': categories,
        'locations': locations,
        'users': users,
        'departments': departments,
        'device': None,
    })


@login_required
@permission_required('device_edit')
def device_edit(request, pk):
    device = get_object_or_404(Device, pk=pk)
    
    if request.method == 'POST':
        old_values = {
            'name': device.name,
            'category': str(device.category),
            'status': device.status,
            'user': str(device.user) if device.user else '',
            'location': str(device.location) if device.location else '',
        }
        old_workstation_id = device.workstation_id
        
        device.name = request.POST.get('name')
        device.asset_no = request.POST.get('asset_no', device.asset_no)
        device.device_no = request.POST.get('device_no')
        device.serial_no = request.POST.get('serial_no')
        device.model = request.POST.get('model')
        device.category_id = request.POST.get('category')
        device.secret_level = request.POST.get('secret_level', 'public')
        user_id = request.POST.get('user') or None
        device.user_id = user_id
        device.department_id = request.POST.get('department') or None
        device.status = 'normal' if user_id else 'unused'
        device.location_id = request.POST.get('location') or None
        device.workstation_id = request.POST.get('workstation') or None
        device.location_text = request.POST.get('location_text') or ''
        device.purchase_date = request.POST.get('purchase_date') or None
        device.enable_date = request.POST.get('enable_date') or None
        device.install_date = request.POST.get('install_date') or None
        device.mac_address = request.POST.get('mac_address')
        device.ip_address = request.POST.get('ip_address')
        device.os_name = request.POST.get('os_name')
        device.disk_serial = request.POST.get('disk_serial')
        device.purpose = request.POST.get('purpose')
        device.remarks = request.POST.get('remarks')
        device.is_fixed = request.POST.get('is_fixed') == 'on'
        device.asset_card_no = request.POST.get('asset_card_no')
        device.is_secret = request.POST.get('is_secret') == 'on'
        device.secret_category = request.POST.get('secret_category')
        
        # Handle file uploads
        if request.FILES.get('qrcode'):
            device.qrcode = request.FILES.get('qrcode')
        elif request.POST.get('qrcode_clear') == '1':
            device.qrcode.delete(save=False)
            device.qrcode = ''
        
        if request.FILES.get('photo'):
            if device.photo:
                device.photo.delete(save=False)
            save_photo_with_asset_no(device, request.FILES.get('photo'))
            device.photo_updated_at = timezone.now()
        elif request.POST.get('photo_clear') == '1':
            device.photo.delete(save=False)
            device.photo = ''
            device.photo_updated_at = None
        
        device.save()
        
        # 根据工位的area_id更新设备位置
        if device.workstation_id:
            if device.workstation.area_id:
                # 工位关联在4级内，设备位置记录4级位置
                device.location = device.workstation.area
            else:
                # 工位没有关联区域，设备位置记录工位的location（3级楼层）
                device.location = device.workstation.location
            device.save(update_fields=['location'])
        
        if device.workstation_id != old_workstation_id:
            if device.workstation_id:
                device.workstation.status = 'occupied'
                device.workstation.save()
            if old_workstation_id:
                old_workstation = Workstation.objects.get(pk=old_workstation_id)
                if not old_workstation.devices.exists():
                    old_workstation.status = 'available'
                    old_workstation.save()
        
        for field, old_val in old_values.items():
            new_val = str(getattr(device, field))
            if old_val != new_val:
                AssetLog.objects.create(
                    device=device,
                    user=request.user,
                    action='update',
                    field_name=field,
                    old_value=old_val,
                    new_value=new_val,
                )
        
        messages.success(request, '设备更新成功')
        return redirect('device_list')
    
    categories = AssetCategory.objects.filter(parent__isnull=True).prefetch_related('children')
    locations = AssetLocation.objects.all()
    users = User.objects.all()
    departments = Department.objects.all()
    return render(request, 'assets/device_form.html', {
        'device': device,
        'categories': categories,
        'locations': locations,
        'users': users,
        'departments': departments,
    })


@login_required
def device_detail(request, pk):
    device = get_object_or_404(Device.objects.select_related('category', 'location', 'user', 'department', 'workstation'), pk=pk)
    
    if device.status == 'scrapped' and device.scrap_pre_scrap_data:
        try:
            pre_data = json.loads(device.scrap_pre_scrap_data)
            device.pre_department = Department.objects.filter(id=pre_data.get('department_id')).first() if pre_data.get('department_id') else None
            device.pre_user = User.objects.filter(id=pre_data.get('user_id')).first() if pre_data.get('user_id') else None
            device.pre_location = AssetLocation.objects.filter(id=pre_data.get('location_id')).first() if pre_data.get('location_id') else None
            device.pre_workstation = Workstation.objects.filter(id=pre_data.get('workstation_id')).first() if pre_data.get('workstation_id') else None
        except:
            device.pre_department = None
            device.pre_user = None
            device.pre_location = None
            device.pre_workstation = None
    
    logs = device.logs.all()[:20]
    all_fields = DeviceField.objects.all().order_by('sort')
    
    ws_location_id = None
    ws_id = None
    if device.workstation_id:
        ws_id = device.workstation_id
        ws_location_id = device.workstation.location_id
    elif device.location_id and device.location.level == 4 and device.location.parent_id:
        ws_location_id = device.location.parent_id
    
    return render(request, 'assets/device_detail.html', {
        'device': device, 'logs': logs, 'all_fields': all_fields,
        'ws_location_id': ws_location_id, 'ws_id': ws_id,
    })


def asset_view(request, asset_no):
    device = get_object_or_404(Device.objects.select_related('category', 'location', 'user', 'department', 'workstation'), asset_no=asset_no)
    
    if device.status == 'scrapped' and device.scrap_pre_scrap_data:
        try:
            pre_data = json.loads(device.scrap_pre_scrap_data)
            device.pre_department = Department.objects.filter(id=pre_data.get('department_id')).first() if pre_data.get('department_id') else None
            device.pre_user = User.objects.filter(id=pre_data.get('user_id')).first() if pre_data.get('user_id') else None
            device.pre_location = AssetLocation.objects.filter(id=pre_data.get('location_id')).first() if pre_data.get('location_id') else None
            device.pre_workstation = Workstation.objects.filter(id=pre_data.get('workstation_id')).first() if pre_data.get('workstation_id') else None
        except:
            device.pre_department = None
            device.pre_user = None
            device.pre_location = None
            device.pre_workstation = None
    
    all_fields = DeviceField.objects.all().order_by('sort')
    
    ws_location_id = None
    ws_id = None
    if device.workstation_id:
        ws_id = device.workstation_id
        ws_location_id = device.workstation.location_id
    elif device.location_id and device.location.level == 4 and device.location.parent_id:
        ws_location_id = device.location.parent_id
    
    return render(request, 'assets/asset_view.html', {
        'device': device, 'all_fields': all_fields,
        'ws_location_id': ws_location_id, 'ws_id': ws_id,
    })


@login_required
@permission_required('device_delete')
def device_delete(request, pk):
    if request.method == 'POST':
        device = get_object_or_404(Device, pk=pk)
        AssetLog.objects.create(
            device=device,
            user=request.user,
            action='delete',
            old_value=str(device),
        )
        device.delete()
        messages.success(request, '设备删除成功')
    return redirect('device_list')


@login_required
@permission_required('device_batch_delete')
def device_batch_delete(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        count = Device.objects.filter(id__in=ids).delete()[0]
        messages.success(request, f'已删除 {count} 台设备')
    return redirect('device_list')


@login_required
def device_batch_assign(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        user_id = request.POST.get('user_id')
        device_ids = [int(i) for i in ids if i]
        
        devices = Device.objects.filter(id__in=device_ids)
        for device in devices:
            device.user_id = user_id
            device.status = 'normal'
            device.save()
            AssetLog.objects.create(
                device=device,
                user=request.user,
                action='assign',
                new_value=f'分配给 {User.objects.get(pk=user_id).realname}',
            )
        
        messages.success(request, f'已分配 {len(device_ids)} 台设备')
    return redirect('device_list')


@login_required
def device_batch_fault(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        device_ids = [int(i) for i in ids if i]
        
        devices = Device.objects.filter(id__in=device_ids)
        for device in devices:
            device.status = 'fault'
            device.save()
            AssetLog.objects.create(
                device=device,
                user=request.user,
                action='repair',
                new_value='标记为故障',
            )
        
        messages.success(request, f'已标记 {len(device_ids)} 台设备为故障')
    return redirect('device_list')


@login_required
def device_batch_scrap(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        device_ids = [int(i) for i in ids if i]
        
        devices = Device.objects.filter(id__in=device_ids)
        workstation_ids = []
        for device in devices:
            pre_scrap_data = {
                'department_id': device.department_id,
                'user_id': device.user_id,
                'location_id': device.location_id,
                'workstation_id': device.workstation_id,
            }
            
            old_status = device.status
            device.status = 'scrapped'
            device.scrap_date = timezone.now()
            device.scrap_pre_scrap_data = json.dumps(pre_scrap_data)
            device.department_id = None
            device.user_id = None
            device.location_id = None
            device.workstation_id = None
            device.save()
            
            if pre_scrap_data.get('workstation_id'):
                workstation_ids.append(pre_scrap_data['workstation_id'])
            
            AssetLog.objects.create(
                device=device,
                user=request.user,
                action='scrap',
                field_name='status',
                old_value=old_status,
                new_value='scrapped',
            )
        
        if workstation_ids:
            Workstation.objects.filter(id__in=workstation_ids).update(status='available')
        
        messages.success(request, f'已报废 {len(device_ids)} 台设备')
    return redirect('device_scrap_list')


@login_required
def device_print_label(request, pk):
    device = get_object_or_404(Device, pk=pk)
    return render(request, 'assets/device_label.html', {'device': device})


@login_required
def device_scan(request, device_id):
    device = get_object_or_404(Device.objects.select_related('category', 'location', 'user'), pk=device_id)
    return render(request, 'assets/device_scan.html', {'device': device})


def generate_asset_no(category):
    prefix = category.get_full_code()
    last_device = Device.objects.filter(asset_no__startswith=prefix).order_by('-asset_no').first()
    
    if last_device:
        try:
            last_num = int(last_device.asset_no.split('-')[-1])
            new_num = last_num + 1
        except:
            new_num = 1
    else:
        new_num = 1
    
    return f"{prefix}-{new_num:03d}"


def save_photo_with_asset_no(device, photo_file):
    """
    保存设备照片：
    1. 压缩图片到1MB以内
    2. 文件名为资产编号（不添加随机后缀）
    3. 如果设备已有照片，覆盖旧文件而非创建新文件
    """
    if not photo_file:
        return
    
    from PIL import Image
    from django.core.files.base import ContentFile
    import io
    
    ext = os.path.splitext(photo_file.name)[1].lower() or '.jpg'
    filename = f"{device.asset_no}{ext}"
    
    # 先删除旧文件（从磁盘真正删除，避免Django添加随机后缀）
    if device.photo:
        try:
            old_path = device.photo.path
            if os.path.isfile(old_path):
                os.remove(old_path)
        except Exception:
            pass
        device.photo.name = ''
    
    # 打开并处理图片
    image = Image.open(photo_file)
    original_format = image.format or 'JPEG'
    
    # 转换颜色模式（PNG支持透明，JPEG不支持）
    if original_format == 'PNG' and image.mode in ('RGBA', 'P'):
        # 保持PNG格式以保留透明度
        target_format = 'PNG'
    else:
        target_format = 'JPEG'
        if image.mode != 'RGB':
            image = image.convert('RGB')
    
    max_size = 1024 * 1024  # 1MB
    max_width = 1600
    min_quality = 20
    
    # 第一步：先缩小尺寸，宽度限制在1600以内
    if image.width > max_width:
        ratio = max_width / image.width
        new_size = (max_width, int(image.height * ratio))
        image = image.resize(new_size, Image.LANCZOS)
    
    # PNG转为JPEG（PNG太大，不适合设备照片）
    if target_format == 'PNG':
        target_format = 'JPEG'
        if image.mode != 'RGB':
            image = image.convert('RGB')
        ext = '.jpg'
        filename = f"{device.asset_no}{ext}"
    
    def save_to_buffer(img, q):
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=q, optimize=True)
        return buf
    
    # 第二步：如果还是大于1MB，逐步降低质量
    quality = 85
    buffer = save_to_buffer(image, quality)
    
    while buffer.tell() > max_size and quality > min_quality:
        quality -= 5
        buffer = save_to_buffer(image, quality)
    
    buffer.seek(0)
    device.photo.save(filename, ContentFile(buffer.read()), save=True)


@login_required
def category_list(request):
    categories = AssetCategory.objects.filter(
        parent__isnull=True
    ).prefetch_related(
        'children__children__children',
        'children__children',
        'children',
        'devices'
    ).order_by('sort', 'code', 'id')
    return render(request, 'assets/category_list.html', {'categories': categories})


@login_required
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        code = request.POST.get('code')
        parent_id = request.POST.get('parent')
        description = request.POST.get('description')
        sort = request.POST.get('sort', 0)
        
        level = 1
        if parent_id:
            parent = AssetCategory.objects.filter(pk=parent_id).first()
            if parent:
                level = parent.level + 1
        
        if code and AssetCategory.objects.filter(code=code, parent_id=parent_id or None).exists():
            messages.error(request, '同级分类中编码已存在')
        else:
            AssetCategory.objects.create(
                name=name,
                code=code,
                parent_id=parent_id if parent_id else None,
                level=level,
                description=description,
                sort=sort
            )
            messages.success(request, '分类创建成功')
            return redirect('category_list')
    
    all_cats = AssetCategory.objects.order_by('sort', 'id').all()
    flat_cats = build_category_tree(all_cats)
    return render(request, 'assets/category_form.html', {'categories': flat_cats})


def build_category_tree(all_cats):
    """Build category tree and return flat list in tree order"""
    # Group by parent_id
    by_parent = {}
    for cat in all_cats:
        pid = cat.parent_id or 0
        by_parent.setdefault(pid, []).append(cat)
    
    # Sort each level by sort, id
    for pid in by_parent:
        by_parent[pid].sort(key=lambda c: (c.sort or 0, c.id))
    
    # Recursive flatten
    result = []
    def flatten(parent_id):
        for cat in by_parent.get(parent_id, []):
            result.append(cat)
            flatten(cat.id)
    flatten(0)
    return result


@login_required
def category_edit(request, pk):
    category = get_object_or_404(AssetCategory, pk=pk)
    
    if request.method == 'POST':
        category.name = request.POST.get('name')
        category.code = request.POST.get('code')
        parent_id = request.POST.get('parent') or None
        
        if parent_id:
            parent = AssetCategory.objects.filter(pk=parent_id).first()
            if parent:
                category.level = parent.level + 1
        else:
            category.level = 1
        
        category.parent_id = parent_id
        category.description = request.POST.get('description')
        category.sort = request.POST.get('sort', 0)
        
        # Check for duplicate code within same parent
        if category.code and AssetCategory.objects.filter(code=category.code, parent_id=parent_id or None).exclude(pk=pk).exists():
            messages.error(request, '同级分类中编码已存在')
        else:
            category.save()
            messages.success(request, '分类更新成功')
            return redirect('category_list')
    
    all_cats = AssetCategory.objects.order_by('sort', 'id').exclude(pk=pk)
    flat_cats = build_category_tree(all_cats)
    return render(request, 'assets/category_form.html', {'category': category, 'categories': flat_cats})


@login_required
def category_delete(request, pk):
    if request.method in ['POST', 'GET']:
        category = get_object_or_404(AssetCategory, pk=pk)
        if category.children.exists() or category.devices.exists():
            messages.error(request, '该分类下有子分类或设备，无法删除')
        else:
            category.delete()
            messages.success(request, '分类删除成功')
    return redirect('category_list')


@login_required
def location_list(request):
    locations = AssetLocation.objects.filter(
        parent__isnull=True
    ).prefetch_related(
        'children__children__children',
        'children__children',
        'children',
        'workstations',
        'devices'
    ).order_by('sort', 'code', 'id')
    return render(request, 'assets/location_list.html', {'locations': locations})


@login_required
def location_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        parent_id = request.POST.get('parent')
        level = int(request.POST.get('level', 1))
        park_code = request.POST.get('park_code', '').strip().upper()
        building_code = request.POST.get('building_code', '').strip().upper()
        floor_code = request.POST.get('floor_code', '').strip()
        room_code = request.POST.get('room_code', '').strip()
        floor_count = int(request.POST.get('floor_count', 1))
        basement_count = int(request.POST.get('basement_count', 0))
        has_rooftop = request.POST.get('has_rooftop') == 'on'
        description = request.POST.get('description', '')
        sort = int(request.POST.get('sort', 0))
        
        parent = AssetLocation.objects.get(pk=parent_id) if parent_id else None
        if parent:
            level = parent.level + 1
            park_code = parent.code if parent.code else park_code
        
        if level == 1:
            code = park_code if park_code else f"LOC{AssetLocation.objects.count() + 1:03d}"
        elif level == 2:
            parent_code = parent.code if parent and parent.code else ''
            code = f"{parent_code}{building_code}" if parent_code and building_code else f"LOC{AssetLocation.objects.count() + 1:03d}"
        elif level == 3:
            prefix = parent.code if parent and parent.code else "LOC"
            max_floor = AssetLocation.objects.filter(parent=parent).count() if parent else 0
            code = f"{prefix}{max_floor + 1:02d}"
        elif level == 4:
            prefix = parent.code if parent and parent.code else "LOC"
            code = f"{prefix}-{room_code}" if room_code else f"{prefix}-{AssetLocation.objects.count() + 1:03d}"
        else:
            code = f"LOC{AssetLocation.objects.count() + 1:03d}"
        
        if AssetLocation.objects.filter(code=code).exists():
            messages.error(request, '位置编码已存在')
        else:
            new_location = AssetLocation.objects.create(
                name=name,
                code=code,
                parent=parent,
                level=level,
                park_code=park_code,
                building_code=building_code,
                floor_code=floor_code,
                floor_count=floor_count,
                basement_count=basement_count,
                has_rooftop=has_rooftop,
                description=description,
                sort=sort
            )
            
            if level == 2 and (floor_count > 0 or basement_count > 0 or has_rooftop):
                sort_order = 0
                for i in range(1, basement_count + 1):
                    floor_name = f"B{i}层"
                    floor_code_val = f"B{i}"
                    floor_code_full = f"{code}{floor_code_val}"
                    sort_order += 1
                    AssetLocation.objects.create(
                        name=floor_name,
                        code=floor_code_full,
                        parent=new_location,
                        level=3,
                        park_code=park_code,
                        building_code=building_code,
                        floor_code=floor_code_val,
                        description=f"{name} {floor_name}",
                        sort=sort_order
                    )
                
                for i in range(1, floor_count + 1):
                    floor_name = f"{i}楼"
                    floor_code_val = f"{i:02d}"
                    floor_code_full = f"{code}{floor_code_val}"
                    sort_order += 1
                    AssetLocation.objects.create(
                        name=floor_name,
                        code=floor_code_full,
                        parent=new_location,
                        level=3,
                        park_code=park_code,
                        building_code=building_code,
                        floor_code=floor_code_val,
                        description=f"{name} {floor_name}",
                        sort=sort_order
                    )
                
                if has_rooftop:
                    sort_order += 1
                    AssetLocation.objects.create(
                        name="天台",
                        code=f"{code}C1",
                        parent=new_location,
                        level=3,
                        park_code=park_code,
                        building_code=building_code,
                        floor_code="C1",
                        description=f"{name} 天台",
                        sort=sort_order
                    )
                
                total_floors = basement_count + floor_count + (1 if has_rooftop else 0)
                msg_parts = []
                if basement_count > 0:
                    msg_parts.append(f"地下{basement_count}层")
                if floor_count > 0:
                    msg_parts.append(f"地上{floor_count}层")
                if has_rooftop:
                    msg_parts.append("天台")
                messages.success(request, f'位置创建成功，已自动生成 {total_floors} 个楼层（{"，".join(msg_parts)}）')
            else:
                messages.success(request, '位置创建成功')
            
            return redirect('location_list')
    
    locations = AssetLocation.objects.all()
    return render(request, 'assets/location_form.html', {'locations': locations})


@login_required
def location_edit(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    
    if request.method == 'POST':
        location.name = request.POST.get('name')
        parent_id = request.POST.get('parent') or None
        if parent_id:
            parent = AssetLocation.objects.get(pk=parent_id)
            location.level = parent.level + 1
            location.parent = parent
        
        park_code = request.POST.get('park_code', '').strip().upper()
        building_code = request.POST.get('building_code', '').strip().upper()
        floor_code = request.POST.get('floor_code', '').strip()
        room_code = request.POST.get('room_code', '').strip()
        
        if location.level == 1:
            location.code = park_code if park_code else location.code
        elif location.level == 2:
            location.code = f"{park_code}{building_code}" if park_code and building_code else location.code
        elif location.level == 4 and room_code:
            parent = location.parent
            prefix = parent.code if parent and parent.code else "LOC"
            location.code = f"{prefix}-{room_code}"
        
        location.park_code = park_code
        location.building_code = building_code
        location.floor_code = floor_code
        location.room_code = room_code
        location.floor_count = int(request.POST.get('floor_count', 1))
        location.basement_count = int(request.POST.get('basement_count', 0))
        location.has_rooftop = request.POST.get('has_rooftop') == 'on'
        location.map_width = int(request.POST.get('map_width', 1200))
        location.map_height = int(request.POST.get('map_height', 800))
        location.grid_size = int(request.POST.get('grid_size', 50))
        location.has_map = request.POST.get('has_map') == 'on'
        location.default_doorstop_width = float(request.POST.get('default_doorstop_width', 15))
        location.default_snap_threshold = float(request.POST.get('default_snap_threshold', 10))
        location.default_snap_enabled = request.POST.get('default_snap_enabled') == 'on'
        location.description = request.POST.get('description', '')
        location.sort = int(request.POST.get('sort', 0))
        location.save()
        
        # 如果是2级设施，根据新的楼层数生成缺失的3级楼层
        if location.level == 2 and (location.floor_count > 0 or location.basement_count > 0 or location.has_rooftop):
            existing_floors = location.children.filter(level=3)
            existing_count = existing_floors.count()
            
            # 计算需要的总楼层数
            total_needed = location.basement_count + location.floor_count + (1 if location.has_rooftop else 0)
            
            # 如果现有楼层数少于需要的楼层数，则生成新的楼层
            if existing_count < total_needed:
                sort_order = existing_count
                park_code = location.park_code or ''
                building_code = location.building_code or ''
                
                # 生成地下楼层
                for i in range(1, location.basement_count + 1):
                    floor_name = f"B{i}层"
                    floor_code_val = f"B{i}"
                    floor_code_full = f"{location.code}{floor_code_val}"
                    # 检查是否已存在
                    if not AssetLocation.objects.filter(parent=location, floor_code=floor_code_val).exists():
                        sort_order += 1
                        AssetLocation.objects.create(
                            name=floor_name,
                            code=floor_code_full,
                            parent=location,
                            level=3,
                            park_code=park_code,
                            building_code=building_code,
                            floor_code=floor_code_val,
                            description=f"{location.name} {floor_name}",
                            sort=sort_order
                        )
                
                # 生成地上楼层
                for i in range(1, location.floor_count + 1):
                    floor_name = f"{i}楼"
                    floor_code_val = f"{i:02d}"
                    floor_code_full = f"{location.code}{floor_code_val}"
                    # 检查是否已存在
                    if not AssetLocation.objects.filter(parent=location, floor_code=floor_code_val).exists():
                        sort_order += 1
                        AssetLocation.objects.create(
                            name=floor_name,
                            code=floor_code_full,
                            parent=location,
                            level=3,
                            park_code=park_code,
                            building_code=building_code,
                            floor_code=floor_code_val,
                            description=f"{location.name} {floor_name}",
                            sort=sort_order
                        )
                
                # 生成天台
                if location.has_rooftop:
                    floor_code_val = "RT"
                    if not AssetLocation.objects.filter(parent=location, floor_code=floor_code_val).exists():
                        floor_name = "天台"
                        floor_code_full = f"{location.code}{floor_code_val}"
                        sort_order += 1
                        AssetLocation.objects.create(
                            name=floor_name,
                            code=floor_code_full,
                            parent=location,
                            level=3,
                            park_code=park_code,
                            building_code=building_code,
                            floor_code=floor_code_val,
                            description=f"{location.name} {floor_name}",
                            sort=sort_order
                        )
        
        messages.success(request, '位置更新成功')
        return redirect('location_list')
    
    locations = AssetLocation.objects.exclude(pk=pk)
    return render(request, 'assets/location_form.html', {'location': location, 'locations': locations})


from django.views.decorators.http import require_POST


@login_required
@require_POST
def location_delete(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    if location.children.exists() or location.devices.exists() or location.workstations.exists():
        messages.error(request, f'该位置下有子位置、设备或工位，无法删除')
    else:
        # 删除对应的MapElement
        if location.level == 4:
            MapElement.objects.filter(location=location.parent, element_type='region', label=location.name).delete()
            # 清空关联工位的area字段
            Workstation.objects.filter(area=location).update(area=None)
        
        location.delete()
        messages.success(request, '位置删除成功')
    return redirect('location_list')


@login_required
@csrf_exempt
@require_POST
def api_location_reorder(request):
    """批量更新位置排序"""
    try:
        data = json.loads(request.body)
        order_list = data.get('order', [])
        
        for item in order_list:
            location_id = item.get('id')
            sort_order = item.get('sort', 0)
            AssetLocation.objects.filter(pk=location_id).update(sort=sort_order)
        
        return JsonResponse({'success': True, 'message': '排序保存成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def device_map(request):
    # Get all locations with hierarchy
    all_locs = AssetLocation.objects.select_related('parent__parent', 'parent').prefetch_related('workstations', 'devices')
    
    # Build hierarchy: parks -> buildings -> floors
    parks = []
    for park in all_locs.filter(level=1).order_by('sort', 'id'):
        park_data = {
            'id': park.id, 'name': park.name, 'code': park.code,
            'buildings': []
        }
        for building in all_locs.filter(parent=park, level=2).order_by('sort', 'id'):
            building_data = {
                'id': building.id, 'name': building.name, 'code': building.code,
                'floors': []
            }
            for floor in all_locs.filter(parent=building, level=3, has_map=True).order_by('sort', 'id'):
                building_data['floors'].append({
                    'id': floor.id, 'name': floor.name, 'code': floor.code,
                    'workstation_count': floor.workstations.count(),
                    'device_count': floor.devices.count(),
                })
            if building_data['floors']:
                park_data['buildings'].append(building_data)
        if park_data['buildings']:
            parks.append(park_data)
    
    return render(request, 'assets/device_map.html', {'parks': parks})


@login_required
def workstation_manage(request):
    # Get all locations with hierarchy
    all_locs = AssetLocation.objects.select_related('parent__parent', 'parent').prefetch_related('workstations__devices__category', 'workstations__devices__user', 'workstations__devices__department')
    
    # Build hierarchy: parks -> buildings -> floors -> workstations
    parks = []
    for park in all_locs.filter(level=1).order_by('sort', 'id'):
        park_data = {
            'id': park.id, 'name': park.name, 'code': park.code,
            'buildings': []
        }
        for building in all_locs.filter(parent=park, level=2).order_by('sort', 'id'):
            building_data = {
                'id': building.id, 'name': building.name, 'code': building.code,
                'floors': []
            }
            for floor in all_locs.filter(parent=building, level=3).order_by('sort', 'id'):
                workstations = floor.workstations.order_by('workstation_code').prefetch_related('devices__category', 'devices__user', 'devices__department')
                ws_list = []
                for ws in workstations:
                    devices = ws.devices.all()
                    # Auto-generate name
                    display_name = ws.name
                    if not display_name:
                        pc = devices.filter(category__name='台式机').first() or devices.filter(category__name__icontains='计算机').first()
                        if pc and pc.user:
                            display_name = f'{pc.user.realname}的工位'
                        elif devices.exists() and devices.first().department:
                            display_name = f'{devices.first().department.name}工位'
                    ws_list.append({
                        'id': ws.id, 'code': ws.workstation_code, 'name': display_name or '-',
                        'x': ws.x, 'y': ws.y, 'status': ws.status, 'status_display': ws.get_status_display(),
                        'devices': [{'id': d.id, 'asset_no': d.asset_no, 'name': d.name, 'status': d.status} for d in devices],
                        'device_count': devices.count(),
                    })
                if ws_list or floor.has_map:
                    building_data['floors'].append({
                        'id': floor.id, 'name': floor.name, 'code': floor.code,
                        'has_map': floor.has_map,
                        'workstations': ws_list,
                        'workstation_count': len(ws_list),
                    })
            if building_data['floors']:
                park_data['buildings'].append(building_data)
        if park_data['buildings']:
            parks.append(park_data)
    
    return render(request, 'assets/workstation_manage.html', {'parks': parks})


@login_required
def field_list(request):
    fields = DeviceField.objects.all()
    return render(request, 'assets/field_list.html', {'fields': fields})


@login_required
def field_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        field_key = request.POST.get('field_key')
        field_type = request.POST.get('field_type')
        is_required = request.POST.get('is_required') == 'on'
        is_visible = request.POST.get('is_visible') == 'on'
        is_card_visible = request.POST.get('is_card_visible') == 'on'
        options = request.POST.get('options')
        default_value = request.POST.get('default_value')
        sort = request.POST.get('sort', 0)
        
        DeviceField.objects.create(
            name=name,
            field_key=field_key,
            field_type=field_type,
            is_required=is_required,
            is_visible=is_visible,
            is_card_visible=is_card_visible,
            options=options,
            default_value=default_value,
            sort=sort,
        )
        messages.success(request, '字段创建成功')
        return redirect('field_list')
    
    return render(request, 'assets/field_form.html')


@login_required
def field_edit(request, pk):
    field = get_object_or_404(DeviceField, pk=pk)
    
    if request.method == 'POST':
        field.name = request.POST.get('name')
        field.field_type = request.POST.get('field_type')
        field.is_required = request.POST.get('is_required') == 'on'
        field.is_visible = request.POST.get('is_visible') == 'on'
        field.is_card_visible = request.POST.get('is_card_visible') == 'on'
        field.options = request.POST.get('options')
        field.default_value = request.POST.get('default_value')
        field.sort = request.POST.get('sort', 0)
        
        if not field.is_system:
            field.field_key = request.POST.get('field_key')
        
        field.save()
        messages.success(request, '字段更新成功')
        return redirect('field_list')
    
    return render(request, 'assets/field_form.html', {'field': field})


@login_required
def field_delete(request, pk):
    if request.method == 'POST':
        field = get_object_or_404(DeviceField, pk=pk)
        if field.is_system:
            messages.error(request, '系统字段不能删除')
        else:
            field.delete()
            messages.success(request, '字段删除成功')
    return redirect('field_list')


@login_required
@permission_required('software')
def software_list(request):
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    license_type = request.GET.get('license_type', '')
    
    software_list = Software.objects.select_related('category').all()
    
    if search:
        software_list = software_list.filter(
            Q(asset_no__icontains=search) |
            Q(name__icontains=search) |
            Q(device_no__icontains=search) |
            Q(serial_no__icontains=search) |
            Q(vendor__icontains=search)
        )
    if category_id:
        software_list = software_list.filter(category_id=category_id)
    if license_type:
        software_list = software_list.filter(license_type=license_type)
    
    # 获取页码大小，默认20
    page_size = int(request.GET.get('page_size', 20))
    if page_size not in [20, 50, 100, 200]:
        page_size = 20
    
    paginator = Paginator(software_list, page_size)
    page = request.GET.get('page', 1)
    software_list = paginator.get_page(page)
    
    # 计算分页范围
    current_page = software_list.number
    total_pages = paginator.num_pages
    
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        page_range.extend([1, 2])
        start = max(3, current_page - 2)
        end = min(total_pages - 1, current_page + 2)
        if start > 3:
            page_range.append('...')
        page_range.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_range.append('...')
        page_range.extend([total_pages - 1, total_pages])
    
    categories = SoftwareCategory.objects.all()
    
    # 获取卡片可见字段
    card_visible_fields = SoftwareField.objects.filter(is_card_visible=True).order_by('sort')
    
    return render(request, 'assets/software_list.html', {
        'software_list': software_list,
        'categories': categories,
        'software_fields': SoftwareField.objects.filter(is_visible=True),
        'card_visible_fields': card_visible_fields,
        'page_range': page_range,
        'page_size': page_size,
    })


@login_required
def software_detail(request, pk):
    software = get_object_or_404(Software, pk=pk)
    return render(request, 'assets/software_detail.html', {'software': software})


@login_required
def parse_software_license_count(request):
    license_count_str = request.POST.get('license_count')
    if license_count_str == '' or license_count_str is None:
        return None
    try:
        val = int(license_count_str)
        if val == -1:
            return -1  # -1 表示无限制
        return val
    except:
        return None


@permission_required('software_create')
def software_create(request):
    if request.method == 'POST':
        software = Software.objects.create(
            asset_no=request.POST.get('asset_no') or None,
            device_no=request.POST.get('device_no') or None,
            serial_no=request.POST.get('serial_no') or None,
            name=request.POST.get('name'),
            category_id=request.POST.get('category') or None,
            version=request.POST.get('version') or None,
            vendor=request.POST.get('vendor') or None,
            license_type=request.POST.get('license_type', 'perpetual'),
            license_count=parse_software_license_count(request),
            purchase_date=request.POST.get('purchase_date') or None,
            expire_date=request.POST.get('expire_date') or None,
            price=request.POST.get('price') or None,
            description=request.POST.get('description') or None,
            is_fixed=request.POST.get('is_fixed') == 'on',
            asset_card_no=request.POST.get('asset_card_no') or None,
        )
        messages.success(request, '软件创建成功')
        return redirect('software_list')
    
    categories = SoftwareCategory.objects.all()
    return render(request, 'assets/software_form.html', {'categories': categories})


@login_required
def software_edit(request, pk):
    software = get_object_or_404(Software, pk=pk)
    
    if request.method == 'POST':
        software.asset_no = request.POST.get('asset_no') or None
        software.device_no = request.POST.get('device_no') or None
        software.serial_no = request.POST.get('serial_no') or None
        software.name = request.POST.get('name')
        software.category_id = request.POST.get('category') or None
        software.version = request.POST.get('version') or None
        software.vendor = request.POST.get('vendor') or None
        software.license_type = request.POST.get('license_type', 'perpetual')
        software.license_count = parse_software_license_count(request)
        software.purchase_date = request.POST.get('purchase_date') or None
        software.expire_date = request.POST.get('expire_date') or None
        software.price = request.POST.get('price') or None
        software.description = request.POST.get('description') or None
        software.is_fixed = request.POST.get('is_fixed') == 'on'
        software.asset_card_no = request.POST.get('asset_card_no') or None
        software.save()
        
        messages.success(request, '软件更新成功')
        return redirect('software_list')
    
    categories = SoftwareCategory.objects.all()
    return render(request, 'assets/software_form.html', {'software': software, 'categories': categories})


@login_required
@csrf_exempt
def software_delete(request, pk):
    try:
        if request.method == 'POST':
            software = get_object_or_404(Software, pk=pk)
            software.delete()
            return JsonResponse({'success': True, 'message': '删除成功'})
        return JsonResponse({'success': False, 'message': '请求方式错误'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'删除失败: {str(e)}'})


@login_required
@csrf_exempt
def software_batch_delete(request):
    try:
        if request.method == 'POST':
            ids = request.POST.get('ids', '').split(',')
            software_ids = [int(i) for i in ids if i]
            Software.objects.filter(id__in=software_ids).delete()
            return JsonResponse({'success': True, 'message': f'已删除 {len(software_ids)} 个软件'})
        return JsonResponse({'success': False, 'message': '请求方式错误'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'删除失败: {str(e)}'})


def process_software_import_task(task_id, file_content, update_existing, user_id):
    """后台处理软件导入任务"""
    import openpyxl
    from io import BytesIO
    
    progress = software_import_progress[task_id]
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
    except Exception as e:
        progress['status'] = 'error'
        progress['errors'].append(f'文件读取失败: {str(e)}')
        return
    
    headers = [cell.value for cell in ws[1]]
    field_map = {}
    for idx, header in enumerate(headers):
        if header:
            for field in SoftwareField.objects.all():
                if field.name == header:
                    field_map[idx] = field.field_key
                    break
    
    user = User.objects.filter(pk=user_id).first()
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    progress['total'] = len(rows)
    
    for idx, row in enumerate(rows, start=2):
        progress['current'] = idx - 1
        progress['current_asset_no'] = str(row[2]).strip() if row[2] else ''
        
        try:
            data = {}
            for col_idx, field_key in field_map.items():
                if col_idx >= len(row):
                    continue
                cell_value = row[col_idx]
                
                if field_key == 'license_count' and (cell_value == '-1' or cell_value == -1):
                    data[field_key] = -1  # -1 表示无限制
                elif field_key in ['purchase_date', 'expire_date'] and cell_value:
                    if isinstance(cell_value, datetime):
                        data[field_key] = cell_value.date()
                    elif isinstance(cell_value, str):
                        data[field_key] = cell_value
                    else:
                        data[field_key] = str(cell_value)
                elif field_key in ['is_fixed']:
                    data[field_key] = str(cell_value).lower() in ['true', '1', '是', 'yes']
                else:
                    data[field_key] = cell_value
            
            software_name = data.get('name')
            if not software_name:
                progress['error'] += 1
                progress['errors'].append(f'第{idx}行：软件名称不能为空')
                continue
            
            asset_no = data.get('asset_no')
            
            if asset_no:
                software = Software.objects.filter(asset_no=asset_no).first()
                
                if software:
                    if update_existing:
                        for key, value in data.items():
                            if value is not None and value != '':
                                setattr(software, key, value)
                        software.save()
                        progress['update'] += 1
                    else:
                        progress['skip'] += 1
                else:
                    Software.objects.create(**data)
                    progress['success'] += 1
            else:
                Software.objects.create(**data)
                progress['success'] += 1
                
        except Exception as e:
            progress['error'] += 1
            progress['errors'].append(f'第{idx}行：{str(e)}')
        
        time.sleep(0.01)
    
    progress['current'] = progress['total']
    progress['status'] = 'completed'


@login_required
def software_import(request):
    if request.method == 'POST':
        import openpyxl
        
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'message': '请选择要导入的文件'})
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': '请上传Excel文件(.xlsx或.xls)'})
        
        update_existing = request.POST.get('update_existing') == 'true'
        
        task_id = str(uuid.uuid4())[:8]
        file_content = file.read()
        
        software_import_progress[task_id] = {
            'status': 'processing',
            'total': 0,
            'current': 0,
            'success': 0,
            'update': 0,
            'skip': 0,
            'error': 0,
            'errors': [],
            'current_asset_no': ''
        }
        
        thread = threading.Thread(
            target=process_software_import_task,
            args=(task_id, file_content, update_existing, request.user.id)
        )
        thread.daemon = True
        thread.start()
        
        return JsonResponse({'success': True, 'task_id': task_id})
    
    return render(request, 'assets/software_import.html')


@login_required
def software_export(request):
    import openpyxl
    from django.utils import timezone as tz
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '软件台账'
    
    visible_fields = SoftwareField.objects.filter(is_visible=True)
    headers = [f.name for f in visible_fields]
    ws.append(headers)
    
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    software_list = Software.objects.all()
    
    for s in software_list:
        row = []
        for field in visible_fields:
            key = field.field_key
            if key == 'license_type':
                row.append(s.get_license_type_display())
            elif key == 'license_count':
                row.append('无限制' if s.license_count is None else s.license_count)
            elif key == 'is_fixed':
                row.append('是' if s.is_fixed else '否')
            elif key in ['purchase_date', 'expire_date']:
                row.append(getattr(s, key).strftime('%Y-%m-%d') if getattr(s, key) else '')
            elif key in ['created_at', 'updated_at']:
                row.append(getattr(s, key).strftime('%Y-%m-%d %H:%M') if getattr(s, key) else '')
            else:
                row.append(getattr(s, key) or '')
        ws.append(row)
    
    tz_value = get_config_value('timezone', 'Asia/Shanghai')
    now = datetime.now(tz.timezone(tz_value))
    filename = f'软件台账_{now.strftime("%Y%m%d_%H%M%S")}_{random.randint(1000,9999)}.xlsx'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def software_field_list(request):
    fields = SoftwareField.objects.all()
    return render(request, 'assets/software_field_list.html', {'fields': fields})


@login_required
def software_field_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        field_key = request.POST.get('field_key')
        field_type = request.POST.get('field_type')
        is_required = request.POST.get('is_required') == 'on'
        is_visible = request.POST.get('is_visible') == 'on'
        is_card_visible = request.POST.get('is_card_visible') == 'on'
        options = request.POST.get('options')
        default_value = request.POST.get('default_value')
        sort = request.POST.get('sort', 0)
        
        SoftwareField.objects.create(
            name=name,
            field_key=field_key,
            field_type=field_type,
            is_required=is_required,
            is_visible=is_visible,
            is_card_visible=is_card_visible,
            options=options,
            default_value=default_value,
            sort=sort,
        )
        messages.success(request, '字段创建成功')
        return redirect('software_field_list')
    
    return render(request, 'assets/software_field_form.html')


@login_required
def software_field_edit(request, pk):
    field = get_object_or_404(SoftwareField, pk=pk)
    
    if request.method == 'POST':
        field.name = request.POST.get('name')
        field.field_type = request.POST.get('field_type')
        field.is_required = request.POST.get('is_required') == 'on'
        field.is_visible = request.POST.get('is_visible') == 'on'
        field.is_card_visible = request.POST.get('is_card_visible') == 'on'
        field.options = request.POST.get('options')
        field.default_value = request.POST.get('default_value')
        field.sort = request.POST.get('sort', 0)
        
        if not field.is_system:
            field.field_key = request.POST.get('field_key')
        
        field.save()
        messages.success(request, '字段更新成功')
        return redirect('software_field_list')
    
    return render(request, 'assets/software_field_form.html', {'field': field})


@login_required
def software_field_delete(request, pk):
    if request.method == 'POST':
        field = get_object_or_404(SoftwareField, pk=pk)
        if field.is_system:
            messages.error(request, '系统字段不能删除')
        else:
            field.delete()
            messages.success(request, '字段删除成功')
    return redirect('software_field_list')


@login_required
@permission_required('consumable')
def consumable_list(request):
    consumables = Consumable.objects.all().order_by('created_at')
    return render(request, 'assets/consumable_list.html', {'consumables': consumables})


@login_required
@permission_required('consumable_create')
def consumable_create(request):
    if request.method == 'POST':
        Consumable.objects.create(
            name=request.POST.get('name'),
            category_id=request.POST.get('category') or None,
            specification=request.POST.get('specification'),
            applicable_models=request.POST.get('applicable_models'),
            stock_quantity=0,
            min_stock=request.POST.get('min_stock') or 0,
            description=request.POST.get('description'),
        )
        messages.success(request, '耗材创建成功')
        return redirect('consumable_list')
    
    categories = ConsumableCategory.objects.all()
    return render(request, 'assets/consumable_form.html', {'categories': categories})


@login_required
def consumable_receive_view(request):
    if request.method == 'POST':
        consumable_id = request.POST.get('consumable')
        quantity = int(request.POST.get('quantity', 0))
        purpose = request.POST.get('purpose', '')
        
        consumable = Consumable.objects.get(pk=consumable_id)
        consumable.stock_quantity += quantity
        consumable.save()
        
        ConsumableRecord.objects.create(
            consumable=consumable,
            quantity=quantity,
            record_type='receive',
            purpose=purpose,
            user=request.user,
            department=request.user.department if hasattr(request.user, 'department') else None,
        )
        messages.success(request, '入库成功')
        return redirect('consumable_list')
    
    consumables = Consumable.objects.all().order_by('name')
    consumables_json = [{'id': c.id, 'name': c.name, 'stock_quantity': c.stock_quantity} for c in consumables]
    selected_id = request.GET.get('consumable_id')
    return render(request, 'assets/consumable_receive.html', {
        'consumables': consumables, 
        'consumables_json': consumables_json,
        'selected_id': int(selected_id) if selected_id else None
    })


@login_required
def consumable_use_view(request):
    if request.method == 'POST':
        consumable_id = request.POST.get('consumable')
        quantity = int(request.POST.get('quantity', 1))
        user_id = request.POST.get('user')
        purpose = request.POST.get('purpose', '')
        
        consumable = Consumable.objects.get(pk=consumable_id)
        if consumable.stock_quantity < quantity:
            messages.error(request, '库存不足')
            return redirect('consumable_use')
        
        consumable.stock_quantity -= quantity
        consumable.save()
        
        selected_user = User.objects.get(pk=user_id)
        ConsumableRecord.objects.create(
            consumable=consumable,
            user=selected_user,
            department=selected_user.department if hasattr(selected_user, 'department') else None,
            quantity=quantity,
            record_type='领用',
            purpose=purpose,
            approved_by=request.user,
        )
        messages.success(request, '领用成功')
        return redirect('consumable_list')
    
    consumables = Consumable.objects.all().order_by('name')
    users = User.objects.filter(is_active=True).select_related('department').order_by('emp_no')
    selected_id = request.GET.get('consumable_id')
    return render(request, 'assets/consumable_use.html', {
        'consumables': consumables, 
        'users': users, 
        'current_user': request.user,
        'selected_id': int(selected_id) if selected_id else None
    })


def consumable_users_api(request):
    q = request.GET.get('q', '')
    users = User.objects.filter(is_active=True)
    if q:
        users = users.filter(Q(emp_no__icontains=q) | Q(realname__icontains=q) | Q(department__name__icontains=q))
    users = users.select_related('department')[:20]
    results = [{'id': u.id, 'text': f"{u.emp_no}-{u.realname}({u.department.name if u.department else ''})", 'display': f"{u.emp_no}-{u.realname}"} for u in users]
    return JsonResponse(results, safe=False)


@login_required
def consumable_detail(request, pk):
    consumable = get_object_or_404(Consumable, pk=pk)
    records = ConsumableRecord.objects.filter(consumable=consumable).select_related('user', 'department', 'approved_by').order_by('-created_at')
    return render(request, 'assets/consumable_detail.html', {
        'consumable': consumable,
        'records': records
    })


@login_required
def consumable_edit(request, pk):
    consumable = get_object_or_404(Consumable, pk=pk)
    if request.method == 'POST':
        consumable.name = request.POST.get('name')
        consumable.category_id = request.POST.get('category') or None
        consumable.specification = request.POST.get('specification')
        consumable.applicable_models = request.POST.get('applicable_models')
        consumable.min_stock = request.POST.get('min_stock') or 0
        consumable.description = request.POST.get('description')
        consumable.save()
        messages.success(request, '耗材更新成功')
        return redirect('consumable_detail', pk=consumable.pk)
    
    categories = ConsumableCategory.objects.all()
    return render(request, 'assets/consumable_form.html', {
        'consumable': consumable,
        'categories': categories
    })


@login_required
def consumable_delete(request, pk):
    consumable = get_object_or_404(Consumable, pk=pk)
    if request.method == 'POST':
        consumable.delete()
        messages.success(request, '耗材删除成功')
        return redirect('consumable_list')
    return render(request, 'assets/consumable_confirm_delete.html', {'consumable': consumable})


@login_required
def api_users_search(request):
    q = request.GET.get('q', '')
    department_id = request.GET.get('department_id', '')
    users = User.objects.filter(is_active=True)
    if department_id:
        users = users.filter(department_id=department_id)
    if q:
        users = users.filter(Q(emp_no__icontains=q) | Q(realname__icontains=q) | Q(department__name__icontains=q))
    users = users.select_related('department')[:20]
    results = [{'id': u.id, 'text': f"{u.emp_no}-{u.realname}（{u.department.name if u.department else ''}）", 'department_id': u.department_id or ''} for u in users]
    return JsonResponse(results, safe=False)


@login_required
def service_list(request):
    requests = ServiceRequest.objects.all()
    return render(request, 'assets/service_list.html', {'requests': requests})


@login_required
def service_create(request):
    if request.method == 'POST':
        last_request = ServiceRequest.objects.order_by('-id').first()
        next_no = f"SR{timezone.now().strftime('%Y%m%d')}{int(last_request.id or 0) + 1:04d}" if last_request else f"SR{timezone.now().strftime('%Y%m%d')}0001"
        
        ServiceRequest.objects.create(
            request_no=next_no,
            title=request.POST.get('title'),
            service_type_id=request.POST.get('service_type'),
            description=request.POST.get('description'),
            priority=request.POST.get('priority', 'normal'),
            requester=request.user,
            device_id=request.POST.get('device') or None,
        )
        messages.success(request, '服务请求创建成功')
        return redirect('service_list')
    
    service_types = ServiceType.objects.all()
    devices = Device.objects.all()
    return render(request, 'assets/service_form.html', {'service_types': service_types, 'devices': devices})


@login_required
def api_get_code(request):
    category_id = request.GET.get('category_id')
    if category_id:
        try:
            category = AssetCategory.objects.get(pk=category_id)
            code = generate_asset_no(category)
            return JsonResponse({'code': code})
        except AssetCategory.DoesNotExist:
            pass
    return JsonResponse({'code': ''})


@login_required
def api_category_list(request):
    categories = AssetCategory.objects.order_by('level', 'id').all()
    result = []
    for category in categories:
        result.append({
            'id': category.id,
            'name': category.name,
            'level': category.level,
            'parent_id': category.parent_id,
            'code': category.code,
            'description': category.description
        })
    return JsonResponse({'success': True, 'data': result})


@login_required
def api_category_save(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'}), 400
    
    data = json.loads(request.body)
    category_id = data.get('id')
    
    if category_id:
        category = get_object_or_404(AssetCategory, pk=category_id)
    else:
        category = AssetCategory()
    
    if not data.get('name'):
        return JsonResponse({'success': False, 'message': '分类名称不能为空'}), 400
    
    if not data.get('code'):
        return JsonResponse({'success': False, 'message': '分类编码不能为空'}), 400
    
    category.name = data.get('name')
    category.code = data.get('code')
    category.description = data.get('description', '')
    category.sort = data.get('sort', 0)
    
    parent_id = data.get('parent_id')
    if parent_id:
        parent = AssetCategory.objects.get(pk=parent_id)
        category.parent = parent
        category.level = parent.level + 1
        if category.level > 4:
            return JsonResponse({'success': False, 'message': '分类最多支持4级'}), 400
    else:
        category.parent = None
        category.level = 1
    
    try:
        if not category_id:
            category.save()
        else:
            category.save()
        return JsonResponse({'success': True, 'message': '保存成功', 'id': category.id})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'保存失败: {str(e)}'}), 500


@login_required
def api_category_delete(request, id):
    category = get_object_or_404(AssetCategory, pk=id)
    
    if AssetCategory.objects.filter(parent_id=id).first():
        return JsonResponse({'success': False, 'message': '该分类下有子分类，无法删除'}), 400
    
    if Device.objects.filter(category_id=id).first():
        return JsonResponse({'success': False, 'message': '该分类下有设备，无法删除'}), 400
    
    try:
        category.delete()
        return JsonResponse({'success': True, 'message': '删除成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'删除失败: {str(e)}'}), 500


@login_required
def api_generate_asset_number(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'}, status=400)
    
    data = json.loads(request.body)
    category_id = data.get('category_id')
    
    if not category_id:
        return JsonResponse({'success': False, 'message': '请选择分类'}, status=400)
    
    try:
        category = AssetCategory.objects.get(pk=category_id)
        
        # 使用 get_full_code() 获取完整的前缀（如 XACD-Z-001-001）
        asset_number_prefix = category.get_full_code()
        
        max_device = Device.objects.filter(
            asset_no__startswith=f"{asset_number_prefix}-"
        ).order_by('-asset_no').first()
        
        if max_device:
            last_part = max_device.asset_no.split('-')[-1]
            try:
                new_seq = int(last_part) + 1
                new_seq_str = f"{new_seq:03d}"
            except ValueError:
                new_seq_str = "001"
        else:
            new_seq_str = "001"
        
        new_asset_number = f"{asset_number_prefix}-{new_seq_str}"
        
        return JsonResponse({
            'success': True,
            'asset_number': new_asset_number
        })
    except AssetCategory.DoesNotExist:
        return JsonResponse({'success': False, 'message': '分类不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'生成资产编号失败: {str(e)}'}, status=500)


@csrf_exempt
@login_required
def api_get_category_by_asset_no(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '无效请求'}, status=400)
    
    data = json.loads(request.body)
    asset_no = data.get('asset_no', '').strip()
    
    if not asset_no:
        return JsonResponse({'success': False, 'message': '请输入资产编号'}, status=400)
    
    parts = asset_no.split('-')
    if len(parts) < 2:
        return JsonResponse({'success': False, 'message': '资产编号格式不正确'}, status=400)
    
    prefix = '-'.join(parts[:-1])
    
    category = AssetCategory.find_by_asset_prefix(prefix)
    
    if not category:
        return JsonResponse({'success': False, 'message': '未找到匹配的分类'}, status=404)
    
    return JsonResponse({
        'success': True,
        'category_id': category.id,
        'category_name': category.name,
        'category_level': category.level,
        'full_code': category.get_full_code()
    })


@login_required
def device_stats(request):
    stats = {
        'total': Device.objects.count(),
        'normal': Device.objects.filter(status='normal').count(),
        'fault': Device.objects.filter(status='fault').count(),
        'scrapped': Device.objects.filter(status='scrapped').count(),
        'unused': Device.objects.filter(status='unused').count(),
    }
    return JsonResponse(stats)


@login_required
def location_tree(request):
    def build_tree(location):
        children = location.children.all().order_by('sort', 'code')
        return {
            'id': location.id,
            'name': location.name,
            'code': location.code,
            'level': location.level,
            'parent_id': location.parent_id,
            'park_code': location.park_code,
            'building_code': location.building_code,
            'floor_code': location.floor_code,
            'floor_count': location.floor_count,
            'room_code': location.room_code,
            'has_map': location.has_map,
            'map_width': location.map_width,
            'map_height': location.map_height,
            'children': [build_tree(child) for child in children]
        }
    
    roots = AssetLocation.objects.filter(parent__isnull=True).order_by('sort', 'code')
    tree_data = [build_tree(loc) for loc in roots]
    return JsonResponse({'data': tree_data})


def map_data(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    
    elements = MapElement.objects.filter(location=location).values()
    workstations = Workstation.objects.filter(location=location).select_related('location')
    devices = Device.objects.filter(location=location, status__in=['normal', 'fault', 'unused']).select_related('category', 'user', 'workstation')
    
    try:
        background = location.background
        bg_data = {
            'bg_type': background.bg_type,
            'scale': background.scale,
            'offset_x': background.offset_x,
            'offset_y': background.offset_y,
            'opacity': background.opacity,
        }
        if background.file_data:
            bg_data['file_data'] = background.file_data.decode('utf-8') if isinstance(background.file_data, bytes) else None
    except MapBackground.DoesNotExist:
        bg_data = None
    
    workstation_list = []
    for ws in workstations:
        ws_devices = devices.filter(workstation=ws)
        workstation_list.append({
            'id': ws.id,
            'workstation_code': ws.workstation_code,
            'name': ws.name,
            'x': ws.x,
            'y': ws.y,
            'width': ws.width,
            'height': ws.height,
            'status': ws.status,
            'area_id': ws.area_id,
            'devices': [
                {
                    'id': d.id,
                    'asset_no': d.asset_no,
                    'name': d.name,
                    'status': d.status,
                    'user': d.user.realname if d.user else None,
                }
                for d in ws_devices
            ]
        })
    
    device_list = devices.filter(workstation__isnull=True).values(
        'id', 'asset_no', 'name', 'status', 'category__name', 'user__realname'
    )
    
    return JsonResponse({
        'location': {
            'id': location.id,
            'name': location.name,
            'code': location.code,
            'level': location.level,
            'map_width': location.map_width,
            'map_height': location.map_height,
        },
        'background': bg_data,
        'elements': [{'element_type': {'wall': 'line', 'window': 'dashed'}.get(e.get('element_type'), e.get('element_type')), **{k: v for k, v in e.items() if k != 'element_type'}} for e in elements],
        'workstations': workstation_list,
        'devices': list(device_list),
        'area_bindings': [{
            'id': b.id,
            'location_id': b.location_id,
            'location_name': b.location.name,
            'area_name': b.area_name,
            'area_points': b.area_points,
            'area_color': b.area_color,
        } for b in LocationAreaBinding.objects.filter(parent_location=location)],
    })


@csrf_exempt
@login_required
def map_element_save(request):
    from apps.assets.models import Workstation
    
    if request.method == 'POST':
        location_id = request.POST.get('location_id')
        
        elements_data = request.POST.get('elements')
        if elements_data:
            import json
            try:
                elements_list = json.loads(elements_data)
            except json.JSONDecodeError:
                return JsonResponse({'success': False, 'message': '数据格式错误'})
            
            try:
                saved_ids = []
                for el_data in elements_list:
                    el_id = el_data.get('id')
                    data = {
                        'location_id': location_id,
                        'element_type': {'line': 'wall', 'dashed': 'window'}.get(el_data.get('type', 'wall'), el_data.get('type', 'wall')),
                        'x': float(el_data.get('x', 0)),
                        'y': float(el_data.get('y', 0)),
                        'x2': float(el_data.get('x2')) if el_data.get('x2') is not None else None,
                        'y2': float(el_data.get('y2')) if el_data.get('y2') is not None else None,
                        'width': float(el_data.get('width', 0)),
                        'height': float(el_data.get('height', 0)),
                        'rotation': float(el_data.get('rotation', 0)),
                        'color': el_data.get('color', '#000000'),
                        'thickness': int(el_data.get('thickness', 2)),
                        'points': el_data.get('points', ''),
                        'properties': el_data.get('properties', ''),
                        'label': el_data.get('label', ''),
                        'sort_order': int(el_data.get('sort_order', 0)),
                        'door_direction': el_data.get('door_direction', 'right'),
                        'door_width': float(el_data.get('door_width', 60)),
                        'door_open_angle': int(el_data.get('door_open_angle', 90)),
                        'doorstop_width': float(el_data.get('doorstop_width', 15)),
                        'auto_doorstop': el_data.get('auto_doorstop', True),
                        'snap_enabled': el_data.get('snap_enabled', True),
                        'snap_threshold': float(el_data.get('snap_threshold', 10)),
                    }
                    
                    if el_id:
                        try:
                            el_id_int = int(el_id)
                            element = MapElement.objects.filter(pk=el_id_int).first()
                            if element:
                                for key, value in data.items():
                                    setattr(element, key, value)
                                element.save()
                                saved_ids.append(element.id)
                        except (ValueError, TypeError):
                            pass  # 忽略无效的 ID
                    else:
                        element = MapElement.objects.create(**data)
                        saved_ids.append(element.id)
                    
                    # Handle region type - create or update 4-level location
                    if data['element_type'] == 'region' and data['points']:
                        parent_location = AssetLocation.objects.filter(pk=location_id).first()
                        if parent_location and parent_location.level == 3:
                            region_name = data['label'] or '区域'
                            
                            # Get old label from el_data (set by editor when renaming)
                            old_label = el_data.get('old_label')
                            
                            # Try to find existing location by old label first, then new label
                            existing_location = None
                            if old_label:
                                existing_location = AssetLocation.objects.filter(parent=parent_location, level=4, name=old_label).first()
                            if not existing_location:
                                existing_location = AssetLocation.objects.filter(parent=parent_location, level=4, name=region_name).first()
                            
                            if existing_location:
                                # Update existing location with new name and area points
                                existing_location.name = region_name
                                existing_location.area_points = data['points']
                                existing_location.description = f"{parent_location.name} {region_name}"
                                existing_location.save()
                            else:
                                base_code = parent_location.code
                                num = 1
                                while AssetLocation.objects.filter(code=f"{base_code}-{num:03d}").exists():
                                    num += 1
                                new_code = f"{base_code}-{num:03d}"
                                # Create new 4-level location
                                new_location = AssetLocation.objects.create(
                                    name=region_name,
                                    code=new_code,
                                    parent=parent_location,
                                    level=4,
                                    park_code=parent_location.park_code,
                                    building_code=parent_location.building_code,
                                    floor_code=parent_location.floor_code,
                                    area_points=data['points'],
                                    description=f"{parent_location.name} {region_name}",
                                    sort=num
                                )
                            # 更新所有已存在的工位的区域关联
                            all_workstations = Workstation.objects.filter(location=parent_location)
                            for ws in all_workstations:
                                update_workstation_area(ws)
                    
                    # Sync workstation to Workstation table
                    if data['element_type'] == 'workstation':
                        location = AssetLocation.objects.filter(pk=location_id).first()
                        if location:
                            ws, created = Workstation.objects.get_or_create(
                                location=location,
                                workstation_code=data['label'] or f'WS-{element.id}',
                                defaults={
                                    'x': data['x'],
                                    'y': data['y'],
                                    'width': data['width'] or 60,
                                    'height': data['height'] or 40,
                                    'status': 'available',
                                }
                            )
                            if not created:
                                ws.x = data['x']
                                ws.y = data['y']
                                ws.width = data['width'] or 60
                                ws.height = data['height'] or 40
                                ws.save()
                            # 更新工位区域关联
                            update_workstation_area(ws)
                
                # 安全地构建 existing_ids
                existing_ids = []
                for el in elements_list:
                    el_id = el.get('id')
                    if el_id:
                        try:
                            existing_ids.append(int(el_id))
                        except (ValueError, TypeError):
                            pass  # 忽略无效的 ID
                
                # Delete workstations that correspond to deleted workstation-type elements
                deleted_ws_elements = MapElement.objects.filter(location_id=location_id, element_type='workstation').exclude(id__in=saved_ids)
                deleted_ws_labels = [e.label for e in deleted_ws_elements if e.label and e.label.strip()]
                if deleted_ws_labels:
                    try:
                        Workstation.objects.filter(location_id=location_id, workstation_code__in=deleted_ws_labels).delete()
                    except Exception:
                        pass  # 忽略删除失败的异常
                
                # Handle deleted region elements - clear area association and delete 4-level location
                deleted_region_elements = MapElement.objects.filter(location_id=location_id, element_type='region').exclude(id__in=saved_ids)
                for region_el in deleted_region_elements:
                    # Find and delete corresponding 4-level location
                    parent_location = AssetLocation.objects.filter(pk=location_id).first()
                    if parent_location:
                        region_location = AssetLocation.objects.filter(parent=parent_location, level=4, name=region_el.label).first()
                        if region_location:
                            # Clear area association for workstations
                            Workstation.objects.filter(area=region_location).update(area=None)
                            # Delete LocationAreaBinding records
                            LocationAreaBinding.objects.filter(location=region_location).delete()
                            # Delete the 4-level location
                            region_location.delete()
                
                MapElement.objects.filter(location_id=location_id).exclude(id__in=saved_ids).delete()
                
                return JsonResponse({'success': True, 'message': f'保存成功，共 {len(saved_ids)} 个元素', 'saved_ids': saved_ids})
            except Exception as e:
                return JsonResponse({'success': False, 'message': f'保存失败: {str(e)}'})
        
        element_id = request.POST.get('element_id')
        
        data = {
            'location_id': location_id,
            'element_type': request.POST.get('element_type'),
            'x': float(request.POST.get('x', 0)),
            'y': float(request.POST.get('y', 0)),
            'x2': float(request.POST.get('x2')) if request.POST.get('x2') else None,
            'y2': float(request.POST.get('y2')) if request.POST.get('y2') else None,
            'width': float(request.POST.get('width', 0)),
            'height': float(request.POST.get('height', 0)),
            'rotation': float(request.POST.get('rotation', 0)),
            'color': request.POST.get('color', '#000000'),
            'thickness': int(request.POST.get('thickness', 2)),
            'points': request.POST.get('points', ''),
            'properties': request.POST.get('properties', ''),
            'sort_order': int(request.POST.get('sort_order', 0)),
            'door_direction': request.POST.get('door_direction', 'right'),
            'door_width': float(request.POST.get('door_width', 60)),
            'door_open_angle': int(request.POST.get('door_open_angle', 90)),
            'doorstop_width': float(request.POST.get('doorstop_width', 15)),
            'auto_doorstop': request.POST.get('auto_doorstop', 'true').lower() == 'true',
            'snap_enabled': request.POST.get('snap_enabled', 'true').lower() == 'true',
            'snap_threshold': float(request.POST.get('snap_threshold', 10)),
        }
        
        if element_id:
            element = get_object_or_404(MapElement, pk=element_id)
            for key, value in data.items():
                setattr(element, key, value)
            element.save()
            message = '元素更新成功'
        else:
            element = MapElement.objects.create(**data)
            message = '元素创建成功'
        
        return JsonResponse({'success': True, 'message': message, 'element_id': element.id})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


@login_required
def map_element_delete(request, pk):
    if request.method == 'POST':
        element = get_object_or_404(MapElement, pk=pk)
        element.delete()
        return JsonResponse({'success': True, 'message': '元素删除成功'})
    return JsonResponse({'success': False, 'message': '无效请求'})


@login_required
def map_background_upload(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    
    if request.method == 'POST':
        bg_type = request.POST.get('bg_type', 'image')
        file_data = request.FILES.get('file')
        scale = float(request.POST.get('scale', 1.0))
        offset_x = float(request.POST.get('offset_x', 0))
        offset_y = float(request.POST.get('offset_y', 0))
        opacity = float(request.POST.get('opacity', 1.0))
        
        bg_data = {
            'location': location,
            'bg_type': bg_type,
            'scale': scale,
            'offset_x': offset_x,
            'offset_y': offset_y,
            'opacity': opacity,
        }
        
        if file_data:
            bg_data['file_data'] = file_data.read()
        
        MapBackground.objects.update_or_create(
            location=location,
            defaults=bg_data
        )
        
        AssetLocation.objects.filter(pk=pk).update(has_map=True)
        
        return JsonResponse({'success': True, 'message': '底图上传成功'})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


@login_required
def workstation_list(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    
    # 根据位置级别使用不同的查询条件
    if location.level == 4:
        # 4级区域：查询area字段
        workstations = Workstation.objects.filter(area=location).prefetch_related('devices__category', 'devices__user', 'devices__department')
    else:
        # 3级楼层：查询location字段
        workstations = Workstation.objects.filter(location=location).prefetch_related('devices__category', 'devices__user', 'devices__department')
    
    ws_data = []
    for ws in workstations:
        devices = ws.devices.select_related('user', 'department', 'category').order_by('category__name', 'id')
        
        display_name = ws.name
        if not display_name:
            # Find earliest PC device
            pc_device = devices.filter(category__name='台式机').first() or devices.filter(category__name__icontains='计算机').first()
            if pc_device and pc_device.user:
                display_name = f'{pc_device.user.realname}的工位'
            elif devices.exists():
                # Use earliest non-PC device's department
                first_device = devices.first()
                if first_device.department:
                    display_name = f'{first_device.department.name}工位'
        
        ws_data.append({
            'workstation': ws,
            'display_name': display_name,
            'full_path': location.get_full_path(),
            'devices': [{'id': d.id, 'asset_no': d.asset_no, 'name': d.name, 'status': d.status, 'status_display': d.get_status_display(), 'department': d.department.name if d.department else '', 'user': d.user.realname if d.user else ''} for d in devices],
            'device_count': devices.count(),
        })
    
    return render(request, 'assets/workstation_list.html', {
        'location': location,
        'ws_data': ws_data,
    })


@login_required
def api_workstations_by_location(request, location_id):
    location = get_object_or_404(AssetLocation, pk=location_id)
    workstations = Workstation.objects.filter(location=location).values('id', 'name', 'workstation_code', 'area_id')
    return JsonResponse({
        'success': True,
        'workstations': list(workstations)
    })


@login_required
def workstation_create(request):
    if request.method == 'POST':
        location_id = request.POST.get('location_id')
        location = get_object_or_404(AssetLocation, pk=location_id)
        
        code = generate_workstation_code(location)
        
        workstation = Workstation.objects.create(
            location=location,
            workstation_code=code,
            name=request.POST.get('name'),
            x=float(request.POST.get('x', 0)),
            y=float(request.POST.get('y', 0)),
            width=float(request.POST.get('width', 30)),
            height=float(request.POST.get('height', 20)),
            status=request.POST.get('status', 'available'),
            description=request.POST.get('description', ''),
        )
        
        messages.success(request, f'工位创建成功: {code}')
        return redirect('workstation_list', location_id=location_id)
    
    location_id = request.GET.get('location_id')
    location = get_object_or_404(AssetLocation, pk=location_id) if location_id else None
    locations = AssetLocation.objects.filter(level=4)
    return render(request, 'assets/workstation_form.html', {
        'location': location,
        'locations': locations,
    })


@login_required
def workstation_edit(request, pk):
    workstation = get_object_or_404(Workstation.objects.select_related('location'), pk=pk)
    
    if not workstation.name:
        devices = workstation.devices.select_related('user', 'department', 'category').order_by('id')
        pc_device = devices.filter(category__name='台式机').first() or devices.filter(category__name__icontains='计算机').first()
        if pc_device and pc_device.user:
            workstation.name = f'{pc_device.user.realname}的工位'
        elif devices.exists():
            first_device = devices.first()
            if first_device.department:
                workstation.name = f'{first_device.department.name}工位'
    
    if request.method == 'POST':
        workstation.name = request.POST.get('name', '')
        workstation.status = request.POST.get('status', 'available')
        workstation.description = request.POST.get('description', '')
        workstation.save()
        
        messages.success(request, '工位更新成功')
        return redirect('workstation_list', pk=workstation.location_id)
    
    return render(request, 'assets/workstation_form.html', {
        'workstation': workstation,
    })


@login_required
def workstation_delete(request, pk):
    if request.method == 'POST':
        workstation = get_object_or_404(Workstation, pk=pk)
        # Check if workstation has devices bound
        if workstation.devices.exists():
            return JsonResponse({'success': False, 'message': '该工位有绑定设备，无法删除'})
        workstation.delete()
        return JsonResponse({'success': True, 'message': '工位删除成功'})
    return JsonResponse({'success': False, 'message': '无效请求'})


@login_required
def workstation_batch_create(request):
    if request.method == 'POST':
        location_id = request.POST.get('location_id')
        location = get_object_or_404(AssetLocation, pk=location_id)
        
        rows = int(request.POST.get('rows', 1))
        cols = int(request.POST.get('cols', 1))
        start_x = float(request.POST.get('start_x', 0))
        start_y = float(request.POST.get('start_y', 0))
        spacing_x = float(request.POST.get('spacing_x', 50))
        spacing_y = float(request.POST.get('spacing_y', 40))
        width = float(request.POST.get('width', 30))
        height = float(request.POST.get('height', 20))
        
        prefix = location.code or f"L{location.id}"
        
        created = []
        for row in range(rows):
            for col in range(cols):
                code = generate_workstation_code(location, row * cols + col + 1)
                ws = Workstation.objects.create(
                    location=location,
                    workstation_code=code,
                    name=f"{prefix}-{row+1:02d}{col+1:02d}",
                    x=start_x + col * spacing_x,
                    y=start_y + row * spacing_y,
                    width=width,
                    height=height,
                    status='available',
                )
                created.append(ws.workstation_code)
        
        messages.success(request, f'成功创建 {len(created)} 个工位')
        return redirect('workstation_list', location_id=location_id)
    
    location_id = request.GET.get('location_id')
    location = get_object_or_404(AssetLocation, pk=location_id) if location_id else None
    locations = AssetLocation.objects.filter(level=4)
    return render(request, 'assets/workstation_batch.html', {'location': location, 'locations': locations})


def generate_workstation_code(location, seq=None):
    prefix = location.code or f"L{location.id}"
    
    if seq is None:
        last_ws = Workstation.objects.filter(location=location).order_by('-id').first()
        seq = (last_ws.id or 0) + 1 if last_ws else 1
    
    return f"{prefix}-{seq:04d}"


@login_required
def api_device_bind_workstation(request):
    if request.method == 'POST':
        device_id = request.POST.get('device_id')
        workstation_id = request.POST.get('workstation_id')
        
        device = get_object_or_404(Device, pk=device_id)
        
        if workstation_id:
            workstation = get_object_or_404(Workstation, pk=workstation_id)
            device.workstation = workstation
            # 根据工位区域自动更新设备位置
            if workstation.area_id:
                device.location = workstation.area
                parts = []
                loc = workstation.area
            else:
                device.location = workstation.location
                parts = []
                loc = workstation.location
            while loc:
                parts.insert(0, loc.name)
                loc = loc.parent
            parts.append(workstation.workstation_code)
            device.location_text = '-'.join(parts)
            workstation.status = 'occupied'
            workstation.save()
        else:
            device.workstation = None
        
        device.save()
        
        AssetLog.objects.create(
            device=device,
            user=request.user,
            action='update',
            field_name='workstation',
            old_value='',
            new_value=str(device.workstation) if device.workstation else '未绑定',
        )
        
        return JsonResponse({'success': True, 'message': '设备绑定成功'})
    
    return JsonResponse({'success': False, 'message': '无效请求'})


@login_required
def location_map(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    elements = location.map_elements.all()
    workstations = location.workstations.all()
    
    try:
        background = location.background
    except MapBackground.DoesNotExist:
        background = None
    
    all_locations = AssetLocation.objects.filter(level__lte=4).order_by('level', 'sort', 'id')
    
    def build_tree(loc):
        children = loc.children.all()
        return {
            'id': loc.id,
            'name': loc.name,
            'code': loc.code,
            'level': loc.level,
            'park_code': loc.park_code,
            'building_code': loc.building_code,
            'floor_code': loc.floor_code,
            'floor_count': loc.floor_count,
            'has_map': loc.has_map,
            'children': [build_tree(c) for c in children]
        }
    
    roots = AssetLocation.objects.filter(parent__isnull=True)
    location_tree = [build_tree(loc) for loc in roots]
    location_tree_json = json.dumps(location_tree)
    
    return render(request, 'assets/location_map.html', {
        'location': location,
        'elements': elements,
        'workstations': workstations,
        'background': background,
        'location_tree_json': location_tree_json,
        'all_locations': all_locations,
    })


@login_required
def location_map_edit(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    
    elements_qs = location.map_elements.all()
    elements_list = []
    for el in elements_qs:
        elements_list.append({
            'i': el.id,
            't': 'line' if el.element_type == 'wall' else ('dashed' if el.element_type == 'window' else el.element_type),
            'x': el.x,
            'y': el.y,
            'x2': el.x2,
            'y2': el.y2,
            'w': el.width,
            'h': el.height,
            'c': el.color or '#000',
            'th': el.thickness,
            'dd': el.door_direction or 'right',
            'dw': el.door_width or 60,
            'da': el.door_open_angle or 90,
            'pt': el.points or '',
            'label': el.label or '',
            'angle': el.rotation or 0,
        })
    elements_json = json.dumps(elements_list)
    
    return render(request, 'assets/map_editor.html', {
        'location': location,
        'elements_json': elements_json,
    })


@login_required
@require_http_methods(["GET", "POST"])
def location_area_binding(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    
    if request.method == 'GET':
        child_locations = AssetLocation.objects.filter(parent=location, level=4)
        bindings = LocationAreaBinding.objects.filter(parent_location=location)
        binding_map = {b.location_id: b for b in bindings}
        
        return render(request, 'assets/location_area_binding.html', {
            'location': location,
            'child_locations': child_locations,
            'bindings': bindings,
            'binding_map': binding_map,
        })
    
    location_id = request.POST.get('location_id')
    area_points = request.POST.get('area_points', '[]')
    area_color = request.POST.get('area_color', '#3498db')
    area_name = request.POST.get('area_name', '')
    
    child_location = AssetLocation.objects.get(pk=location_id)
    
    binding, created = LocationAreaBinding.objects.update_or_create(
        location=child_location,
        parent_location=location,
        defaults={
            'area_points': area_points,
            'area_color': area_color,
            'area_name': area_name or child_location.name,
        }
    )
    
    messages.success(request, f'区域绑定成功: {binding.area_name}')
    return redirect('location_area_binding', pk=pk)


@login_required
def location_area_binding_delete(request, pk, binding_id):
    binding = get_object_or_404(LocationAreaBinding, pk=binding_id, parent_location_id=pk)
    binding.delete()
    messages.success(request, '区域绑定已删除')
    return redirect('location_area_binding', pk=pk)


@login_required
def api_location_area_bindings(request, pk):
    location = get_object_or_404(AssetLocation, pk=pk)
    bindings = LocationAreaBinding.objects.filter(parent_location=location)
    
    data = [{
        'id': b.id,
        'location_id': b.location_id,
        'location_name': b.location.name,
        'area_name': b.area_name,
        'area_points': b.area_points,
        'area_color': b.area_color,
    } for b in bindings]
    
    return JsonResponse({'success': True, 'bindings': data})


@login_required
def api_devices_unbound(request):
    devices = Device.objects.filter(
        workstation__isnull=True, status='normal'
    ).select_related('category', 'user')[:50]
    data = [{
        'id': d.id,
        'asset_no': d.asset_no,
        'name': d.name,
        'status': d.status,
        'category': d.category.name if d.category else '',
        'user': d.user.realname if d.user else '',
    } for d in devices]
    return JsonResponse({'success': True, 'devices': data})


@login_required
def api_devices_search(request):
    q = request.GET.get('q', '').strip()
    ids_param = request.GET.get('ids', '').strip()
    
    if ids_param:
        ids = [int(i) for i in ids_param.split(',') if i]
        devices = Device.objects.filter(id__in=ids).select_related('category', 'user', 'workstation')
    elif q:
        devices = Device.objects.filter(
            Q(asset_no__icontains=q) | Q(name__icontains=q) | Q(serial_no__icontains=q)
        ).select_related('category', 'user', 'workstation')[:20]
    else:
        return JsonResponse({'success': True, 'devices': []})
    
    data = [{
        'id': d.id,
        'asset_no': d.asset_no,
        'name': d.name,
        'status': d.status,
        'status_display': d.get_status_display(),
        'category': d.category.name if d.category else '',
        'user': d.user.realname if d.user else '',
        'workstation_id': d.workstation_id,
        'asset_card_no': d.asset_card_no or '',
    } for d in devices]
    return JsonResponse({'success': True, 'devices': data})


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def api_workstation_bind_device(request, pk):
    try:
        workstation = get_object_or_404(Workstation, pk=pk)
        body = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        device_id = body.get('device_id')
        device = get_object_or_404(Device, pk=device_id)
        device.workstation = workstation
        # 根据工位区域自动更新设备位置
        if workstation.area_id:
            device.location = workstation.area
            parts = []
            loc = workstation.area
        else:
            device.location = workstation.location
            parts = []
            loc = workstation.location
        while loc:
            parts.insert(0, loc.name)
            loc = loc.parent
        parts.append(workstation.workstation_code)
        device.location_text = '-'.join(parts)
        device.save()
        workstation.status = 'occupied'
        workstation.save()
        AssetLog.objects.create(
            device=device, user=request.user, action='assign',
            field_name='workstation', old_value='', new_value=str(workstation),
        )
        return JsonResponse({'success': True, 'message': '绑定成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
@login_required
@require_http_methods(["POST"])
def api_workstation_unbind_device(request, pk):
    try:
        workstation = get_object_or_404(Workstation, pk=pk)
        body = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        device_id = body.get('device_id')
        device = get_object_or_404(Device, pk=device_id)
        old_ws = str(device.workstation) if device.workstation else ''
        device.workstation = None
        device.save()
        if not workstation.devices.exists():
            workstation.status = 'available'
            workstation.save()
        AssetLog.objects.create(
            device=device, user=request.user, action='transfer',
            field_name='workstation', old_value=old_ws, new_value='未绑定',
        )
        return JsonResponse({'success': True, 'message': '解绑成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def api_users_by_department(request):
    department_id = request.GET.get('department_id')
    if department_id:
        users = User.objects.filter(department_id=department_id, is_active=True).select_related('department')
    else:
        users = User.objects.filter(is_active=True).select_related('department')
    result = [{
        'id': u.id,
        'emp_no': u.emp_no,
        'realname': u.realname,
        'department': u.department.name if u.department else ''
    } for u in users]
    return JsonResponse({'success': True, 'users': result})


@login_required
@csrf_exempt
def api_device_assign(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST请求'})
    try:
        device = get_object_or_404(Device, pk=pk)
        body = json.loads(request.body) if request.content_type == 'application/json' else request.POST
        user_id = body.get('user_id')
        user = get_object_or_404(User, pk=user_id)
        old_user = str(device.user) if device.user else ''
        old_dept = str(device.department) if device.department else ''
        old_status = device.get_status_display()
        device.user = user
        device.department = user.department
        device.status = 'normal'
        device.save()
        AssetLog.objects.create(device=device, user=request.user, action='assign',
            old_value=f'{old_user}/{old_dept}/{old_status}',
            new_value=f'{user.realname}/{user.department.name if user.department else ""}/使用')
        return JsonResponse({'success': True, 'message': '分配成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
def api_device_revoke(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST请求'})
    try:
        device = get_object_or_404(Device, pk=pk)
        old_user = str(device.user) if device.user else ''
        old_dept = str(device.department) if device.department else ''
        old_status = device.get_status_display()
        device.user = None
        device.department = None
        device.status = 'unused'
        device.save()
        AssetLog.objects.create(device=device, user=request.user, action='revoke',
            old_value=f'{old_user}/{old_dept}/{old_status}',
            new_value='未分配/无/闲置')
        return JsonResponse({'success': True, 'message': '回收成功'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
def api_regenerate_all_qrcodes(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '仅支持POST请求'})
    
    from django.http import StreamingHttpResponse
    
    def generate_progress():
        devices = Device.objects.all()
        total = devices.count()
        processed = 0
        errors = []
        
        yield f'data: {json.dumps({"status": "start", "total": total, "processed": 0})}\n\n'
        
        app_url = get_config_value('app_url', 'http://127.0.0.1:8000').rstrip('/')
        for device in devices:
            try:
                qr = qrcode.QRCode(version=1, box_size=10, border=1)
                qr.add_data(f'{app_url}/assets/view/{device.asset_no}')
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                buffer = BytesIO()
                img.save(buffer, 'PNG')
                asset_no = device.asset_no or f'DEV-{device.id}'
                if device.qrcode:
                    device.qrcode.delete(save=False)
                device.qrcode.save(f'QR-{asset_no}.png', buffer)
                processed += 1
                
                yield f'data: {json.dumps({"status": "processing", "total": total, "processed": processed, "current": asset_no})}\n\n'
            except Exception as e:
                errors.append({'device_id': device.id, 'asset_no': device.asset_no, 'error': str(e)})
                processed += 1
                yield f'data: {json.dumps({"status": "processing", "total": total, "processed": processed, "current": device.asset_no, "error": str(e)})}\n\n'
        
        yield f'data: {json.dumps({"status": "complete", "total": total, "processed": processed, "errors": errors})}\n\n'
    
    response = StreamingHttpResponse(generate_progress(), content_type='text/event-stream')
    response['Cache-Control'] = 'no-cache'
    return response


# 标签管理相关视图
def get_device_all_fields():
    """获取设备所有可用字段（系统字段+自定义字段）"""
    system_fields = [
        {'field_key': 'asset_no', 'label': '资产编号', 'type': 'text'},
        {'field_key': 'device_no', 'label': '设备编号', 'type': 'text'},
        {'field_key': 'serial_no', 'label': '序列号', 'type': 'text'},
        {'field_key': 'name', 'label': '设备名称', 'type': 'text'},
        {'field_key': 'model', 'label': '型号', 'type': 'text'},
        {'field_key': 'category', 'label': '资产分类', 'type': 'foreignkey'},
        {'field_key': 'status', 'label': '设备状态', 'type': 'choice'},
        {'field_key': 'secret_level', 'label': '密级', 'type': 'choice'},
        {'field_key': 'user', 'label': '使用人', 'type': 'foreignkey'},
        {'field_key': 'department', 'label': '所属部门', 'type': 'foreignkey'},
        {'field_key': 'location', 'label': '位置', 'type': 'foreignkey'},
        {'field_key': 'workstation', 'label': '工位', 'type': 'foreignkey'},
        {'field_key': 'location_text', 'label': '所在位置描述', 'type': 'text'},
        {'field_key': 'purchase_date', 'label': '购入日期', 'type': 'date'},
        {'field_key': 'enable_date', 'label': '启用时间', 'type': 'date'},
        {'field_key': 'install_date', 'label': '安装时间', 'type': 'date'},
        {'field_key': 'mac_address', 'label': 'MAC地址', 'type': 'text'},
        {'field_key': 'ip_address', 'label': 'IP地址', 'type': 'text'},
        {'field_key': 'os_name', 'label': '操作系统', 'type': 'text'},
        {'field_key': 'os_version', 'label': '系统版本', 'type': 'text'},
        {'field_key': 'disk_serial', 'label': '硬盘序列号', 'type': 'text'},
        {'field_key': 'purpose', 'label': '用途', 'type': 'text'},
        {'field_key': 'remarks', 'label': '备注', 'type': 'textarea'},
        {'field_key': 'is_fixed', 'label': '固资在账', 'type': 'boolean'},
        {'field_key': 'asset_card_no', 'label': '卡片编号', 'type': 'text'},
        {'field_key': 'is_secret', 'label': '保密台账', 'type': 'boolean'},
        {'field_key': 'secret_category', 'label': '台账分类', 'type': 'text'},
        {'field_key': 'qrcode', 'label': '二维码', 'type': 'image'},
        {'field_key': 'photo', 'label': '设备照片', 'type': 'image'},
    ]
    
    custom_fields = list(DeviceField.objects.all().values('field_key', 'name', 'field_type'))
    custom_fields = [{'field_key': f['field_key'], 'label': f['name'], 'type': f['field_type']} for f in custom_fields]
    
    return system_fields + custom_fields


def get_device_field_value(device, field_key):
    """获取设备字段的值"""
    if field_key == 'location':
        obj = getattr(device, 'location', None)
        if obj:
            return obj.get_full_path() if hasattr(obj, 'get_full_path') else str(obj)
        return ''
    elif field_key == 'workstation':
        obj = getattr(device, 'workstation', None)
        if obj:
            return obj.workstation_code if hasattr(obj, 'workstation_code') else str(obj)
        return ''
    elif field_key in ['category', 'user', 'department', 'created_by']:
        obj = getattr(device, field_key, None)
        if obj:
            if hasattr(obj, 'name'):
                return obj.name
            elif hasattr(obj, 'realname'):
                return obj.realname
            return str(obj)
        return ''
    elif field_key in ['purchase_date', 'enable_date', 'install_date', 'created_at', 'updated_at']:
        value = getattr(device, field_key, None)
        return value.strftime('%Y-%m-%d') if value else ''
    elif field_key in ['is_fixed', 'is_secret']:
        return '是' if getattr(device, field_key, False) else '否'
    elif field_key == 'status':
        return device.get_status_display()
    elif field_key == 'secret_level':
        return device.get_secret_level_display()
    else:
        return str(getattr(device, field_key, '') or '')


def get_device_data_dict(device):
    """获取设备所有字段的值字典"""
    if not device:
        return {}
    
    data = {}
    fields = get_device_all_fields()
    for field in fields:
        field_key = field['field_key']
        data[field_key] = get_device_field_value(device, field_key)
    
    return data


@login_required
def label_settings(request):
    """标签管理页面 - 直接编辑默认模板"""
    # 获取或创建默认模板
    template = LabelTemplate.objects.filter(is_default=True).first()
    if not template:
        template = LabelTemplate.objects.first()
    
    # 如果没有模板，自动创建一个默认模板
    if not template:
        template = LabelTemplate.objects.create(
            name='默认标签模板',
            size_type='40x60',
            width=60,  # 宽度
            height=40,  # 高度
            fields_config=[],  # 将由前端初始化
            layout_config={'border': True, 'border_width': 1},
            is_default=True,
        )
    
    # 如果是POST请求，保存模板
    if request.method == 'POST':
        template.size_type = request.POST.get('size_type', '40x60')
        
        # 根据尺寸类型自动计算宽度和高度
        size_type = template.size_type
        if size_type == '40x60':
            template.width = 60  # 宽度60mm
            template.height = 40  # 高度40mm
        elif size_type == '50x80':
            template.width = 80  # 宽度80mm
            template.height = 50  # 高度50mm
        else:
            template.width = int(request.POST.get('width', 60))
            template.height = int(request.POST.get('height', 40))
        
        # 安全解析 JSON，避免空字符串导致错误
        fields_config_str = request.POST.get('fields_config', '[]')
        layout_config_str = request.POST.get('layout_config', '{}')
        
        try:
            template.fields_config = json.loads(fields_config_str) if fields_config_str else []
        except json.JSONDecodeError:
            template.fields_config = []
        
        try:
            template.layout_config = json.loads(layout_config_str) if layout_config_str else {}
        except json.JSONDecodeError:
            template.layout_config = {}
        
        template.save()
        messages.success(request, '标签模板保存成功')
        return redirect('label_settings')
    
    # 获取随机设备用于预览
    device = Device.objects.order_by('?').first()
    
    all_fields = get_device_all_fields()
    
    # 将 JSON 字段转换为字符串，确保前端正确解析
    fields_config_json = json.dumps(template.fields_config) if template.fields_config else '[]'
    layout_config_json = json.dumps(template.layout_config) if template.layout_config else '{}'
    
    return render(request, 'assets/label_template_form.html', {
        'template': template,
        'all_fields': json.dumps(all_fields),
        'fields_config_json': fields_config_json,
        'layout_config_json': layout_config_json,
        'device': device,
        'device_data': json.dumps(get_device_data_dict(device)) if device else '{}',
    })


@login_required
def api_get_label_templates(request):
    templates = LabelTemplate.objects.all().values('id', 'name', 'size_type', 'width', 'height', 'is_default')
    return JsonResponse({'success': True, 'templates': list(templates)})


@login_required
def device_print_label(request, pk):
    device = get_object_or_404(Device, pk=pk)
    template = LabelTemplate.objects.filter(is_default=True).first()
    
    if not template:
        template = LabelTemplate.objects.first()
    
    # 如果没有模板，自动创建默认模板
    if not template:
        template = LabelTemplate.objects.create(
            name='默认标签模板',
            size_type='40x60',
            width=60,
            height=40,
            fields_config=[],
            layout_config={'border': True, 'border_width': 1},
            is_default=True,
        )
    
    buffer = generate_label_pdf_buffer([device], template)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="label_{device.asset_no}.pdf"'
    return response


@login_required
def device_batch_print(request):
    ids_str = request.GET.get('ids', '')
    if not ids_str:
        messages.error(request, '请选择要打印的设备')
        return redirect('device_list')
    
    ids = [int(id) for id in ids_str.split(',') if id.isdigit()]
    devices = Device.objects.filter(id__in=ids)
    
    if not devices.exists():
        messages.error(request, '未找到选中的设备')
        return redirect('device_list')
    
    template_id = request.GET.get('template_id')
    if template_id:
        template = get_object_or_404(LabelTemplate, pk=template_id)
    else:
        template = LabelTemplate.objects.filter(is_default=True).first()
        if not template:
            template = LabelTemplate.objects.first()
    
    # 如果没有模板，自动创建默认模板
    if not template:
        template = LabelTemplate.objects.create(
            name='默认标签模板',
            size_type='40x60',
            width=60,
            height=40,
            fields_config=[],
            layout_config={'border': True, 'border_width': 1},
            is_default=True,
        )
    
    if request.method == 'POST':
        template_id = request.POST.get('template_id')
        if template_id:
            template = get_object_or_404(LabelTemplate, pk=template_id)
    
    buffer = generate_label_pdf_buffer(devices, template)
    
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="labels_batch_{len(devices)}.pdf"'
    return response


def generate_label_pdf_buffer(devices, template):
    """生成标签PDF并返回Buffer - 与前端预览保持一致"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from django.contrib.staticfiles import finders
    
    # 使用微软雅黑字体（与前端 Microsoft YaHei 一致）
    font_name = 'Helvetica'  # 后备字体
    font_bold_name = 'Helvetica-Bold'  # 后备粗体
    
    font_path = finders.find('fonts/msyh.ttc')
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont('MsYaHei', font_path))
            font_name = 'MsYaHei'
        except:
            pass
    
    font_bold_path = finders.find('fonts/msyhbd.ttc')
    if font_bold_path:
        try:
            pdfmetrics.registerFont(TTFont('MsYaHei-Bold', font_bold_path))
            font_bold_name = 'MsYaHei-Bold'
        except:
            pass
    
    # 转换为列表，确保可以多次遍历
    devices_list = list(devices)
    
    buffer = BytesIO()
    
    # 前端使用的缩放因子
    scale = 5
    # 像素到PDF点的转换 (96DPI -> 72DPI)
    px_to_pt = 72 / 96
    
    # 页面尺寸：与前端canvas一致，然后转换为PDF点
    # 前端：canvas = width * 5 × height * 5 (像素)
    page_width = template.width * scale * px_to_pt
    page_height = template.height * scale * px_to_pt
    
    # 始终使用纵向页面（标签是竖向放置的）
    page_size = (page_width, page_height)
    
    c = canvas.Canvas(buffer, pagesize=page_size)
    
    for i, device in enumerate(devices_list):
        # 每个标签一页
        if i > 0:
            c.showPage()
        
        # 绘制边框
        border = template.layout_config.get('border', True)
        if border:
            border_width = template.layout_config.get('border_width', 1)
            c.setLineWidth(border_width)
            c.rect(0, 0, page_width, page_height)
        
        # 绘制字段
        for field_config in template.fields_config:
            if not field_config.get('show', True):
                continue
            
            field_key = field_config.get('field_key', '')
            # 坐标：与前端一致 (x * 5 * px_to_pt, y * 5 * px_to_pt)
            field_x = field_config.get('x', 5) * scale * px_to_pt
            field_y = field_config.get('y', 10) * scale * px_to_pt
            
            if field_key == 'qrcode':
                if device.qrcode:
                    # 二维码大小：与前端一致 (size * 5 / 2 * px_to_pt)
                    qr_size = field_config.get('size', 20) * scale / 2 * px_to_pt
                    try:
                        c.drawImage(device.qrcode.path, field_x, field_y, qr_size, qr_size)
                    except:
                        pass
            else:
                value = get_device_field_value(device, field_key)
                if value:
                    # 字体大小：与前端一致 (font_size * scale / 2.5 * px_to_pt)
                    font_size = field_config.get('font_size', 9) * scale / 2.5 * px_to_pt
                    bold = field_config.get('bold', False)
                    
                    if bold:
                        c.setFont(font_bold_name, font_size)
                    else:
                        c.setFont(font_name, font_size)
                    
                    label_prefix = field_config.get('label', '')
                    show_label = field_config.get('show_label', True)
                    
                    if show_label and label_prefix:
                        text = f"{label_prefix}: {value}"
                    else:
                        text = value
                    
                    c.drawString(field_x, field_y, text)
                    
                    c.drawString(field_x, field_y, text)
    
    c.save()
    buffer.seek(0)
    return buffer


# 设备导入相关视图
@login_required
def device_download_template(request):
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '设备导入模板'
    
    headers = ['资产编号', '设备编号', '设备名称', '型号', '序列号', '密级', 'MAC地址', 'IP地址', '操作系统', '安装时间', '硬盘序列号', '购入日期', '启用时间', '用途', '备注', '固资在账', '卡片编号', '保密台账', '台账分类']
    ws.append(headers)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 10)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="device_import_template.xlsx"'
    wb.save(response)
    return response


def software_download_template(request):
    import openpyxl
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '软件导入模板'
    
    headers = ['资产编号', '编号', '软件名称', '版本', '供应商', '授权类型', '授权数量', '价格', '购买日期', '到期日期', '固资在账', '卡片编号', '描述']
    ws.append(headers)
    
    ws.append(['SW-001', '编号001', 'Microsoft Office', '2024', '微软', '永久授权', '-1', '5000', '2024-01-01', '2025-12-31', '是', 'KA-001', '办公软件'])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 10)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="software_import_template.xlsx"'
    wb.save(response)
    return response


@login_required
def device_export(request):
    import openpyxl
    from django.utils import timezone as tz
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '设备台账'
    
    headers = [
        '资产分类', '资产编号', '设备编号', '设备名称', '型号', '序列号', '密级',
        '所属部门', '使用人', '位置', '工位', '设备状态', 'MAC地址', 'IP地址',
        '操作系统', '安装时间', '硬盘序列号', '购入日期', '启用时间', '用途',
        '备注', '固资在账', '卡片编号', '保密台账', '台账分类'
    ]
    ws.append(headers)
    
    # 设置表头样式
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 获取设备数据
    devices = Device.objects.select_related(
        'category', 'location', 'user', 'department', 'workstation'
    ).exclude(status='scrapped').order_by('id')
    
    # 获取搜索条件
    search = request.GET.get('search', '')
    category_id = request.GET.get('category', '')
    location_id = request.GET.get('location', '')
    status = request.GET.get('status', '')
    
    if search:
        devices = devices.filter(
            Q(asset_no__icontains=search) | 
            Q(name__icontains=search) | 
            Q(serial_no__icontains=search) |
            Q(device_no__icontains=search) |
            Q(model__icontains=search) |
            Q(user__realname__icontains=search) |
            Q(department__name__icontains=search) |
            Q(mac_address__icontains=search) |
            Q(ip_address__icontains=search) |
            Q(remarks__icontains=search)
        )
    if category_id:
        devices = devices.filter(category_id=category_id)
    if location_id:
        devices = devices.filter(location_id=location_id)
    if status:
        devices = devices.filter(status=status)
    
    # 密级映射
    secret_level_map = {
        'public': '公开',
        'internal': '内部',
        'confidential': '秘密',
        'secret': '机密',
        'top_secret': '绝密',
        'commercial_secret': '商密',
    }
    
    # 状态映射
    status_map = {
        'normal': '使用',
        'fault': '故障',
        'scrapped': '报废',
        'unused': '闲置',
    }
    
    # 写入数据
    for device in devices:
        row = [
            device.category.name if device.category else '',
            device.asset_no or '',
            device.device_no or '',
            device.name or '',
            device.model or '',
            device.serial_no or '',
            secret_level_map.get(device.secret_level, device.secret_level or ''),
            device.department.name if device.department else '',
            device.user.realname if device.user else '',
            device.location.get_full_path() if device.location else '',
            device.workstation.workstation_code if device.workstation else '',
            status_map.get(device.status, device.status or ''),
            device.mac_address or '',
            device.ip_address or '',
            device.os_name or '',
            device.install_date.strftime('%Y-%m-%d') if device.install_date else '',
            device.disk_serial or '',
            device.purchase_date.strftime('%Y-%m-%d') if device.purchase_date else '',
            device.enable_date.strftime('%Y-%m-%d') if device.enable_date else '',
            device.purpose or '',
            device.remarks or '',
            '是' if device.is_fixed else '否',
            device.asset_card_no or '',
            '是' if device.is_secret else '否',
            device.secret_category or '',
        ]
        ws.append(row)
    
    # 设置数据区域边框和对齐
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value or ''))
                # 中文字符按2计算宽度
                for char in str(cell.value or ''):
                    if '\u4e00' <= char <= '\u9fff':
                        cell_len += 1
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws.column_dimensions[column].width = min(max(max_length + 2, 8), 40)
    
    # 生成文件名带时间码和随机码（使用系统配置的时区）
    try:
        import pytz
        timezone_str = get_config_value('timezone', 'Asia/Shanghai')
        local_tz = pytz.timezone(timezone_str)
        now = tz.now().astimezone(local_tz)
    except:
        now = tz.localtime(tz.now())
    random_code = ''.join(random.choices(string.ascii_letters + string.digits, k=4))
    filename = f'device-export-{now.strftime("%Y%m%d")}-{now.strftime("%H%M%S")}-{random_code}.xlsx'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def process_device_import_task(task_id, file_content, update_existing, user_id):
    """后台处理设备导入任务"""
    import openpyxl
    from io import BytesIO
    
    progress = device_import_progress[task_id]
    
    try:
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
    except Exception as e:
        progress['status'] = 'error'
        progress['errors'].append(f'文件读取失败: {str(e)}')
        return
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    progress['total'] = len(rows)
    
    app_url = get_config_value('app_url', 'http://127.0.0.1:8000').rstrip('/')
    user = User.objects.filter(pk=user_id).first()
    
    secret_level_map = {
        '公开': 'public', '内部': 'internal', '秘密': 'confidential',
        '机密': 'secret', '绝密': 'top_secret', '商密': 'commercial_secret',
    }
    
    for idx, row in enumerate(rows, start=2):
        progress['current'] = idx - 1
        
        asset_no = str(row[0]).strip() if row[0] else ''
        device_no = str(row[1]).strip() if row[1] else ''
        name = str(row[2]).strip() if row[2] else ''
        model = str(row[3]).strip() if row[3] else ''
        serial_no = str(row[4]).strip() if row[4] else ''
        secret_level_text = str(row[5]).strip() if row[5] else ''
        mac_address = str(row[6]).strip() if row[6] else ''
        ip_address = str(row[7]).strip() if row[7] else ''
        os_name = str(row[8]).strip() if row[8] else ''
        install_date = row[9] if row[9] else None
        disk_serial = str(row[10]).strip() if row[10] else ''
        purchase_date = row[11] if row[11] else None
        enable_date = row[12] if row[12] else None
        purpose = str(row[13]).strip() if row[13] else ''
        remarks = str(row[14]).strip() if row[14] else ''
        is_fixed_text = str(row[15]).strip() if row[15] else ''
        asset_card_no = str(row[16]).strip() if row[16] else ''
        is_secret_text = str(row[17]).strip() if row[17] else ''
        secret_category = str(row[18]).strip() if row[18] else ''
        
        progress['current_asset_no'] = asset_no
        
        if not asset_no:
            progress['error'] += 1
            progress['errors'].append(f'第{idx}行：资产编号不能为空')
            continue
        
        if not name:
            progress['error'] += 1
            progress['errors'].append(f'第{idx}行：设备名称不能为空')
            continue
        
        # 根据资产编号自动匹配资产分类
        parts = asset_no.split('-')
        category = None
        if len(parts) >= 2:
            prefix = '-'.join(parts[:-1])
            category = AssetCategory.find_by_asset_prefix(prefix)
        
        if not category:
            progress['error'] += 1
            progress['errors'].append(f'第{idx}行：未找到匹配的资产分类（资产编号：{asset_no}）')
            continue
        
        # 处理密级
        secret_level = secret_level_map.get(secret_level_text, 'public')
        
        # 处理日期
        if isinstance(install_date, datetime):
            install_date = install_date.date()
        if isinstance(purchase_date, datetime):
            purchase_date = purchase_date.date()
        if isinstance(enable_date, datetime):
            enable_date = enable_date.date()
        
        # 处理布尔值
        is_fixed = is_fixed_text in ['是', '1', 'true', 'True', 'Y', 'y']
        is_secret = is_secret_text in ['是', '1', 'true', 'True', 'Y', 'y']
        
        # 检查设备是否已存在
        device = Device.objects.filter(asset_no=asset_no).first()
        
        if device:
            if update_existing:
                device.device_no = device_no or device.device_no
                device.name = name or device.name
                device.model = model or device.model
                device.serial_no = serial_no or device.serial_no
                device.secret_level = secret_level
                device.mac_address = mac_address or device.mac_address
                device.ip_address = ip_address or device.ip_address
                device.os_name = os_name or device.os_name
                device.install_date = install_date or device.install_date
                device.disk_serial = disk_serial or device.disk_serial
                device.purchase_date = purchase_date or device.purchase_date
                device.enable_date = enable_date or device.enable_date
                device.purpose = purpose or device.purpose
                device.remarks = remarks or device.remarks
                device.is_fixed = is_fixed
                device.asset_card_no = asset_card_no or device.asset_card_no
                device.is_secret = is_secret
                device.secret_category = secret_category or device.secret_category
                device.category = category
                device.save()
                progress['update'] += 1
            else:
                progress['skip'] += 1
        else:
            try:
                device = Device.objects.create(
                    asset_no=asset_no,
                    device_no=device_no,
                    name=name,
                    model=model,
                    serial_no=serial_no,
                    category=category,
                    secret_level=secret_level,
                    mac_address=mac_address,
                    ip_address=ip_address,
                    os_name=os_name,
                    install_date=install_date,
                    disk_serial=disk_serial,
                    purchase_date=purchase_date,
                    enable_date=enable_date,
                    purpose=purpose,
                    remarks=remarks,
                    is_fixed=is_fixed,
                    asset_card_no=asset_card_no,
                    is_secret=is_secret,
                    secret_category=secret_category,
                    status='unused',
                    created_by=user,
                )
                
                # 生成二维码
                qr = qrcode.QRCode(version=1, box_size=10, border=1)
                qr.add_data(f'{app_url}/assets/view/{device.asset_no}')
                qr.make(fit=True)
                img = qr.make_image(fill_color="black", back_color="white")
                buffer = BytesIO()
                img.save(buffer, 'PNG')
                device.qrcode.save(f'QR-{asset_no}.png', buffer)
                
                progress['success'] += 1
            except Exception as e:
                progress['error'] += 1
                progress['errors'].append(f'第{idx}行：创建失败 - {str(e)}')
        
        time.sleep(0.01)
    
    progress['current'] = progress['total']
    progress['status'] = 'completed'


@login_required
@csrf_exempt
def device_import(request):
    if request.method == 'POST':
        import openpyxl
        
        file = request.FILES.get('file')
        if not file:
            return JsonResponse({'success': False, 'message': '请选择要导入的文件'})
        
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'success': False, 'message': '请上传Excel文件(.xlsx或.xls)'})
        
        update_existing = request.POST.get('update_existing') == 'true'
        
        task_id = str(uuid.uuid4())[:8]
        file_content = file.read()
        
        device_import_progress[task_id] = {
            'status': 'processing',
            'total': 0,
            'current': 0,
            'success': 0,
            'update': 0,
            'skip': 0,
            'error': 0,
            'errors': [],
            'current_asset_no': ''
        }
        
        thread = threading.Thread(
            target=process_device_import_task,
            args=(task_id, file_content, update_existing, request.user.id)
        )
        thread.daemon = True
        thread.start()
        
        return JsonResponse({'success': True, 'task_id': task_id})
    
    return render(request, 'assets/device_import.html')


@login_required
def device_import_progress_api(request):
    task_id = request.GET.get('task_id')
    if task_id and task_id in device_import_progress:
        return JsonResponse(device_import_progress[task_id])
    return JsonResponse({'status': 'not_found'})


@login_required
def software_import_progress_api(request):
    task_id = request.GET.get('task_id')
    if task_id and task_id in software_import_progress:
        return JsonResponse(software_import_progress[task_id])
    return JsonResponse({'status': 'not_found'})


@login_required
@permission_required('service_contract')
def service_contract_list(request):
    search = request.GET.get('search', '')
    service_type = request.GET.get('service_type', '')
    
    contracts = ServiceContract.objects.all().order_by('-created_at')
    
    if search:
        contracts = contracts.filter(
            Q(name__icontains=search) |
            Q(service_type__icontains=search) |
            Q(description__icontains=search)
        )
    if service_type:
        contracts = contracts.filter(service_type=service_type)
    
    page_size = int(request.GET.get('page_size', 20))
    if page_size not in [20, 50, 100, 200]:
        page_size = 20
    
    paginator = Paginator(contracts, page_size)
    page = request.GET.get('page', 1)
    contracts = paginator.get_page(page)
    
    current_page = contracts.number
    total_pages = paginator.num_pages
    
    page_range = []
    if total_pages <= 7:
        page_range = list(range(1, total_pages + 1))
    else:
        page_range.extend([1, 2])
        start = max(3, current_page - 2)
        end = min(total_pages - 1, current_page + 2)
        if start > 3:
            page_range.append('...')
        page_range.extend(range(start, end + 1))
        if end < total_pages - 1:
            page_range.append('...')
        page_range.extend([total_pages - 1, total_pages])
    
    service_types = ['维保服务', '技术支持', '售后服务', '软件更新服务', '培训服务', '其他']
    
    return render(request, 'assets/service_contract_list.html', {
        'contracts': contracts,
        'service_types': service_types,
        'page_range': page_range,
        'page_size': page_size,
    })


@login_required
@permission_required('service_contract_create')
def service_contract_create(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        service_type = request.POST.get('service_type')
        price = request.POST.get('price')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        description = request.POST.get('description')
        
        if not name or not start_date or not end_date:
            messages.error(request, '服务名称和服务期限为必填项')
            return redirect('service_contract_create')
        
        ServiceContract.objects.create(
            name=name,
            service_type=service_type or '',
            price=price or None,
            start_date=start_date,
            end_date=end_date,
            description=description or '',
        )
        messages.success(request, '服务创建成功')
        return redirect('service_contract_list')
    
    service_types = ['维保服务', '技术支持', '售后服务', '软件更新服务', '培训服务', '其他']
    return render(request, 'assets/service_contract_form.html', {'service_types': service_types})


@login_required
def service_contract_detail(request, pk):
    contract = get_object_or_404(ServiceContract, pk=pk)
    from datetime import date
    today = date.today()
    return render(request, 'assets/service_contract_detail.html', {
        'contract': contract,
        'today': today
    })


@login_required
def service_contract_edit(request, pk):
    contract = get_object_or_404(ServiceContract, pk=pk)
    
    if request.method == 'POST':
        contract.name = request.POST.get('name')
        contract.service_type = request.POST.get('service_type') or ''
        contract.price = request.POST.get('price') or None
        contract.start_date = request.POST.get('start_date')
        contract.end_date = request.POST.get('end_date')
        contract.description = request.POST.get('description') or ''
        contract.save()
        messages.success(request, '服务更新成功')
        return redirect('service_contract_detail', pk=pk)
    
    service_types = ['维保服务', '技术支持', '售后服务', '软件更新服务', '培训服务', '其他']
    return render(request, 'assets/service_contract_form.html', {
        'contract': contract,
        'service_types': service_types
    })


@login_required
def service_contract_renew(request, pk):
    contract = get_object_or_404(ServiceContract, pk=pk)
    
    if request.method == 'POST':
        new_start_date = request.POST.get('new_start_date')
        new_end_date = request.POST.get('new_end_date')
        new_price = request.POST.get('new_price')
        new_description = request.POST.get('new_description')
        
        if not new_start_date or not new_end_date:
            messages.error(request, '新的服务期限为必填项')
            return redirect('service_contract_renew', pk=pk)
        
        contract.start_date = new_start_date
        contract.end_date = new_end_date
        if new_price:
            contract.price = new_price
        if new_description:
            contract.description = new_description
        contract.save()
        
        messages.success(request, '服务续期成功')
        return redirect('service_contract_detail', pk=pk)
    
    return render(request, 'assets/service_contract_renew.html', {'contract': contract})


@login_required
@require_http_methods(["POST"])
def service_contract_delete(request, pk):
    contract = get_object_or_404(ServiceContract, pk=pk)
    contract.delete()
    messages.success(request, '服务删除成功')
    return redirect('service_contract_list')


# ============ 更新卡片编号功能 ============

# 更新卡片编号进度存储
update_card_no_progress = {}


def add_log(task_id, level, message):
    """添加处理日志"""
    import datetime
    if task_id in update_card_no_progress:
        update_card_no_progress[task_id]['logs'].append({
            'time': datetime.datetime.now().strftime('%H:%M:%S'),
            'level': level,  # info, success, warning, error, ai
            'message': message
        })


@login_required
def update_card_no(request):
    """更新卡片编号 - 页面渲染"""
    return render(request, 'assets/update_card_no.html')


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def update_card_no_api(request):
    """更新卡片编号 - 文件上传处理"""
    import openpyxl
    
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'success': False, 'message': '请选择要上传的文件'})
    
    if not file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'success': False, 'message': '请上传Excel文件(.xlsx或.xls)'})
    
    task_id = str(uuid.uuid4())[:8]
    file_content = file.read()
    
    update_card_no_progress[task_id] = {
        'status': 'processing',
        'total': 0,
        'current': 0,
        'updated': 0,
        'skipped': 0,
        'error': 0,
        'errors': [],
        'current_asset_no': '',
        'logs': []
    }
    
    thread = threading.Thread(
        target=process_update_card_no_task,
        args=(task_id, file_content, request.user.id)
    )
    thread.daemon = True
    thread.start()
    
    return JsonResponse({'success': True, 'task_id': task_id})


@login_required
def update_card_no_progress_api(request):
    """更新卡片编号进度查询API"""
    task_id = request.GET.get('task_id')
    if task_id and task_id in update_card_no_progress:
        return JsonResponse(update_card_no_progress[task_id])
    return JsonResponse({'status': 'not_found'})


def expand_asset_numbers(base_no, count):
    """展开合并写法的资产编号"""
    import re
    
    results = []
    
    # 处理 / 分隔的格式: XACD-Z-001-001-001/002/003
    slash_match = re.match(r'^(.+)-(\d+)/(.+)$', base_no)
    if slash_match:
        prefix = slash_match.group(1)
        first_num = slash_match.group(2)
        rest_nums = slash_match.group(3).split('/')
        results.append(f"{prefix}-{first_num}")
        for num in rest_nums:
            results.append(f"{prefix}-{num}")
        return results[:count] if count else results
    
    # 处理 ~ 连号格式: XACD-Z-001-001-006~009
    range_match = re.match(r'^(.+)-(\d+)~(\d+)$', base_no)
    if range_match:
        prefix = range_match.group(1)
        start = int(range_match.group(2))
        end = int(range_match.group(3))
        num_len = len(range_match.group(2))
        for i in range(start, end + 1):
            results.append(f"{prefix}-{str(i).zfill(num_len)}")
        return results[:count] if count else results
    
    # 普通格式，直接返回
    return [base_no]


def call_ai_parse_card_numbers(excel_data):
    """调用AI解析资产编号和卡片编号，使用系统设置中配置的大模型"""
    prompt = f"""请从以下Excel表格数据中提取资产编号和卡片编号的对应关系。

规则：
1. 自动识别包含"资产编号"、"卡片编号"、"资产数量"的列
2. 资产编号格式类似: XACD-Z-001-001-001
3. 合并写法如 XACD-Z-001-001-001/002/003 表示多个编号
4. 连号写法如 XACD-Z-001-001-006~009 表示连续编号
5. 结合资产数量字段判断拆解数量
6. 返回JSON格式，包含展开后的资产编号和对应卡片编号：
{{
  "mappings": [
    {{"asset_no": "XACD-Z-001-001-001", "card_no": "CARD-001"}},
    {{"asset_no": "XACD-Z-001-001-002", "card_no": "CARD-001"}},
    ...
  ]
}}

表格数据：
{excel_data}"""
    
    try:
        ai_content = call_llm([
            {"role": "system", "content": "你是一个专业的数据解析助手，擅长从表格数据中提取设备资产编号和卡片编号的对应关系。"},
            {"role": "user", "content": prompt}
        ])
        
        json_match = re.search(r'\{[\s\S]*\}', ai_content)
        if json_match:
            return json.loads(json_match.group())
        return None
    except Exception as e:
        return None


def process_update_card_no_task(task_id, file_content, user_id):
    """后台处理更新卡片编号任务"""
    import openpyxl
    from io import BytesIO
    
    progress = update_card_no_progress[task_id]
    
    try:
        add_log(task_id, 'info', '开始读取Excel文件...')
        
        # 读取Excel
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active
        row_count = ws.max_row - 1 if ws.max_row else 0
        add_log(task_id, 'success', f'Excel文件读取成功，共{row_count}行数据')
        
        # 自动识别列
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else '')
        
        asset_no_col = None
        card_no_col = None
        quantity_col = None
        
        for idx, header in enumerate(headers):
            if '资产编号' in header or '资产编码' in header:
                asset_no_col = idx
            elif '卡片编号' in header or '卡片码' in header:
                card_no_col = idx
            elif '数量' in header or '资产数量' in header:
                quantity_col = idx
        
        col_info = []
        if asset_no_col is not None:
            col_info.append(f'资产编号({chr(65+asset_no_col)}列)')
        if card_no_col is not None:
            col_info.append(f'卡片编号({chr(65+card_no_col)}列)')
        if quantity_col is not None:
            col_info.append(f'资产数量({chr(65+quantity_col)}列)')
        add_log(task_id, 'info', f'识别到列：{"、".join(col_info)}')
        
        if asset_no_col is None:
            progress['status'] = 'completed'
            progress['errors'].append('未找到资产编号列')
            add_log(task_id, 'error', '未找到资产编号列，处理终止')
            return
        
        if card_no_col is None:
            progress['status'] = 'completed'
            progress['errors'].append('未找到卡片编号列')
            add_log(task_id, 'error', '未找到卡片编号列，处理终止')
            return
        
        # 构建Excel数据映射
        add_log(task_id, 'info', '开始解析Excel数据...')
        excel_mappings = {}
        expanded_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > max(asset_no_col, card_no_col):
                asset_no = str(row[asset_no_col]).strip() if row[asset_no_col] else ''
                card_no = str(row[card_no_col]).strip() if row[card_no_col] else ''
                quantity = int(row[quantity_col]) if quantity_col and row[quantity_col] else None
                
                if asset_no:
                    # 展开合并写法
                    expanded = expand_asset_numbers(asset_no, quantity)
                    if len(expanded) > 1:
                        expanded_count += 1
                        add_log(task_id, 'info', f'展开合并写法：{asset_no} → {", ".join(expanded[:5])}{"..." if len(expanded) > 5 else ""}')
                    for no in expanded:
                        excel_mappings[no] = card_no
        
        if expanded_count > 0:
            add_log(task_id, 'success', f'共展开{expanded_count}个合并写法的资产编号')
        add_log(task_id, 'success', f'Excel数据解析完成，共{len(excel_mappings)}条资产编号映射')
        
        if not excel_mappings:
            progress['status'] = 'completed'
            progress['errors'].append('Excel中未找到有效的资产编号数据')
            add_log(task_id, 'error', 'Excel中未找到有效的资产编号数据，处理终止')
            return
        
        # 获取AI配置
        if is_llm_enabled():
            config = get_llm_config()
            model_name = config['model_name']
            add_log(task_id, 'ai', f'获取AI配置成功：模型={model_name}')
            add_log(task_id, 'ai', f'调用AI解析数据，发送{len(excel_mappings)}条映射记录...')
            
            excel_text = '\n'.join([f"{k}|{v}" for k, v in excel_mappings.items()])
            ai_result = call_ai_parse_card_numbers(excel_text)
            
            if ai_result and 'mappings' in ai_result:
                ai_mappings = {item['asset_no']: item['card_no'] for item in ai_result['mappings']}
                add_log(task_id, 'success', f'AI解析成功，返回{len(ai_mappings)}条映射记录')
                add_log(task_id, 'ai', f'AI返回示例：{", ".join([f"{k}→{v}" for k, v in list(ai_mappings.items())[:3]])}...')
                excel_mappings = ai_mappings
            else:
                add_log(task_id, 'warning', 'AI解析失败或返回格式错误，使用本地解析结果')
        else:
            add_log(task_id, 'info', 'LLM未启用或未配置API Key，使用本地解析')
        
        # 遍历系统设备
        add_log(task_id, 'info', '开始遍历系统设备...')
        devices = Device.objects.all()
        total = devices.count()
        progress['total'] = total
        add_log(task_id, 'info', f'系统中共{total}个设备待处理')
        
        for device in devices:
            progress['current'] += 1
            progress['current_asset_no'] = device.asset_no
            
            try:
                if device.asset_no in excel_mappings:
                    new_card_no = excel_mappings[device.asset_no]
                    
                    if not new_card_no:
                        # Excel中卡片编号为空，跳过
                        progress['skipped'] += 1
                        add_log(task_id, 'warning', f'设备{device.asset_no}：Excel卡片编号为空，跳过')
                    elif device.asset_card_no:
                        # 设备已有卡片编号，跳过但记录比对结果
                        if device.asset_card_no == new_card_no:
                            add_log(task_id, 'info', f'设备{device.asset_no}：卡片编号相同，{new_card_no}，跳过')
                        else:
                            add_log(task_id, 'info', f'设备{device.asset_no}：卡片编号不同，现有:{device.asset_card_no}，Excel:{new_card_no}，跳过')
                        progress['skipped'] += 1
                    else:
                        # 设备卡片编号为空，更新
                        device.asset_card_no = new_card_no
                        device.is_fixed = True
                        device.save(update_fields=['asset_card_no', 'is_fixed'])
                        progress['updated'] += 1
                        add_log(task_id, 'success', f'设备{device.asset_no}：更新卡片编号为{new_card_no}，固资在账设为"是"')
                else:
                    # 表格中未找到该设备，跳过
                    progress['skipped'] += 1
                    add_log(task_id, 'warning', f'设备{device.asset_no}：表格中未找到，跳过')
            except Exception as e:
                progress['error'] += 1
                progress['errors'].append(f"{device.asset_no}: {str(e)}")
                add_log(task_id, 'error', f'设备{device.asset_no}：处理失败 - {str(e)}')
            
            # 每处理10个设备更新一次进度
            if progress['current'] % 10 == 0:
                time.sleep(0.1)
        
        progress['status'] = 'completed'
        progress['current_asset_no'] = '处理完成'
        add_log(task_id, 'success', f'处理完成！共处理{total}个设备，更新{progress["updated"]}个，跳过{progress["skipped"]}个，失败{progress["error"]}个')
        
    except Exception as e:
        progress['status'] = 'completed'
        progress['errors'].append(f"处理失败: {str(e)}")
        add_log(task_id, 'error', f'处理失败: {str(e)}')
