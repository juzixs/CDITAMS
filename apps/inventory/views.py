from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q, F, Case, When, IntegerField
import json
import random
import re
import os
import threading
import uuid
import time
from io import BytesIO

from .models import InventoryPlan, InventoryTask, InventoryRecord, InventoryTaskDevice
from apps.assets.models import Device, AssetLocation, AssetCategory
from apps.assets.views import save_photo_with_asset_no
from apps.accounts.models import User, Department
from apps.accounts.decorators import permission_required
from apps.settings.llm_service import is_llm_enabled, call_llm, get_llm_config, call_llm_vision_two_step
from apps.settings.views import get_config_value


# 设备导入进度存储
inventory_import_progress = {}


def inventory_add_log(task_id, level, message):
    """添加处理日志"""
    import datetime
    if task_id in inventory_import_progress:
        inventory_import_progress[task_id]['logs'].append({
            'time': datetime.datetime.now().strftime('%H:%M:%S'),
            'level': level,
            'message': message
        })


def parse_ai_json_response(ai_content):
    """解析AI返回的JSON内容，支持多种格式"""
    if not ai_content:
        return None, 'AI返回内容为空'
    
    # 清理内容，移除可能的markdown标记
    ai_content = ai_content.strip()
    if ai_content.startswith('```'):
        lines = ai_content.split('\n')
        ai_content = '\n'.join(lines[1:-1] if lines[-1] == '```' else lines[1:])
    ai_content = ai_content.strip()
    
    # 尝试直接解析
    try:
        return json.loads(ai_content), None
    except:
        pass
    
    # 尝试提取JSON对象
    match = re.search(r'\{[\s\S]*\}', ai_content)
    if match:
        try:
            return json.loads(match.group()), None
        except:
            pass
    
    # 尝试提取JSON数组
    match = re.search(r'\[[\s\S]*\]', ai_content)
    if match:
        try:
            return json.loads(match.group()), None
        except:
            pass
    
    return None, '无法解析JSON格式'


# ==================== 任务管理 ====================

@login_required
@permission_required('inventory_task')
def task_list(request):
    """任务列表页"""
    status = request.GET.get('status', '')
    task_type = request.GET.get('task_type', '')
    search = request.GET.get('search', '')
    
    tasks = InventoryTask.objects.select_related('created_by', 'assignee').all()
    
    if status:
        tasks = tasks.filter(status=status)
    if task_type:
        tasks = tasks.filter(task_type=task_type)
    if search:
        tasks = tasks.filter(
            Q(task_no__icontains=search) |
            Q(name__icontains=search)
        )
    
    return render(request, 'inventory/task_list.html', {
        'tasks': tasks,
        'status_filter': status,
        'task_type_filter': task_type,
        'search': search,
    })


@login_required
@permission_required('inventory_task_create')
def task_create(request):
    """创建盘点任务"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        task_type = request.POST.get('task_type', 'full')
        is_fixed_only = request.POST.get('is_fixed_only') == 'on'
        scheduled_start = request.POST.get('scheduled_start')
        scheduled_end = request.POST.get('scheduled_end')
        remarks = request.POST.get('remarks', '').strip()
        
        # 动态盘点任务直接进入执行状态
        initial_status = 'in_progress' if task_type == 'dynamic' else 'pending'
        
        task = InventoryTask.objects.create(
            name=name,
            task_type=task_type,
            is_fixed_only=is_fixed_only,
            status=initial_status,
            created_by=request.user,
            scheduled_start=scheduled_start if scheduled_start else None,
            scheduled_end=scheduled_end if scheduled_end else None,
            remarks=remarks,
        )
        
        # 全盘和动态盘点自动生成设备列表
        if task_type in ('full', 'dynamic'):
            generate_task_devices(task)
        
        messages.success(request, f'盘点任务创建成功: {task.task_no}')
        
        # 动态盘点任务直接跳转到执行页面
        if task_type == 'dynamic':
            return redirect('inventory_task_execute', pk=task.pk)
        return redirect('inventory_task_detail', pk=task.pk)
    
    return render(request, 'inventory/task_create.html')


@login_required
def task_detail(request, pk):
    """任务详情页（含待盘/已盘列表）"""
    task = get_object_or_404(InventoryTask.objects.select_related('created_by', 'assignee'), pk=pk)
    
    # 动态盘点任务自动同步设备列表
    if task.task_type == 'dynamic' and task.status == 'in_progress':
        sync_dynamic_task_devices(task)
    
    # 待盘点设备 - 有位置的在前，无位置的在后，按位置排序
    pending_devices = InventoryTaskDevice.objects.filter(
        task=task, status='pending'
    ).select_related(
        'device__category', 'device__location', 'device__user', 'device__department', 'device__workstation'
    ).annotate(
        has_location=Case(
            When(device__location_text='', then=1),
            default=0,
            output_field=IntegerField(),
        )
    ).order_by('has_location', 'device__location_text', 'device__id')
    
    # 已盘点设备
    checked_devices = InventoryTaskDevice.objects.filter(
        task=task, status='checked'
    ).select_related(
        'device__category', 'device__location', 'device__user', 'device__department', 'device__workstation'
    ).order_by('-checked_at')
    
    # 统计
    total_count = InventoryTaskDevice.objects.filter(task=task).count()
    pending_count = pending_devices.count()
    checked_count = checked_devices.count()
    
    return render(request, 'inventory/task_detail.html', {
        'task': task,
        'pending_devices': pending_devices,
        'checked_devices': checked_devices,
        'total_count': total_count,
        'pending_count': pending_count,
        'checked_count': checked_count,
    })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
@permission_required('inventory_task_delete')
def task_delete(request, pk):
    """删除任务"""
    task = get_object_or_404(InventoryTask, pk=pk)
    if task.status == 'completed':
        return JsonResponse({'success': False, 'message': '已完成的任务不能删除'})
    
    task.delete()
    messages.success(request, '任务已删除')
    return JsonResponse({'success': True})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def task_start(request, pk):
    """开始任务"""
    task = get_object_or_404(InventoryTask, pk=pk)
    if task.status != 'pending':
        return JsonResponse({'success': False, 'message': '只有待执行的任务才能开始'})
    
    device_count = InventoryTaskDevice.objects.filter(task=task).count()
    if device_count == 0:
        return JsonResponse({'success': False, 'message': '请先添加待盘点设备'})
    
    task.status = 'in_progress'
    task.device_count = device_count
    task.save()
    
    return JsonResponse({'success': True, 'message': '任务已开始'})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def task_complete(request, pk):
    """完成任务"""
    task = get_object_or_404(InventoryTask, pk=pk)
    if task.status != 'in_progress':
        return JsonResponse({'success': False, 'message': '只有执行中的任务才能完成'})
    
    # 动态盘点任务：只统计已盘点设备，删除未盘点的设备
    if task.task_type == 'dynamic':
        # 删除未盘点的设备记录
        InventoryTaskDevice.objects.filter(task=task, status='pending').delete()
        # 更新设备数量为已盘点数量
        task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
    
    task.status = 'completed'
    task.completed_at = timezone.now()
    task.save()
    
    return JsonResponse({'success': True, 'message': '任务已完成'})


# ==================== 设备抽选 ====================

@login_required
def get_device_pool(request, task_id):
    """获取设备池（根据任务类型和筛选条件）"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    # 基础查询：排除报废设备
    devices = Device.objects.exclude(status='scrapped').select_related('category', 'location', 'user')
    
    # 固资在账筛选
    if task.is_fixed_only:
        devices = devices.filter(is_fixed=True)
    
    return devices


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_random_select(request, task_id):
    """随机抽选设备"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        data = json.loads(request.body)
        select_type = data.get('select_type', 'percent')  # percent 或 count
        select_value = int(data.get('select_value', 0))
        
        if select_value <= 0:
            return JsonResponse({'success': False, 'message': '请输入有效的抽选值'})
        
        # 获取设备池
        devices = get_device_pool(request, task_id)
        
        # 排除已添加的设备
        existing_device_ids = InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True)
        devices = devices.exclude(id__in=existing_device_ids)
        
        device_list = list(devices)
        total_count = len(device_list)
        
        if total_count == 0:
            return JsonResponse({'success': False, 'message': '没有可抽选的设备'})
        
        # 计算抽选数量
        if select_type == 'percent':
            select_count = max(1, int(total_count * select_value / 100))
        else:
            select_count = min(select_value, total_count)
        
        # 随机抽选
        selected_devices = random.sample(device_list, select_count)
        
        # 添加到任务设备列表
        current_max_order = InventoryTaskDevice.objects.filter(task=task).count()
        for i, device in enumerate(selected_devices):
            InventoryTaskDevice.objects.create(
                task=task,
                device=device,
                status='pending',
                sort_order=current_max_order + i + 1,
                added_by=request.user,
            )
        
        # 更新设备计数
        task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
        task.save()
        
        return JsonResponse({
            'success': True,
            'message': f'已随机抽选 {select_count} 台设备',
            'selected_count': select_count,
            'total_count': task.device_count,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def api_device_search(request, task_id):
    """搜索设备（人工抽选）"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    q = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 20))
    
    # 获取设备池
    devices = get_device_pool(request, task_id)
    
    # 排除已添加的设备
    existing_device_ids = InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True)
    devices = devices.exclude(id__in=existing_device_ids)
    
    # 搜索
    if q:
        devices = devices.filter(
            Q(asset_no__icontains=q) |
            Q(name__icontains=q) |
            Q(serial_no__icontains=q) |
            Q(device_no__icontains=q) |
            Q(model__icontains=q) |
            Q(user__realname__icontains=q) |
            Q(department__name__icontains=q) |
            Q(location_text__icontains=q) |
            Q(asset_card_no__icontains=q) |
            Q(secret_level__icontains=q) |
            Q(status__icontains=q)
        ).distinct()
    
    # 分页
    paginator = Paginator(devices, page_size)
    page_obj = paginator.get_page(page)
    
    # 序列化
    device_list = []
    for d in page_obj:
        device_list.append({
            'id': d.id,
            'asset_no': d.asset_no,
            'device_no': d.device_no or '',
            'name': d.name,
            'model': d.model or '',
            'secret_level': d.get_secret_level_display(),
            'secret_level_code': d.secret_level,
            'department': d.department.name if d.department else '',
            'user': d.user.realname if d.user else '',
            'location': d.location.get_full_path() if d.location else '',
            'location_text': d.location_text or '',
            'status': d.get_status_display(),
            'status_code': d.status,
            'is_fixed': d.is_fixed,
            'asset_card_no': d.asset_card_no or '',
            'category': d.category.name if d.category else '',
        })
    
    return JsonResponse({
        'success': True,
        'devices': device_list,
        'total': paginator.count,
        'page': page_obj.number,
        'total_pages': paginator.num_pages,
    })


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_add_devices(request, task_id):
    """添加设备到待盘列表"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        data = json.loads(request.body)
        device_ids = data.get('device_ids', [])
        
        if not device_ids:
            return JsonResponse({'success': False, 'message': '请选择设备'})
        
        # 排除已添加的
        existing_ids = set(InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True))
        new_ids = [did for did in device_ids if did not in existing_ids]
        
        if not new_ids:
            return JsonResponse({'success': False, 'message': '所选设备已在待盘列表中'})
        
        # 获取最大排序号
        current_max_order = InventoryTaskDevice.objects.filter(task=task).count()
        
        # 批量创建
        devices = Device.objects.filter(id__in=new_ids)
        task_devices = []
        for i, device in enumerate(devices):
            task_devices.append(InventoryTaskDevice(
                task=task,
                device=device,
                status='pending',
                sort_order=current_max_order + i + 1,
                added_by=request.user,
            ))
        InventoryTaskDevice.objects.bulk_create(task_devices)
        
        # 更新计数
        task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
        task.save()
        
        return JsonResponse({
            'success': True,
            'message': f'已添加 {len(new_ids)} 台设备',
            'added_count': len(new_ids),
            'total_count': task.device_count,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_remove_device(request, task_id):
    """从待盘列表移除设备"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        data = json.loads(request.body)
        task_device_id = data.get('task_device_id')
        
        task_device = get_object_or_404(InventoryTaskDevice, pk=task_device_id, task=task)
        
        if task_device.status == 'checked':
            return JsonResponse({'success': False, 'message': '已盘点的设备不能移除'})
        
        task_device.delete()
        
        # 更新计数
        task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
        task.save()
        
        return JsonResponse({'success': True, 'message': '设备已移除'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== AI解析导入 ====================

def expand_asset_no_pattern(asset_no_str):
    """
    解析类似 XACD-Z-001-001-001/003/007 的合并写法
    返回: ['XACD-Z-001-001-001', 'XACD-Z-001-001-003', 'XACD-Z-001-001-007']
    """
    asset_no_str = asset_no_str.strip()
    if not asset_no_str:
        return []
    
    if '/' not in asset_no_str:
        return [asset_no_str]
    
    parts = asset_no_str.split('/')
    base = parts[0].strip()
    
    if not base:
        return []
    
    # 找到最后一个 - 的位置
    last_dash = base.rfind('-')
    if last_dash == -1:
        return [base]
    
    prefix = base[:last_dash + 1]
    base_num = base[last_dash + 1:]
    
    result = [base]
    for part in parts[1:]:
        part = part.strip()
        if not part:
            continue
        # 如果是纯数字，保持与基础编号相同的位数
        if part.isdigit() and base_num.isdigit():
            padded = part.zfill(len(base_num))
            result.append(f"{prefix}{padded}")
        else:
            result.append(f"{prefix}{part}")
    
    return result


def parse_asset_numbers_with_ai(content, content_type='text'):
    """使用AI解析资产编号，使用系统设置中配置的大模型"""
    if not is_llm_enabled():
        return None, '未启用LLM或未配置API Key，请在系统设置-模型设置中配置'
    
    prompt = """请从以下内容中提取所有设备资产编号。

规则：
1. 资产编号格式类似: XACD-Z-001-001-001 或类似格式
2. 合并写法如 XACD-Z-001-001-001/003/007 需拆解为独立编号
3. 有些编号可能显示不全，尽量还原完整编号
4. 同时识别卡片编号
5. 返回JSON格式，包含两个数组:
{
  "asset_numbers": ["编号1", "编号2"],
  "card_numbers": ["卡片编号1"]
}

内容:
"""
    
    try:
        messages_list = [
            {"role": "system", "content": "你是一个专业的数据解析助手，擅长从表格、图片等文档中提取设备资产编号信息。"},
            {"role": "user", "content": prompt + content}
        ]
        
        ai_content = call_llm(messages_list)
        
        json_match = re.search(r'\{[\s\S]*\}', ai_content)
        if json_match:
            data = json.loads(json_match.group())
            return data, None
        else:
            return None, 'AI返回格式异常'
            
    except Exception as e:
        return None, f'AI解析失败: {str(e)}'


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_ai_parse(request):
    """AI解析导入设备"""
    try:
        data = json.loads(request.body)
        content = data.get('content', '')
        content_type = data.get('content_type', 'text')
        
        if not content:
            return JsonResponse({'success': False, 'message': '请提供解析内容'})
        
        result, error = parse_asset_numbers_with_ai(content, content_type)
        
        if error:
            return JsonResponse({'success': False, 'message': error})
        
        # 展开合并写法的编号
        all_asset_numbers = []
        for no in result.get('asset_numbers', []):
            expanded = expand_asset_no_pattern(no)
            all_asset_numbers.extend(expanded)
        
        all_card_numbers = result.get('card_numbers', [])
        
        # 查找匹配的设备
        found_devices = Device.objects.filter(
            Q(asset_no__in=all_asset_numbers) | Q(asset_card_no__in=all_card_numbers)
        ).exclude(status='scrapped').select_related('category', 'location', 'user')
        
        device_list = []
        for d in found_devices:
            device_list.append({
                'id': d.id,
                'asset_no': d.asset_no,
                'name': d.name,
                'model': d.model or '',
                'category': d.category.name if d.category else '',
                'location': d.location.get_full_path() if d.location else '',
                'user': d.user.realname if d.user else '',
            })
        
        # 未找到的编号
        found_asset_nos = set(found_devices.values_list('asset_no', flat=True))
        found_card_nos = set(found_devices.values_list('asset_card_no', flat=True))
        not_found = [no for no in all_asset_numbers if no not in found_asset_nos]
        not_found += [no for no in all_card_numbers if no not in found_card_nos]
        
        return JsonResponse({
            'success': True,
            'devices': device_list,
            'not_found': not_found,
            'parsed_count': len(all_asset_numbers) + len(all_card_numbers),
            'found_count': len(device_list),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_ai_parse_file(request):
    """AI解析文件（表格/图片）"""
    try:
        task_id = request.POST.get('task_id')
        file = request.FILES.get('file')
        
        if not file:
            return JsonResponse({'success': False, 'message': '请上传文件'})
        
        task = get_object_or_404(InventoryTask, pk=task_id)
        
        # 读取文件内容
        file_name = file.name.lower()
        content = ''
        
        if file_name.endswith(('.xlsx', '.xls')):
            # Excel文件
            import openpyxl
            from io import BytesIO
            
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    row_text = ' '.join([str(cell) for cell in row if cell])
                    content += row_text + '\n'
                    
        elif file_name.endswith('.csv'):
            # CSV文件
            content = file.read().decode('utf-8', errors='ignore')
            
        elif file_name.endswith(('.png', '.jpg', '.jpeg', '.bmp')):
            # 图片 - 使用两步AI识别
            image_bytes = file.read()
            app_url = get_config_value('app_url', 'http://127.0.0.1:8000')
            return api_ai_parse_image(request, image_bytes, task, app_url)
            
        else:
            return JsonResponse({'success': False, 'message': '不支持的文件格式'})
        
        # 调用AI解析
        result, error = parse_asset_numbers_with_ai(content)
        
        if error:
            return JsonResponse({'success': False, 'message': error})
        
        return process_ai_parse_result(result, task, request.user)
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


def api_ai_parse_image(request, image_bytes, task, base_url='http://127.0.0.1:8000', image_input_method='http_url'):
    """使用Vision API解析图片（多模态理解），两步调用：先识别再整理JSON"""
    if not is_llm_enabled():
        return JsonResponse({'success': False, 'message': '未启用LLM或未配置API Key，请在系统设置-模型设置中配置'})
    
    try:
        ai_content = call_llm_vision_two_step(
            image_bytes=image_bytes,
            log_callback=None,
            base_url=base_url,
            parse_photo_only=False,
            image_input_method=image_input_method
        )
        
        data, error = parse_ai_json_response(ai_content)
        if data:
            return process_ai_parse_result(data, task, request.user)
        else:
            return JsonResponse({'success': False, 'message': f'AI返回格式异常: {error}'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'图片解析失败: {str(e)}'})


def process_ai_parse_result(result, task, user):
    """处理AI解析结果"""
    # 展开合并写法的编号
    all_asset_numbers = []
    for no in result.get('asset_numbers', []):
        expanded = expand_asset_no_pattern(no)
        all_asset_numbers.extend(expanded)
    
    all_card_numbers = result.get('card_numbers', [])
    
    # 查找匹配的设备
    found_devices = Device.objects.filter(
        Q(asset_no__in=all_asset_numbers) | Q(asset_card_no__in=all_card_numbers)
    ).exclude(status='scrapped').distinct()
    
    # 排除已添加的
    existing_ids = set(InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True))
    new_devices = [d for d in found_devices if d.id not in existing_ids]
    
    # 添加到任务
    current_max_order = InventoryTaskDevice.objects.filter(task=task).count()
    task_devices = []
    for i, device in enumerate(new_devices):
        task_devices.append(InventoryTaskDevice(
            task=task,
            device=device,
            status='pending',
            sort_order=current_max_order + i + 1,
            added_by=user,
        ))
    InventoryTaskDevice.objects.bulk_create(task_devices)
    
    # 更新计数
    task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
    task.save()
    
    # 未找到的编号
    found_asset_nos = set(found_devices.values_list('asset_no', flat=True))
    found_card_nos = set(found_devices.values_list('asset_card_no', flat=True))
    not_found = [no for no in all_asset_numbers if no not in found_asset_nos]
    
    return JsonResponse({
        'success': True,
        'message': f'已添加 {len(new_devices)} 台设备',
        'added_count': len(new_devices),
        'not_found': not_found,
        'parsed_count': len(all_asset_numbers),
        'total_count': task.device_count,
    })


# ==================== 新版导入设备 API ====================

def expand_asset_numbers_for_inventory(base_no, count):
    """展开合并写法的资产编号（与update_card_no相同）"""
    results = []
    
    # 处理 / 分隔的格式: XACD-Z-001-001-001/002/003
    slash_match = re.match(r'^(.+)-(\d+)/(.+)$', base_no)
    if slash_match:
        prefix = slash_match.group(1)
        first_num = slash_match.group(2)
        rest_nums = slash_match.group(3).split('/')
        results.append(f"{prefix}-{first_num}")
        for num in rest_nums:
            results.append(f"{prefix}-{num}")
        return results[:count] if count else results
    
    # 处理 ~ 连号格式: XACD-Z-001-001-006~009
    range_match = re.match(r'^(.+)-(\d+)~(\d+)$', base_no)
    if range_match:
        prefix = range_match.group(1)
        start = int(range_match.group(2))
        end = int(range_match.group(3))
        num_len = len(range_match.group(2))
        for i in range(start, end + 1):
            results.append(f"{prefix}-{str(i).zfill(num_len)}")
        return results[:count] if count else results
    
    # 普通格式，直接返回
    return [base_no]


def parse_inventory_excel(file_content):
    """解析Excel文件，返回资产编号和卡片编号的映射"""
    import openpyxl
    from io import BytesIO
    
    wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
    ws = wb.active
    
    # 1. 扫描前10行找到列名行（支持合并单元格）
    header_row = None
    asset_no_col = None
    card_no_col = None
    quantity_col = None
    
    merged_cells_map = {}
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row
        min_col = merged_range.min_col
        first_cell = ws.cell(row=min_row, column=min_col)
        first_value = str(first_cell.value).strip().lower() if first_cell.value else ''
        for row in range(min_row, min_row + (merged_range.max_row - merged_range.min_row + 1)):
            for col in range(min_col, min_col + (merged_range.max_col - merged_range.min_col + 1)):
                merged_cells_map[(row, col)] = (min_row, min_col, first_value)
    
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_values = []
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if (row_idx, col_idx) in merged_cells_map:
                _, _, merged_value = merged_cells_map[(row_idx, col_idx)]
                row_values.append(merged_value)
            else:
                row_values.append(str(cell_value).strip().lower() if cell_value else '')
        
        has_asset = any('资产编号' in v or '资产编码' in v for v in row_values)
        has_card = any('卡片编号' in v or '卡片码' in v for v in row_values)
        
        if has_asset and has_card:
            header_row = row_idx
            for idx, v in enumerate(row_values):
                if '资产编号' in v or '资产编码' in v:
                    asset_no_col = idx
                elif '卡片编号' in v or '卡片码' in v:
                    card_no_col = idx
                elif '数量' in v or '资产数量' in v:
                    quantity_col = idx
            break
    
    if asset_no_col is None or card_no_col is None:
        return None, '未找到资产编号列或卡片编号列'
    
    # 2. 从列名行下一行开始读取数据
    data_start_row = header_row + 1
    
    # 3. 构建映射
    excel_mappings = {}
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        if len(row) > max(asset_no_col, card_no_col):
            asset_no = str(row[asset_no_col]).strip() if row[asset_no_col] else ''
            card_no = str(row[card_no_col]).strip() if row[card_no_col] else ''
            quantity = None
            if quantity_col is not None and quantity_col < len(row) and row[quantity_col]:
                try:
                    quantity = int(float(row[quantity_col]))
                except:
                    pass
            
            if asset_no:
                expanded = expand_asset_numbers_for_inventory(asset_no, quantity)
                for no in expanded:
                    excel_mappings[no] = card_no
    
    return excel_mappings, None


def parse_excel_with_ai_assist(file_content):
    """使用AI解析Excel/CSV内容（转为文本发送），使用系统设置中配置的大模型"""
    import openpyxl
    from io import BytesIO
    import csv
    
    if not is_llm_enabled():
        return None, '未启用LLM或未配置API Key，请在系统设置-模型设置中配置'
    
    content_text = ''
    file_name = 'file.xlsx'
    
    if hasattr(file_content, 'name'):
        file_name = file_content.name.lower()
    elif isinstance(file_content, bytes):
        if file_content[:4] == b'%PDF':
            return None, '不支持PDF文件'
    
    if file_name.endswith('.csv'):
        if hasattr(file_content, 'read'):
            content = file_content.read().decode('utf-8')
        else:
            content = file_content.decode('utf-8')
        reader = csv.reader(content.splitlines())
        for row in reader:
            content_text += ' | '.join([str(cell) for cell in row if cell]) + '\n'
    else:
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        ws = wb.active
        for row in ws.iter_rows(max_row=50, values_only=True):
            content_text += ' | '.join([str(cell) if cell else '' for cell in row]) + '\n'
    
    prompt = f"""请从以下Excel/CSV表格数据中提取资产编号和卡片编号的对应关系。

规则：
1. 资产编号列可能包含：资产编号、资产编码
2. 卡片编号列可能包含：卡片编号、卡片码
3. 数量列可能包含：资产数量、数量
4. 资产编号可能包含合并写法如 XACD-Z-001-001-001/002/003，需要拆解
5. 连号写法如 006~009 也需要拆解
6. 返回JSON格式：
{{
  "mappings": [
    {{"asset_no": "XACD-Z-001-001-001", "card_no": "CARD-001"}},
    {{"asset_no": "XACD-Z-001-001-002", "card_no": "CARD-001"}}
  ]
}}

请直接分析数据内容，不要询问用户。

表格数据：
{content_text}"""
    
    try:
        ai_content = call_llm([
            {"role": "system", "content": "你是一个专业的数据解析助手，擅长从表格数据中提取设备资产编号和卡片编号的对应关系。"},
            {"role": "user", "content": prompt}
        ])
        
        json_match = re.search(r'\{[\s\S]*\}', ai_content)
        
        if json_match:
            return json.loads(json_match.group()), None
        return None, 'AI返回格式异常'
    except Exception as e:
        return None, f'AI解析失败: {str(e)}'


def call_ai_for_inventory_import(excel_data):
    """调用AI解析资产编号，使用系统设置中配置的大模型"""
    prompt = f"""请从以下Excel表格数据中提取资产编号和卡片编号的对应关系。

规则：
1. 自动识别包含"资产编号"、"卡片编号"、"资产数量"的列
2. 资产编号格式类似: XACD-Z-001-001-001
3. 合并写法如 XACD-Z-001-001-001/002/003 表示多个编号
4. 连号写法如 XACD-Z-001-001-006~009 表示连续编号
5. 结合资产数量字段判断拆解数量
6. 返回JSON格式，包含展开后的资产编号和对应卡片编号：
{{
  "mappings": [
    {{"asset_no": "XACD-Z-001-001-001", "card_no": "CARD-001"}},
    {{"asset_no": "XACD-Z-001-001-002", "card_no": "CARD-001"}},
    ...
  ]
}}

表格数据：
{excel_data}"""
    
    try:
        ai_content = call_llm([
            {"role": "system", "content": "你是一个专业的数据解析助手，擅长从表格数据中提取设备资产编号和卡片编号的对应关系。"},
            {"role": "user", "content": prompt}
        ])
        
        json_match = re.search(r'\{[\s\S]*\}', ai_content)
        if json_match:
            return json.loads(json_match.group())
        return None
    except Exception as e:
        return None


def process_inventory_import_task(task_id, file_content, task, user, parse_photo_only, file_name='unknown', base_url='http://127.0.0.1:8000', image_input_method='http_url'):
    """后台处理导入任务"""
    from io import BytesIO
    import openpyxl
    
    progress = inventory_import_progress[task_id]
    
    try:
        # 使用传入的文件名，兜底用magic bytes检测
        if file_name in ('unknown', 'file.xlsx', 'file.xls', 'file'):
            if isinstance(file_content, bytes):
                if file_content[:8] == b'\x89PNG\r\n\x1a\n':
                    file_name = 'image.png'
                elif file_content[:3] == b'\xFF\xD8\xFF':
                    file_name = 'image.jpg'
                elif file_content[:4] == b'PK\x03\x04':
                    file_name = 'file.xlsx'
                elif file_content[:2] == b'\xD0\xCF':
                    file_name = 'file.xls'
        
        # 判断文件类型
        if file_name.endswith(('.png', '.jpg', '.jpeg', '.bmp')):
            # 图片文件 - 使用两步AI识别
            inventory_add_log(task_id, 'info', f'开始解析图片中的设备... (文件: {file_name})')
            
            if not is_llm_enabled():
                progress['status'] = 'completed'
                inventory_add_log(task_id, 'error', 'LLM未启用或未配置API Key，请在系统设置-模型设置中配置')
                return
            
            config = get_llm_config()
            model_name = config['model_name']
            inventory_add_log(task_id, 'ai', f'调用AI解析图片，使用模型: {model_name}，图片大小: {len(file_content)} bytes')
            
            try:
                ai_content = call_llm_vision_two_step(
                    image_bytes=file_content,
                    log_callback=lambda msg: inventory_add_log(task_id, 'ai', msg),
                    base_url=base_url,
                    parse_photo_only=parse_photo_only,
                    image_input_method=image_input_method
                )
            except Exception as e:
                inventory_add_log(task_id, 'error', f'AI API调用失败: {str(e)}')
                progress['status'] = 'completed'
                return
            
            if not ai_content:
                inventory_add_log(task_id, 'error', 'AI返回内容为空，请检查API配置或模型是否支持图片解析')
                progress['status'] = 'completed'
                progress['errors'].append('AI返回内容为空')
                return
            
            data, error = parse_ai_json_response(ai_content)
            
            if data:
                asset_nos = data.get('asset_numbers', [])
                card_nos = data.get('card_numbers', [])
                
                # 展开合并写法
                all_asset_numbers = []
                for no in asset_nos:
                    expanded = expand_asset_numbers_for_inventory(no, None)
                    all_asset_numbers.extend(expanded)
                
                inventory_add_log(task_id, 'success', f'AI解析成功，识别到 {len(all_asset_numbers)} 个资产编号')
                
                # 查找系统设备
                found_devices = Device.objects.filter(
                    Q(asset_no__in=all_asset_numbers) | Q(asset_card_no__in=card_nos)
                ).exclude(status='scrapped').select_related('category', 'location', 'user', 'department')
                
                # 构建设备列表（两种模式都需要，供前端展示）
                device_list = []
                for d in found_devices:
                    device_list.append({
                        'id': d.id,
                        'asset_no': d.asset_no,
                        'asset_card_no': d.asset_card_no or '',
                        'name': d.name,
                        'category': d.category.name if d.category else '',
                        'location': d.location.get_full_path() if d.location else d.location_text or '',
                        'device_no': d.device_no or '',
                        'model': d.model or '',
                        'secret_level': d.get_secret_level_display(),
                        'department': d.department.name if d.department else '',
                        'user': d.user.realname if d.user else '',
                        'status': d.get_status_display(),
                        'is_fixed': d.is_fixed,
                    })
                progress['device_list'] = device_list
                
                inventory_add_log(task_id, 'success', f'在系统中找到 {len(device_list)} 台匹配设备')
                
                # 自动添加设备到待盘点列表（两种模式都自动添加）
                existing_device_ids = set(InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True))
                added_count = 0
                skipped_count = 0
                
                for d in found_devices:
                    if d.id in existing_device_ids:
                        skipped_count += 1
                        inventory_add_log(task_id, 'warning', f'设备 {d.asset_no}：已在任务中，跳过')
                    else:
                        InventoryTaskDevice.objects.create(
                            task=task,
                            device=d,
                            status='pending',
                            sort_order=added_count + skipped_count + 1,
                            added_by=user,
                        )
                        added_count += 1
                        inventory_add_log(task_id, 'success', f'设备 {d.asset_no}：匹配成功，已添加到任务')
                
                # 更新任务设备计数
                task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
                task.save()
                
                progress['added_count'] = added_count
                progress['skipped_count'] = skipped_count
                inventory_add_log(task_id, 'success', f'导入待盘点完成！已添加 {added_count} 个，已存在 {skipped_count} 个')
            else:
                inventory_add_log(task_id, 'error', f'AI返回格式异常: {error}')
            
            progress['status'] = 'completed'
            return
        
        # Excel文件处理
        inventory_add_log(task_id, 'info', '开始读取Excel文件...')
        
        excel_mappings, error = parse_inventory_excel(file_content)
        
        if error:
            # 本地识别失败，尝试AI解析
            inventory_add_log(task_id, 'warning', f'本地解析失败: {error}，尝试AI解析...')
            
            ai_result, ai_error = parse_excel_with_ai_assist(file_content)
            
            if ai_error or not ai_result or 'mappings' not in ai_result:
                # AI也失败，返回需要手动选择
                progress['status'] = 'need_manual'
                progress['error'] = error
                progress['ai_error'] = ai_error
                inventory_add_log(task_id, 'error', f'本地和AI解析都失败: {ai_error or error}')
                return
            
            # 使用AI返回的映射数据
            ai_mappings = {item['asset_no']: item['card_no'] for item in ai_result['mappings']}
            inventory_add_log(task_id, 'success', f'AI解析成功，返回 {len(ai_mappings)} 条映射记录')
            excel_mappings = ai_mappings
        
        inventory_add_log(task_id, 'success', f'Excel文件读取成功，共 {len(excel_mappings)} 条资产编号映射')
        
        # 获取AI配置
        if is_llm_enabled():
            config = get_llm_config()
            model_name = config['model_name']
            inventory_add_log(task_id, 'ai', f'调用AI解析数据，模型: {model_name}，发送 {len(excel_mappings)} 条映射记录...')
            
            excel_text = '\n'.join([f"{k}|{v}" for k, v in excel_mappings.items()])
            ai_result = call_ai_for_inventory_import(excel_text)
            
            if ai_result and 'mappings' in ai_result:
                ai_mappings = {item['asset_no']: item['card_no'] for item in ai_result['mappings']}
                inventory_add_log(task_id, 'success', f'AI解析成功，返回 {len(ai_mappings)} 条映射记录')
                excel_mappings = ai_mappings
            else:
                inventory_add_log(task_id, 'warning', 'AI解析失败或返回格式错误，使用本地解析结果')
        
        # 获取任务中已存在的设备
        existing_device_ids = set(InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True))
        
        # 遍历系统设备进行匹配
        inventory_add_log(task_id, 'info', '开始匹配系统设备...')
        
        all_devices = Device.objects.exclude(status='scrapped').select_related('category', 'location', 'user', 'department')
        total = all_devices.count()
        progress['total'] = total
        
        added_count = 0
        skipped_count = 0
        not_found = []
        
        for device in all_devices:
            progress['current'] += 1
            
            if device.asset_no in excel_mappings:
                # 匹配成功
                if device.id in existing_device_ids:
                    # 已在任务中
                    skipped_count += 1
                    inventory_add_log(task_id, 'warning', f'设备 {device.asset_no}：已在任务中，跳过')
                else:
                    # 添加到任务
                    InventoryTaskDevice.objects.create(
                        task=task,
                        device=device,
                        status='pending',
                        sort_order=progress['current'],
                        added_by=user,
                    )
                    added_count += 1
                    inventory_add_log(task_id, 'success', f'设备 {device.asset_no}：匹配成功，已添加到任务')
            else:
                # 未匹配
                not_found.append(device.asset_no)
            
            if progress['current'] % 10 == 0:
                time.sleep(0.05)
        
        # 更新任务设备计数
        task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
        task.save()
        
        progress['added_count'] = added_count
        progress['skipped_count'] = skipped_count
        progress['not_found'] = not_found[:100]  # 限制返回数量
        
        progress['status'] = 'completed'
        inventory_add_log(task_id, 'success', f'处理完成！共处理 {total} 个设备，已添加 {added_count} 个，已存在 {skipped_count} 个')
        
    except Exception as e:
        progress['status'] = 'completed'
        progress['errors'].append(str(e))
        inventory_add_log(task_id, 'error', f'处理失败: {str(e)}')


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_import_devices(request, task_id):
    """导入设备API"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    file = request.FILES.get('file')
    if not file:
        return JsonResponse({'success': False, 'message': '请上传文件'})
    
    parse_photo_only = request.POST.get('parse_photo_only', 'false').lower() == 'true'
    image_input_method = request.POST.get('image_input_method', 'http_url')
    
    # 获取原始文件名
    file_name = file.name.lower()
    
    # 获取app_url配置，用于构造图片HTTP URL
    app_url = get_config_value('app_url', 'http://127.0.0.1:8000')
    
    # 所有情况统一走后台异步处理（包括图片+勾选仅解析照片）
    task_id_str = str(uuid.uuid4())[:8]
    file_content = file.read()
    
    inventory_import_progress[task_id_str] = {
        'status': 'processing',
        'total': 0,
        'current': 0,
        'added_count': 0,
        'skipped_count': 0,
        'errors': [],
        'not_found': [],
        'logs': [],
        'device_list': None,
        'file_content': file_content,
    }
    
    thread = threading.Thread(
        target=process_inventory_import_task,
        args=(task_id_str, file_content, task, request.user, parse_photo_only, file_name, app_url, image_input_method)
    )
    thread.daemon = True
    thread.start()
    
    return JsonResponse({'success': True, 'task_id': task_id_str})


def handle_image_parse_only(task, file, user, base_url='http://127.0.0.1:8000', image_input_method='http_url'):
    """处理图片仅解析模式（人工勾选），使用系统设置中配置的大模型"""
    
    if not is_llm_enabled():
        return JsonResponse({'success': False, 'message': '未启用LLM或未配置API Key，请在系统设置-模型设置中配置'})
    
    try:
        image_bytes = file.read()
        
        ai_content = call_llm_vision_two_step(
            image_bytes=image_bytes,
            log_callback=None,  # 同步场景无进度面板
            base_url=base_url,
            parse_photo_only=True,
            image_input_method=image_input_method
        )
        
        if not ai_content:
            return JsonResponse({'success': False, 'message': 'AI返回内容为空，请检查API配置或模型是否支持图片解析'})
        
        data, error = parse_ai_json_response(ai_content)
        
        if not data:
            return JsonResponse({'success': False, 'message': f'AI返回格式异常: {error}'})
        
        asset_nos = data.get('asset_numbers', [])
        card_nos = data.get('card_numbers', [])
        
        # 展开合并写法
        all_asset_numbers = []
        for no in asset_nos:
            expanded = expand_asset_numbers_for_inventory(no, None)
            all_asset_numbers.extend(expanded)
        
        # 查找系统设备
        found_devices = Device.objects.filter(
            Q(asset_no__in=all_asset_numbers) | Q(asset_card_no__in=card_nos)
        ).exclude(status='scrapped').select_related('category', 'location', 'user', 'department')
        
        # 获取已存在的设备
        existing_ids = set(InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True))
        
        device_list = []
        for d in found_devices:
            if d.id not in existing_ids:
                device_list.append({
                    'id': d.id,
                    'asset_no': d.asset_no,
                    'asset_card_no': d.asset_card_no or '',
                    'name': d.name,
                    'category': d.category.name if d.category else '',
                    'location': d.location.get_full_path() if d.location else d.location_text or '',
                    'device_no': d.device_no or '',
                    'model': d.model or '',
                    'secret_level': d.get_secret_level_display(),
                    'department': d.department.name if d.department else '',
                    'user': d.user.realname if d.user else '',
                    'status': d.get_status_display(),
                    'is_fixed': d.is_fixed,
                })
        
        return JsonResponse({
            'success': True,
            'device_list': device_list,
            'parsed_count': len(all_asset_numbers),
            'found_count': len(device_list),
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'图片解析失败: {str(e)}'})


@login_required
def api_import_progress(request, task_id):
    """导入进度查询API"""
    task_id_str = request.GET.get('task_id')
    if task_id_str and task_id_str in inventory_import_progress:
        data = inventory_import_progress[task_id_str].copy()
        # 排除无法序列化的字段
        data.pop('file_content', None)
        return JsonResponse(data)
    return JsonResponse({'status': 'not_found'})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_manual_parse_excel(request, task_id):
    """手动指定列解析Excel"""
    import base64
    import openpyxl
    from io import BytesIO
    
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        data = json.loads(request.body)
        
        task_id_str = data.get('task_id')
        asset_col = int(data.get('asset_col', 0))
        card_col = int(data.get('card_col', 1))
        qty_col = data.get('qty_col')
        if qty_col is not None:
            qty_col = int(qty_col)
            if qty_col < 0:
                qty_col = None
        
        if task_id_str and task_id_str in inventory_import_progress:
            progress = inventory_import_progress[task_id_str]
            file_content = progress.get('file_content')
        else:
            return JsonResponse({'success': False, 'message': '任务已过期，请重新上传文件'})
        
        if not file_content:
            return JsonResponse({'success': False, 'message': '文件内容丢失'})
        
        inventory_add_log(task_id_str, 'info', '手动解析：开始解析Excel...')
        
        wb = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
        ws = wb.active
        
        excel_mappings = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > max(asset_col, card_col):
                asset_no = str(row[asset_col]).strip() if row[asset_col] else ''
                card_no = str(row[card_col]).strip() if row[card_col] else ''
                quantity = None
                if qty_col is not None and qty_col < len(row) and row[qty_col]:
                    try:
                        quantity = int(float(row[qty_col]))
                    except:
                        pass
                
                if asset_no:
                    expanded = expand_asset_numbers_for_inventory(asset_no, quantity)
                    for no in expanded:
                        excel_mappings[no] = card_no
        
        inventory_add_log(task_id_str, 'success', f'手动解析成功，共 {len(excel_mappings)} 条资产编号映射')
        
        existing_device_ids = set(InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True))
        
        all_devices = Device.objects.exclude(status='scrapped').select_related('category', 'location', 'user', 'department')
        total = all_devices.count()
        
        added_count = 0
        skipped_count = 0
        not_found = []
        
        for device in all_devices:
            if device.asset_no in excel_mappings:
                if device.id in existing_device_ids:
                    skipped_count += 1
                    inventory_add_log(task_id_str, 'warning', f'设备 {device.asset_no}：已在任务中，跳过')
                else:
                    InventoryTaskDevice.objects.create(
                        task=task,
                        device=device,
                        status='pending',
                        sort_order=added_count + skipped_count + 1,
                        added_by=request.user,
                    )
                    added_count += 1
                    inventory_add_log(task_id_str, 'success', f'设备 {device.asset_no}：匹配成功，已添加到任务')
            else:
                not_found.append(device.asset_no)
        
        task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
        task.save()
        
        inventory_add_log(task_id_str, 'success', f'处理完成！已添加 {added_count} 个，已存在 {skipped_count} 个')
        
        return JsonResponse({
            'success': True,
            'added_count': added_count,
            'skipped_count': skipped_count,
            'not_found_count': len(not_found[:100]),
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'解析失败: {str(e)}'})


# ==================== 盘点执行 ====================

@login_required
@permission_required('inventory_execute')
def task_execute(request, pk):
    """盘点执行页面"""
    task = get_object_or_404(InventoryTask.objects.select_related('created_by', 'assignee'), pk=pk)
    
    # 动态盘点任务自动同步设备列表
    if task.task_type == 'dynamic' and task.status == 'in_progress':
        sync_dynamic_task_devices(task)
    
    # 待盘点设备 - 有位置的在前，无位置的在后，按位置排序
    pending_devices = InventoryTaskDevice.objects.filter(
        task=task, status='pending'
    ).select_related(
        'device__category', 'device__location', 'device__user', 'device__department', 'device__workstation'
    ).annotate(
        has_location=Case(
            When(device__location_text='', then=1),
            default=0,
            output_field=IntegerField(),
        )
    ).order_by('has_location', 'device__location_text', 'device__id')
    
    # 已盘点设备 - 按盘点时间倒序，新盘点的在前
    checked_devices = InventoryTaskDevice.objects.filter(
        task=task, status='checked'
    ).select_related(
        'device__category', 'device__location', 'device__user', 'device__department', 'device__workstation'
    ).order_by('-checked_at')
    
    # 统计
    total_count = InventoryTaskDevice.objects.filter(task=task).count()
    pending_count = pending_devices.count()
    checked_count = total_count - pending_count
    
    # 获取位置树（用于地图选位）
    from apps.assets.views import get_location_tree_data
    location_tree = get_location_tree_data()
    
    # 获取部门和用户列表（用于分配）
    departments = Department.objects.all()
    
    return render(request, 'inventory/task_execute.html', {
        'task': task,
        'pending_devices': pending_devices,
        'checked_devices': checked_devices,
        'total_count': total_count,
        'pending_count': pending_count,
        'checked_count': checked_count,
        'location_tree': json.dumps(location_tree),
        'departments': departments,
    })


@login_required
def api_device_detail(request, task_id, device_id):
    """获取设备详情（盘点时展示）"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    device = get_object_or_404(
        Device.objects.select_related('category', 'location', 'user', 'department', 'workstation'),
        pk=device_id
    )
    
    data = {
        'id': device.id,
        'asset_no': device.asset_no,
        'device_no': device.device_no or '',
        'name': device.name,
        'model': device.model or '',
        'serial_no': device.serial_no or '',
        'brand': device.brand or '',
        'category': device.category.name if device.category else '',
        'category_id': device.category_id,
        'secret_level': device.get_secret_level_display(),
        'status': device.get_status_display(),
        'status_code': device.status,
        
        # 使用信息
        'department': device.department.name if device.department else '',
        'department_id': device.department_id,
        'user': device.user.realname if device.user else '',
        'user_id': device.user_id,
        'location': device.location.get_full_path() if device.location else '',
        'location_id': device.location_id,
        'workstation': device.workstation.workstation_code if device.workstation else '',
        'workstation_id': device.workstation_id,
        
        # 网络信息
        'mac_address': device.mac_address or '',
        'ip_address': device.ip_address or '',
        
        # 设备信息
        'os_name': device.os_name or '',
        'disk_serial': device.disk_serial or '',
        'purchase_date': str(device.purchase_date) if device.purchase_date else '',
        'enable_date': str(device.enable_date) if device.enable_date else '',
        'install_date': str(device.install_date) if device.install_date else '',
        
        # 其他信息
        'is_fixed': device.is_fixed,
        'asset_card_no': device.asset_card_no or '',
        'is_secret': device.is_secret,
        'secret_category': device.secret_category or '',
        'purpose': device.purpose or '',
        'remarks': device.remarks or '',
        
        # 照片
        'photo_url': device.photo.url if device.photo else '',
        'photo_updated_at': str(device.photo_updated_at) if device.photo_updated_at else '',
    }
    
    # 如果设备有关联工位，获取地图信息
    ws_location_id = None
    if device.workstation_id:
        ws_location_id = device.workstation.location_id
    elif device.location_id:
        loc = device.location
        while loc and loc.level != 3:
            loc = loc.parent
        if loc:
            ws_location_id = loc.id
    
    data['map_location_id'] = ws_location_id
    
    return JsonResponse({'success': True, 'device': data})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_check_device(request, task_id):
    """确认盘点设备"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        # 支持JSON和FormData两种格式
        if request.content_type and 'multipart/form-data' in request.content_type:
            # FormData格式（支持文件上传）
            task_device_id = request.POST.get('task_device_id')
            device_id = request.POST.get('device_id')
            location_status = request.POST.get('location_status', 'in_place')
            asset_status = request.POST.get('asset_status', 'normal')
            remarks = request.POST.get('remarks', '')
            source = request.POST.get('source', 'manual')
            photo_file = request.FILES.get('photo')
        else:
            # JSON格式
            data = json.loads(request.body)
            task_device_id = data.get('task_device_id')
            device_id = data.get('device_id')
            location_status = data.get('location_status', 'in_place')
            asset_status = data.get('asset_status', 'normal')
            remarks = data.get('remarks', '')
            source = data.get('source', 'manual')
            photo_file = None
        
        task_device = get_object_or_404(InventoryTaskDevice, pk=task_device_id, task=task)
        device = task_device.device
        
        # 如果有照片，上传并更新设备照片
        if photo_file:
            save_photo_with_asset_no(device, photo_file)
            device.photo_updated_at = timezone.now()
            device.save(update_fields=['photo', 'photo_updated_at'])
        
        # 创建盘点记录
        record = InventoryRecord.objects.create(
            task=task,
            device=device,
            task_device=task_device,
            device_status=device.status,
            location_status=location_status,
            asset_status=asset_status,
            remarks=remarks,
            checked_by=request.user,
            source=source,
        )
        
        # 如果有备注，更新设备备注字段
        if remarks:
            device.remarks = remarks
            device.save(update_fields=['remarks'])
        
        # 更新任务设备状态
        task_device.status = 'checked'
        task_device.checked_at = timezone.now()
        task_device.save()
        
        # 更新任务计数
        task.checked_count = InventoryTaskDevice.objects.filter(task=task, status='checked').count()
        task.save()
        
        # 如果设备未找到，更新设备状态
        if location_status == 'not_found':
            device.status = 'fault'
            device.fault_reason = '盘点未找到'
            device.save()
        
        return JsonResponse({
            'success': True,
            'message': f'设备 {device.asset_no} 盘点完成',
            'checked_count': task.checked_count,
            'pending_count': InventoryTaskDevice.objects.filter(task=task, status='pending').count(),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_revert_check(request, task_id):
    """退回已盘点设备到待盘点"""
    task = get_object_or_404(InventoryTask, pk=task_id)

    try:
        data = json.loads(request.body)
        task_device_id = data.get('task_device_id')

        task_device = get_object_or_404(InventoryTaskDevice, pk=task_device_id, task=task)

        if task_device.status != 'checked':
            return JsonResponse({'success': False, 'message': '该设备未盘点'})

        # 删除盘点记录
        InventoryRecord.objects.filter(task=task, task_device=task_device).delete()

        # 重置设备状态
        task_device.status = 'pending'
        task_device.checked_at = None
        task_device.save()

        # 更新任务计数
        task.checked_count = InventoryTaskDevice.objects.filter(task=task, status='checked').count()
        task.save()

        return JsonResponse({
            'success': True,
            'message': f'设备 {task_device.device.asset_no} 已退回待盘点',
            'checked_count': task.checked_count,
            'pending_count': InventoryTaskDevice.objects.filter(task=task, status='pending').count(),
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_scan_check(request, task_id):
    """扫码盘点"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        data = json.loads(request.body)
        qrcode_content = data.get('qrcode', '')
        
        # 解析二维码URL，提取资产编号
        # 格式如: https://itams.chidafeiji.com/assets/view/XACD-Z-001-001-001/
        asset_no = None
        
        if '/assets/view/' in qrcode_content:
            # 提取资产编号
            match = re.search(r'/assets/view/([^/]+)', qrcode_content)
            if match:
                asset_no = match.group(1)
        elif '/device/scan/' in qrcode_content:
            # 兼容旧格式
            match = re.search(r'/device/scan/(\d+)', qrcode_content)
            if match:
                device_id = match.group(1)
                device = Device.objects.filter(id=device_id).first()
                if device:
                    asset_no = device.asset_no
        
        if not asset_no:
            return JsonResponse({'success': False, 'message': '无法解析二维码内容'})
        
        # 查找设备
        device = Device.objects.filter(asset_no=asset_no).exclude(status='scrapped').first()
        if not device:
            return JsonResponse({'success': False, 'message': f'未找到设备: {asset_no}'})
        
        # 查找任务设备
        task_device = InventoryTaskDevice.objects.filter(
            task=task, device=device, status='pending'
        ).first()
        
        if not task_device:
            # 动态盘点任务自动添加新设备到任务中
            if task.task_type == 'dynamic' and task.status == 'in_progress':
                # 检查是否已在任务中（可能已盘点）
                existing_task_device = InventoryTaskDevice.objects.filter(
                    task=task, device=device
                ).first()
                
                if existing_task_device:
                    if existing_task_device.status == 'checked':
                        return JsonResponse({
                            'success': False,
                            'message': f'设备 {asset_no} 已盘点完成',
                            'device_already_checked': True,
                        })
                    task_device = existing_task_device
                else:
                    # 添加新设备到任务
                    current_max_order = InventoryTaskDevice.objects.filter(task=task).count()
                    task_device = InventoryTaskDevice.objects.create(
                        task=task,
                        device=device,
                        status='pending',
                        sort_order=current_max_order + 1,
                        added_by=request.user,
                    )
                    task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
                    task.save(update_fields=['device_count'])
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'设备 {asset_no} 不在待盘点列表中',
                    'device_not_in_task': True,
                    'device_id': device.id,
                    'asset_no': asset_no,
                })
        
        # 返回设备信息，准备盘点
        return JsonResponse({
            'success': True,
            'message': f'已找到设备: {asset_no}',
            'task_device_id': task_device.id,
            'device_id': device.id,
            'asset_no': device.asset_no,
            'name': device.name,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def api_update_device_info(request, task_id):
    """更新设备信息（分配、位置）"""
    task = get_object_or_404(InventoryTask, pk=task_id)
    
    try:
        data = json.loads(request.body)
        device_id = data.get('device_id')
        update_type = data.get('update_type')  # 'assign' 或 'location'
        
        device = get_object_or_404(Device, pk=device_id)
        
        if update_type == 'assign':
            # 更新部门和使用人
            user_id = data.get('user_id')
            department_id = data.get('department_id')
            
            old_dept = device.department.name if device.department else ''
            old_user = device.user.realname if device.user else ''
            
            # 如果传入了user_id，自动获取用户的部门
            if user_id:
                user_obj = User.objects.filter(pk=user_id).first()
                if user_obj:
                    device.user_id = user_id
                    device.department_id = user_obj.department_id  # 自动同步用户所属部门
                    device.status = 'normal'
                else:
                    device.user_id = user_id
                    device.department_id = department_id if department_id else None
                    device.status = 'normal'
            else:
                device.user_id = None
                device.department_id = department_id if department_id else None
                device.status = 'unused'
            
            device.save()
            
            from apps.assets.models import AssetLog
            AssetLog.objects.create(
                device=device,
                user=request.user,
                action='assign',
                field_name='user',
                old_value=f'{old_dept}/{old_user}',
                new_value=f'{device.department.name if device.department else ""}/{device.user.realname if device.user else ""}',
                remarks='盘点时更新'
            )
            
            return JsonResponse({
                'success': True,
                'message': '分配信息已更新',
                'department': device.department.name if device.department else '',
                'user': device.user.realname if device.user else '',
            })
            
        elif update_type == 'location':
            # 更新位置和工位
            location_id = data.get('location_id')
            workstation_id = data.get('workstation_id')
            
            old_location = device.location.get_full_path() if device.location else ''
            old_ws = device.workstation.workstation_code if device.workstation else ''
            old_ws_id = device.workstation_id
            
            device.location_id = location_id if location_id else None
            device.workstation_id = workstation_id if workstation_id else None
            
            # 根据工位更新位置
            if workstation_id:
                ws = device.workstation
                if ws.area_id:
                    device.location_id = ws.area_id
                else:
                    device.location_id = ws.location_id
                ws.status = 'occupied'
                ws.save()
            
            device.save()
            
            # 释放旧工位
            if old_ws_id and old_ws_id != workstation_id:
                from apps.assets.models import Workstation
                old_ws_obj = Workstation.objects.filter(id=old_ws_id).first()
                if old_ws_obj and not old_ws_obj.devices.exists():
                    old_ws_obj.status = 'available'
                    old_ws_obj.save()
            
            from apps.assets.models import AssetLog
            AssetLog.objects.create(
                device=device,
                user=request.user,
                action='transfer',
                field_name='location',
                old_value=f'{old_location}/{old_ws}',
                new_value=f'{device.location.get_full_path() if device.location else ""}/{device.workstation.workstation_code if device.workstation else ""}',
                remarks='盘点时更新'
            )
            
            return JsonResponse({
                'success': True,
                'message': '位置信息已更新',
                'location': device.location.get_full_path() if device.location else '',
                'workstation': device.workstation.workstation_code if device.workstation else '',
            })
        
        else:
            return JsonResponse({'success': False, 'message': '无效的更新类型'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== 辅助函数 ====================

def generate_task_devices(task):
    """生成任务设备列表（全盘/动态盘点）"""
    devices = Device.objects.exclude(status='scrapped')
    
    if task.is_fixed_only:
        devices = devices.filter(is_fixed=True)
    
    # 按位置排序，无位置的按ID排序
    devices = devices.order_by('location_text', 'id')
    
    task_devices = []
    for i, device in enumerate(devices):
        task_devices.append(InventoryTaskDevice(
            task=task,
            device=device,
            status='pending',
            sort_order=i + 1,
            added_by=task.created_by,
        ))
    
    InventoryTaskDevice.objects.bulk_create(task_devices)
    
    task.device_count = len(task_devices)
    task.save()


def sync_dynamic_task_devices(task):
    """同步动态盘点任务的设备列表（只添加新设备，不影响已存在设备）"""
    if task.task_type != 'dynamic' or task.status != 'in_progress':
        return
    
    # 获取系统中所有符合条件的设备
    system_devices = Device.objects.exclude(status='scrapped')
    if task.is_fixed_only:
        system_devices = system_devices.filter(is_fixed=True)
    
    # 获取任务中已存在的设备ID
    existing_device_ids = set(
        InventoryTaskDevice.objects.filter(task=task).values_list('device_id', flat=True)
    )
    
    # 找出新增的设备（系统中有但任务中没有）
    new_devices = system_devices.exclude(id__in=existing_device_ids)
    
    if not new_devices.exists():
        return
    
    # 获取当前最大排序号
    current_max_order = InventoryTaskDevice.objects.filter(task=task).count()
    
    # 添加新增的设备到任务中
    task_devices = []
    for i, device in enumerate(new_devices.order_by('location_text', 'id')):
        task_devices.append(InventoryTaskDevice(
            task=task,
            device=device,
            status='pending',
            sort_order=current_max_order + i + 1,
            added_by=task.created_by,
        ))
    
    InventoryTaskDevice.objects.bulk_create(task_devices)
    
    # 更新设备总数
    task.device_count = InventoryTaskDevice.objects.filter(task=task).count()
    task.save(update_fields=['device_count'])


# ==================== 兼容旧接口 ====================

@login_required
@permission_required('inventory_task')
def plan_list(request):
    """盘点计划列表（兼容）"""
    plans = InventoryPlan.objects.all()
    return render(request, 'inventory/plan_list.html', {'plans': plans})


@login_required
def plan_create(request):
    """创建盘点计划（兼容）"""
    if request.method == 'POST':
        plan = InventoryPlan.objects.create(
            name=request.POST.get('name'),
            plan_type=request.POST.get('plan_type'),
            scope=request.POST.get('scope'),
            location_ids=request.POST.get('location_ids', ''),
            category_ids=request.POST.get('category_ids', ''),
            scheduled_start=request.POST.get('scheduled_start'),
            scheduled_end=request.POST.get('scheduled_end'),
            creator=request.user,
            remarks=request.POST.get('remarks'),
        )
        messages.success(request, f'盘点计划创建成功: {plan.plan_no}')
        return redirect('plan_list')
    
    locations = AssetLocation.objects.all()
    categories = AssetCategory.objects.all()
    return render(request, 'inventory/plan_form.html', {'locations': locations, 'categories': categories})


@login_required
def plan_detail(request, pk):
    """盘点计划详情（兼容）"""
    plan = get_object_or_404(InventoryPlan, pk=pk)
    tasks = plan.tasks.all()
    return render(request, 'inventory/plan_detail.html', {'plan': plan, 'tasks': tasks})


@login_required
def inventory_report(request):
    """盘点报表"""
    from django.core.paginator import Paginator
    from django.db.models import Sum, Count, Q
    
    # 获取筛选参数
    task_type = request.GET.get('task_type', '')
    search = request.GET.get('search', '')
    
    # 查询最新一个已完成的全盘任务
    latest_full_task = InventoryTask.objects.filter(
        status='completed',
        task_type='full'
    ).order_by('-completed_at').first()
    
    if latest_full_task:
        # 从该任务获取统计数据
        total_count = latest_full_task.device_count
        checked_count = latest_full_task.checked_count
        pending_count = total_count - checked_count
        completion_rate = round(checked_count / total_count * 100, 1) if total_count > 0 else 0
        
        # 获取该任务的盘点记录
        task_records = InventoryRecord.objects.filter(task=latest_full_task)
        in_place_count = task_records.filter(location_status='in_place').count()
        moved_count = task_records.filter(location_status='moved').count()
        normal_count = task_records.filter(asset_status='normal').count()
        damaged_count = task_records.filter(asset_status='damaged').count()
    else:
        # 没有已完成的全盘任务，显示为0
        total_count = 0
        checked_count = 0
        pending_count = 0
        completion_rate = 0
        in_place_count = 0
        moved_count = 0
        normal_count = 0
        damaged_count = 0
    
    # 盘点报告列表（已完成任务）
    completed_tasks_qs = InventoryTask.objects.filter(status='completed').select_related('created_by').order_by('-completed_at')
    
    # 筛选
    if task_type:
        completed_tasks_qs = completed_tasks_qs.filter(task_type=task_type)
    if search:
        completed_tasks_qs = completed_tasks_qs.filter(Q(task_no__icontains=search) | Q(name__icontains=search))
    
    # 分页
    paginator = Paginator(completed_tasks_qs, 15)
    page = request.GET.get('page', 1)
    reports = paginator.get_page(page)
    
    # 为每个报告获取盘点人信息
    for task in reports:
        checkers = list(InventoryRecord.objects.filter(
            task=task,
            checked_by__isnull=False
        ).values_list('checked_by__realname', flat=True).distinct())
        # 去重
        seen = set()
        unique_checkers = []
        for name in checkers:
            if name not in seen:
                seen.add(name)
                unique_checkers.append(name)
        task.checkers_text = '、'.join(unique_checkers) if unique_checkers else '-'
    
    context = {
        # 统计数据
        'total_count': total_count,
        'checked_count': checked_count,
        'pending_count': pending_count,
        'completion_rate': completion_rate,
        'in_place_count': in_place_count,
        'moved_count': moved_count,
        'normal_count': normal_count,
        'damaged_count': damaged_count,
        # 当前统计的任务
        'latest_task': latest_full_task,
        # 报告列表
        'reports': reports,
        # 筛选参数
        'task_type': task_type,
        'search': search,
    }
    
    return render(request, 'inventory/report.html', context)


@login_required
def task_report(request, pk):
    """单个任务盘点报告"""
    task = get_object_or_404(InventoryTask.objects.select_related('created_by', 'assignee'), pk=pk)
    
    # 获取盘点记录（按资产编号排序）
    records = task.records.select_related('device', 'device__category', 'device__department', 
                                          'device__user', 'device__location', 'checked_by').order_by('device__asset_no')
    
    # 统计数据
    total_count = task.device_count
    checked_count = task.checked_count
    pending_count = total_count - checked_count
    
    # 位置状态统计
    in_place_count = records.filter(location_status='in_place').count()
    moved_count = records.filter(location_status='moved').count()
    
    # 资产状态统计
    normal_count = records.filter(asset_status='normal').count()
    damaged_count = records.filter(asset_status='damaged').count()
    
    # 获取未盘设备列表（按资产编号排序）
    pending_task_devices = task.task_devices.filter(
        status='pending'
    ).select_related('device', 'device__category', 'device__department', 'device__user', 'device__location').order_by('device__asset_no')
    pending_devices = [td.device for td in pending_task_devices]
    
    # 部门统计
    from django.db.models import Count
    department_stats = records.filter(
        device__department__isnull=False
    ).values(
        'device__department__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # 分类统计
    category_stats = records.filter(
        device__category__isnull=False
    ).values(
        'device__category__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    # 盘点人列表（去重）
    checkers = list(records.filter(
        checked_by__isnull=False
    ).values_list('checked_by__realname', flat=True).distinct())
    # 使用set去重，保持去重后的顺序
    seen = set()
    unique_checkers = []
    for name in checkers:
        if name not in seen:
            seen.add(name)
            unique_checkers.append(name)
    checkers_text = '、'.join(unique_checkers) if unique_checkers else '-'
    
    # 完成率
    completion_rate = round(checked_count / total_count * 100, 1) if total_count > 0 else 0
    
    context = {
        'task': task,
        'records': records[:100],  # 限制显示数量
        'pending_devices': pending_devices,
        'total_count': total_count,
        'checked_count': checked_count,
        'pending_count': pending_count,
        'in_place_count': in_place_count,
        'moved_count': moved_count,
        'normal_count': normal_count,
        'damaged_count': damaged_count,
        'completion_rate': completion_rate,
        'department_stats': department_stats,
        'category_stats': category_stats,
        'checkers_text': checkers_text,
    }
    
    return render(request, 'inventory/task_report.html', context)


@login_required
def task_report_print(request, pk):
    """盘点报告PDF打印 - A4横向"""
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from django.contrib.staticfiles import finders
    
    task = get_object_or_404(InventoryTask.objects.select_related('created_by', 'assignee'), pk=pk)
    
    # 获取盘点记录（按资产编号排序）
    records = list(task.records.select_related('device', 'device__category', 'device__department', 
                                               'device__user', 'device__location', 'checked_by').order_by('device__asset_no'))
    
    # 统计数据
    total_count = task.device_count
    checked_count = task.checked_count
    pending_count = total_count - checked_count
    
    # 位置状态统计
    in_place_count = sum(1 for r in records if r.location_status == 'in_place')
    moved_count = sum(1 for r in records if r.location_status == 'moved')
    
    # 资产状态统计
    normal_count = sum(1 for r in records if r.asset_status == 'normal')
    damaged_count = sum(1 for r in records if r.asset_status == 'damaged')
    
    # 获取未盘设备（按资产编号排序）
    pending_task_devices = list(task.task_devices.filter(
        status='pending'
    ).select_related('device', 'device__category', 'device__department', 'device__user', 'device__location').order_by('device__asset_no'))
    
    # 盘点人列表（去重）
    checkers = list(set([r.checked_by.realname for r in records if r.checked_by]))
    checkers_text = '、'.join(checkers) if checkers else '-'
    
    # 完成率
    completion_rate = round(checked_count / total_count * 100, 1) if total_count > 0 else 0
    
    # 注册中文字体
    font_name = 'Helvetica'
    font_bold_name = 'Helvetica-Bold'
    
    font_path = finders.find('fonts/msyh.ttc')
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont('MsYaHei', font_path))
            font_name = 'MsYaHei'
            font_bold_name = 'MsYaHei'
        except:
            pass
    
    # 创建PDF - A4横向
    buffer = BytesIO()
    page_size = landscape(A4)  # 297mm x 210mm
    c = canvas.Canvas(buffer, pagesize=page_size)
    page_width, page_height = page_size
    
    # 页面边距
    left_margin = 10 * mm
    right_margin = 10 * mm
    top_margin = 10 * mm
    bottom_margin = 15 * mm
    
    # 可用宽度
    usable_width = page_width - left_margin - right_margin
    
    # 状态映射
    status_map = {'normal': '使用', 'fault': '故障', 'scrapped': '报废', 'unused': '闲置'}
    
    def draw_page_footer(c, page_num, total_pages=None):
        c.setFont(font_name, 7)
        footer_text = f'驰达IT资产管理系统 - 盘点报告 - {task.task_no}'
        if total_pages:
            footer_text += f'  第 {page_num}/{total_pages} 页'
        c.drawCentredString(page_width / 2, 5 * mm, footer_text)
    
    def draw_table_header(c, y, headers, col_x, col_widths):
        """绘制表格头"""
        c.setFillColor(colors.Color(0.85, 0.85, 0.85))
        c.rect(left_margin, y - 1.5*mm, usable_width, 5*mm, fill=1)
        c.setFillColor(colors.black)
        c.setFont(font_name, 7)
        for i, h in enumerate(headers):
            c.drawString(col_x[i], y, h)
        return y - 5*mm
    
    def draw_table_row(c, y, values, col_x, col_widths, is_odd=False):
        """绘制表格行"""
        if is_odd:
            c.setFillColor(colors.Color(0.97, 0.97, 0.97))
            c.rect(left_margin, y - 1.5*mm, usable_width, 4.5*mm, fill=1)
        c.setFillColor(colors.black)
        c.setFont(font_name, 6.5)
        for i, val in enumerate(values):
            # 截断过长的文本
            max_chars = int(col_widths[i] / 2)
            text = str(val)[:max_chars] if val else '-'
            c.drawString(col_x[i], y, text)
        return y - 4.5*mm
    
    # ========== 第1页：报告概览 ==========
    page_num = 1
    y = page_height - top_margin
    
    # 标题
    c.setFont(font_name, 14)
    c.drawCentredString(page_width / 2, y, '盘 点 报 告')
    y -= 7 * mm
    
    # 分隔线
    c.setStrokeColor(colors.grey)
    c.line(left_margin, y, page_width - right_margin, y)
    y -= 5 * mm
    
    # 任务信息 - 两列布局
    c.setFont(font_name, 8)
    col1_x = left_margin
    col2_x = left_margin + usable_width / 2
    
    c.drawString(col1_x, y, f'任务编号: {task.task_no}')
    c.drawString(col2_x, y, f'任务类型: {task.get_task_type_display()}')
    y -= 4 * mm
    c.drawString(col1_x, y, f'创建人: {task.created_by.realname if task.created_by else "-"}')
    c.drawString(col2_x, y, f'盘点人: {checkers_text}')
    y -= 4 * mm
    range_text = '在账设备' if task.is_fixed_only else '所有设备'
    c.drawString(col1_x, y, f'盘点范围: {range_text}')
    status_text = '已完成' if task.status == 'completed' else '执行中' if task.status == 'in_progress' else '待执行'
    c.drawString(col2_x, y, f'状态: {status_text}')
    y -= 4 * mm
    c.drawString(col1_x, y, f'创建时间: {task.created_at.strftime("%Y-%m-%d %H:%M")}')
    if task.completed_at:
        c.drawString(col2_x, y, f'完成时间: {task.completed_at.strftime("%Y-%m-%d %H:%M")}')
    y -= 6 * mm
    
    # 分隔线
    c.line(left_margin, y, page_width - right_margin, y)
    y -= 5 * mm
    
    # 统计概览
    c.setFont(font_name, 10)
    c.drawString(left_margin, y, '统计概览')
    y -= 5 * mm
    
    c.setFont(font_name, 8)
    c.drawString(left_margin, y, f'应盘设备: {total_count}')
    c.drawString(left_margin + 40*mm, y, f'已盘设备: {checked_count}')
    c.drawString(left_margin + 80*mm, y, f'未盘设备: {pending_count}')
    c.drawString(left_margin + 120*mm, y, f'完成率: {completion_rate}%')
    y -= 8 * mm
    
    # 位置状态和资产状态统计 - 并排显示
    c.setFont(font_name, 9)
    c.drawString(left_margin, y, '位置状态统计')
    c.drawString(left_margin + 60*mm, y, '资产状态统计')
    y -= 5 * mm
    
    c.setFont(font_name, 8)
    c.drawString(left_margin, y, f'在位: {in_place_count}')
    c.drawString(left_margin + 30*mm, y, f'搬离: {moved_count}')
    c.drawString(left_margin + 60*mm, y, f'正常: {normal_count}')
    c.drawString(left_margin + 90*mm, y, f'损坏: {damaged_count}')
    y -= 8 * mm
    
    # 分隔线
    c.line(left_margin, y, page_width - right_margin, y)
    y -= 5 * mm
    
    # ========== 已盘点设备明细表头 ==========
    c.setFont(font_name, 10)
    c.drawString(left_margin, y, f'已盘点设备明细 (共{len(records)}条)')
    y -= 6 * mm
    
    # 表格列定义
    checked_headers = ['#', '资产编号', '设备编号', '设备名称', '型号', '部门', '使用人', 
                       '位置', '状态', '固资', '卡片编号', '位置状态', '资产状态', '盘点人', '盘点时间']
    checked_widths = [8*mm, 30*mm, 16*mm, 23*mm, 25*mm, 16*mm, 12*mm, 60*mm, 10*mm, 7*mm, 
                      21*mm, 12*mm, 12*mm, 10*mm, 24*mm]
    
    # 计算列x坐标
    checked_col_x = []
    x = left_margin
    for w in checked_widths:
        checked_col_x.append(x)
        x += w
    
    # 绘制表头
    y = draw_table_header(c, y, checked_headers, checked_col_x, checked_widths)
    
    # 绘制数据行
    row_num = 0
    for i, record in enumerate(records):
        if y < bottom_margin + 5*mm:
            # 换页
            draw_page_footer(c, page_num)
            c.showPage()
            page_num += 1
            y = page_height - top_margin
            # 重复表头
            c.setFont(font_name, 10)
            c.drawString(left_margin, y, f'已盘点设备明细 (共{len(records)}条) - 续')
            y -= 6 * mm
            y = draw_table_header(c, y, checked_headers, checked_col_x, checked_widths)
        
        row_num += 1
        device = record.device
        location_status = '在位' if record.location_status == 'in_place' else '搬离'
        asset_status = '正常' if record.asset_status == 'normal' else '损坏'
        
        values = [
            str(row_num),
            device.asset_no or '',
            device.device_no or '',
            device.name or '',
            device.model or '',
            device.department.name if device.department else '',
            device.user.realname if device.user else '',
            device.location_text or '',
            status_map.get(device.status, device.status or ''),
            '是' if device.is_fixed else '否',
            device.asset_card_no or '',
            location_status,
            asset_status,
            record.checked_by.realname if record.checked_by else '',
            record.checked_at.strftime('%m-%d %H:%M') if record.checked_at else ''
        ]
        
        y = draw_table_row(c, y, values, checked_col_x, checked_widths, is_odd=(row_num % 2 == 0))
    
    # ========== 未盘设备列表 ==========
    if pending_task_devices:
        y -= 5 * mm
        if y < bottom_margin + 20*mm:
            draw_page_footer(c, page_num)
            c.showPage()
            page_num += 1
            y = page_height - top_margin
        
        # 分隔线
        c.line(left_margin, y, page_width - right_margin, y)
        y -= 5 * mm
        
        c.setFont(font_name, 10)
        c.drawString(left_margin, y, f'未盘设备列表 (共{len(pending_task_devices)}条)')
        y -= 6 * mm
        
        # 未盘设备表头
        pending_headers = ['#', '资产编号', '设备编号', '设备名称', '型号', '部门', '使用人', 
                          '位置', '状态', '固资', '卡片编号', '盘点状态']
        pending_widths = [8*mm, 30*mm, 18*mm, 23*mm, 25*mm, 18*mm, 14*mm, 60*mm, 12*mm, 7*mm, 21*mm, 14*mm]
        
        pending_col_x = []
        x = left_margin
        for w in pending_widths:
            pending_col_x.append(x)
            x += w
        
        y = draw_table_header(c, y, pending_headers, pending_col_x, pending_widths)
        
        pending_num = 0
        for td in pending_task_devices:
            if y < bottom_margin + 5*mm:
                draw_page_footer(c, page_num)
                c.showPage()
                page_num += 1
                y = page_height - top_margin
                c.setFont(font_name, 10)
                c.drawString(left_margin, y, f'未盘设备列表 (共{len(pending_task_devices)}条) - 续')
                y -= 6 * mm
                y = draw_table_header(c, y, pending_headers, pending_col_x, pending_widths)
            
            pending_num += 1
            device = td.device
            
            values = [
                str(pending_num),
                device.asset_no or '',
                device.device_no or '',
                device.name or '',
                device.model or '',
                device.department.name if device.department else '',
                device.user.realname if device.user else '',
                device.location_text or '',
                status_map.get(device.status, device.status or ''),
                '是' if device.is_fixed else '否',
                device.asset_card_no or '',
                '未盘'
            ]
            
            y = draw_table_row(c, y, values, pending_col_x, pending_widths, is_odd=(pending_num % 2 == 0))
    
    # ========== 签字栏 ==========
    y -= 10 * mm
    if y < bottom_margin + 15*mm:
        draw_page_footer(c, page_num)
        c.showPage()
        page_num += 1
        y = page_height - top_margin
    
    c.setFont(font_name, 9)
    c.line(left_margin, y, page_width - right_margin, y)
    y -= 6 * mm
    
    c.drawString(left_margin, y, '盘点人签字: ________________')
    c.drawString(left_margin + 80*mm, y, '审核人签字: ________________')
    c.drawString(left_margin + 160*mm, y, '日期: ________________')
    
    # 最后一页页脚
    draw_page_footer(c, page_num)
    
    c.save()
    buffer.seek(0)
    
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="report_{task.task_no}.pdf"'
    return response


@login_required
def task_report_export(request, pk):
    """盘点报告导出Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.utils import timezone as tz
    
    task = get_object_or_404(InventoryTask.objects.select_related('created_by', 'assignee'), pk=pk)
    
    # 获取盘点记录（按资产编号排序）
    records = list(task.records.select_related('device', 'device__category', 'device__department', 
                                               'device__user', 'device__location', 'checked_by').order_by('device__asset_no'))
    
    # 统计数据
    total_count = task.device_count
    checked_count = task.checked_count
    pending_count = total_count - checked_count
    
    # 位置状态统计
    in_place_count = sum(1 for r in records if r.location_status == 'in_place')
    moved_count = sum(1 for r in records if r.location_status == 'moved')
    
    # 资产状态统计
    normal_count = sum(1 for r in records if r.asset_status == 'normal')
    damaged_count = sum(1 for r in records if r.asset_status == 'damaged')
    
    # 获取未盘设备（按资产编号排序）
    pending_task_devices = list(task.task_devices.filter(
        status='pending'
    ).select_related('device', 'device__category', 'device__department', 'device__user').order_by('device__asset_no'))
    
    # 盘点人列表（去重）
    checkers = list(set([r.checked_by.realname for r in records if r.checked_by]))
    checkers_text = '、'.join(checkers) if checkers else '-'
    
    # 完成率
    completion_rate = round(checked_count / total_count * 100, 1) if total_count > 0 else 0
    
    # 部门统计
    from collections import defaultdict
    dept_stats = defaultdict(int)
    for r in records:
        if r.device.department:
            dept_stats[r.device.department.name] += 1
    
    # 分类统计
    cat_stats = defaultdict(int)
    for r in records:
        if r.device.category:
            cat_stats[r.device.category.name] += 1
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 样式定义
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # ========== 工作表1：统计概览 ==========
    ws1 = wb.active
    ws1.title = '统计概览'
    
    # 任务信息
    ws1['A1'] = '盘点报告'
    ws1['A1'].font = Font(bold=True, size=14)
    ws1.merge_cells('A1:D1')
    
    ws1['A3'] = '任务编号'
    ws1['B3'] = task.task_no
    ws1['A4'] = '任务类型'
    ws1['B4'] = task.get_task_type_display()
    ws1['A5'] = '创建人'
    ws1['B5'] = task.created_by.realname if task.created_by else '-'
    ws1['A6'] = '盘点人'
    ws1['B6'] = checkers_text
    ws1['A7'] = '盘点范围'
    ws1['B7'] = '在账设备' if task.is_fixed_only else '所有设备'
    ws1['A8'] = '状态'
    ws1['B8'] = '已完成' if task.status == 'completed' else '执行中'
    ws1['A9'] = '创建时间'
    ws1['B9'] = task.created_at.strftime('%Y-%m-%d %H:%M')
    ws1['A10'] = '完成时间'
    ws1['B10'] = task.completed_at.strftime('%Y-%m-%d %H:%M') if task.completed_at else '-'
    
    for i in range(3, 11):
        ws1[f'A{i}'].font = Font(bold=True)
    
    # 统计概览
    ws1['D3'] = '统计概览'
    ws1['D3'].font = Font(bold=True, size=12)
    
    ws1['D5'] = '应盘设备'
    ws1['E5'] = total_count
    ws1['D6'] = '已盘设备'
    ws1['E6'] = checked_count
    ws1['D7'] = '未盘设备'
    ws1['E7'] = pending_count
    ws1['D8'] = '完成率'
    ws1['E8'] = f'{completion_rate}%'
    
    for i in range(5, 9):
        ws1[f'D{i}'].font = Font(bold=True)
    
    # 位置状态统计
    ws1['G3'] = '位置状态'
    ws1['G3'].font = Font(bold=True, size=12)
    ws1['G5'] = '在位'
    ws1['H5'] = in_place_count
    ws1['G6'] = '搬离'
    ws1['H6'] = moved_count
    ws1['G5'].font = Font(bold=True)
    ws1['G6'].font = Font(bold=True)
    
    # 资产状态统计
    ws1['J3'] = '资产状态'
    ws1['J3'].font = Font(bold=True, size=12)
    ws1['J5'] = '正常'
    ws1['K5'] = normal_count
    ws1['J6'] = '损坏'
    ws1['K6'] = damaged_count
    ws1['J5'].font = Font(bold=True)
    ws1['J6'].font = Font(bold=True)
    
    # 部门统计
    row = 13
    ws1[f'A{row}'] = '部门盘点统计'
    ws1[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    ws1[f'A{row}'] = '部门'
    ws1[f'B{row}'] = '已盘数量'
    ws1[f'A{row}'].font = header_font
    ws1[f'B{row}'].font = header_font
    row += 1
    for dept_name, count in sorted(dept_stats.items(), key=lambda x: -x[1]):
        ws1[f'A{row}'] = dept_name
        ws1[f'B{row}'] = count
        row += 1
    
    # 分类统计
    row += 1
    ws1[f'A{row}'] = '分类盘点统计'
    ws1[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    ws1[f'A{row}'] = '分类'
    ws1[f'B{row}'] = '已盘数量'
    ws1[f'A{row}'].font = header_font
    ws1[f'B{row}'].font = header_font
    row += 1
    for cat_name, count in sorted(cat_stats.items(), key=lambda x: -x[1]):
        ws1[f'A{row}'] = cat_name
        ws1[f'B{row}'] = count
        row += 1
    
    # 调整列宽
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['J'].width = 12
    ws1.column_dimensions['K'].width = 12
    
    # ========== 工作表2：已盘点设备明细 ==========
    ws2 = wb.create_sheet('已盘点设备明细')
    
    headers2 = ['资产编号', '设备编号', '设备名称', '型号', '所属部门', '使用人', '位置', '状态',
                '固资在账', '卡片编号', '位置状态', '资产状态', '盘点人', '盘点时间', '备注']
    ws2.append(headers2)
    
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    status_map = {'normal': '使用', 'fault': '故障', 'scrapped': '报废', 'unused': '闲置'}
    
    for record in records:
        row = [
            record.device.asset_no or '',
            record.device.device_no or '',
            record.device.name or '',
            record.device.model or '',
            record.device.department.name if record.device.department else '',
            record.device.user.realname if record.device.user else '',
            record.device.location_text or '',
            status_map.get(record.device.status, record.device.status or ''),
            '是' if record.device.is_fixed else '否',
            record.device.asset_card_no or '',
            '在位' if record.location_status == 'in_place' else '搬离',
            '正常' if record.asset_status == 'normal' else '损坏',
            record.checked_by.realname if record.checked_by else '',
            record.checked_at.strftime('%Y-%m-%d %H:%M') if record.checked_at else '',
            record.remarks or ''
        ]
        ws2.append(row)
    
    # 设置边框和对齐
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=len(headers2)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 自动调整列宽
    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value or ''))
                for char in str(cell.value or ''):
                    if '\u4e00' <= char <= '\u9fff':
                        cell_len += 1
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws2.column_dimensions[column].width = min(max(max_length + 2, 8), 20)
    
    # ========== 工作表3：未盘设备列表 ==========
    ws3 = wb.create_sheet('未盘设备列表')
    
    headers3 = ['资产编号', '设备编号', '设备名称', '型号', '所属部门', '使用人', '位置', '状态',
                '固资在账', '卡片编号']
    ws3.append(headers3)
    
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for td in pending_task_devices:
        device = td.device
        row = [
            device.asset_no or '',
            device.device_no or '',
            device.name or '',
            device.model or '',
            device.department.name if device.department else '',
            device.user.realname if device.user else '',
            device.location_text or '',
            status_map.get(device.status, device.status or ''),
            '是' if device.is_fixed else '否',
            device.asset_card_no or ''
        ]
        ws3.append(row)
    
    # 设置边框和对齐
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=len(headers3)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
    
    # 自动调整列宽
    for col in ws3.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                cell_len = len(str(cell.value or ''))
                for char in str(cell.value or ''):
                    if '\u4e00' <= char <= '\u9fff':
                        cell_len += 1
                if cell_len > max_length:
                    max_length = cell_len
            except:
                pass
        ws3.column_dimensions[column].width = min(max(max_length + 2, 8), 20)
    
    # 生成响应
    from django.utils import timezone as tz
    import random, string
    now = tz.localtime(tz.now())
    random_code = ''.join(random.choices(string.ascii_letters + string.digits, k=4))
    filename = f'report_{task.task_no}_{now.strftime("%Y%m%d")}_{random_code}.xlsx'
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ==================== 兼容旧接口 - 盘点审核 ====================

@login_required
def task_verify(request, pk):
    """盘点审核（兼容旧接口）"""
    task = get_object_or_404(InventoryTask, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'confirm':
            task.status = 'completed'
            task.completed_at = timezone.now()
            task.save()
            messages.success(request, '盘点任务审核通过')
        elif action == 'difference':
            record_id = request.POST.get('record_id')
            diff_action = request.POST.get('diff_action')
            diff_remarks = request.POST.get('diff_remarks', '')
            
            record = get_object_or_404(InventoryRecord, pk=record_id)
            
            if diff_action == 'profit':
                messages.success(request, '已确认为盘盈')
            elif diff_action == 'loss':
                messages.success(request, '已确认为盘亏')
            elif diff_action == 'damage':
                record.asset_status = 'damaged'
                record.remarks = diff_remarks
                record.save()
                messages.success(request, '已确认为损毁')
        
        return redirect('task_verify', pk)
    
    records = task.records.select_related('device', 'checked_by').all()
    
    in_place = records.filter(location_status='in_place').count()
    moved = records.filter(location_status='moved').count()
    not_found = records.filter(location_status='not_found').count()
    normal = records.filter(asset_status='normal').count()
    damaged = records.filter(asset_status='damaged').count()
    lost = records.filter(asset_status='lost').count()
    
    stats = {
        'total': task.device_count,
        'checked': task.checked_count,
        'in_place': in_place,
        'moved': moved,
        'not_found': not_found,
        'normal': normal,
        'damaged': damaged,
        'lost': lost,
    }
    
    return render(request, 'inventory/task_verify.html', {'task': task, 'records': records, 'stats': stats})
